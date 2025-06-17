# Complete Enhanced AWS Migration Analysis Platform v7.0 with vROPS Integration
# Requirements: streamlit>=1.28.0, pandas>=1.5.0, plotly>=5.0.0, reportlab>=3.6.0, anthropic>=0.8.0, openpyxl>=3.1.0, requests>=2.28.0, urllib3>=1.26.0

import streamlit as st
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
from plotly.subplots import make_subplots
import math
import boto3
import json
import logging
from datetime import datetime, timedelta
import io
from typing import Dict, List, Tuple, Optional, Any
import numpy as np
import anthropic
import requests
import csv
from io import StringIO
import urllib3
import base64
import ssl
from urllib.parse import quote
import warnings

# Disable SSL warnings for vROPS connections
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
warnings.filterwarnings('ignore', message='Unverified HTTPS request')

# Configure page - MUST be first Streamlit command
st.set_page_config(
    page_title="Enhanced AWS Migration Platform v7.0 with vROPS",
    layout="wide",
    page_icon="🏢",
    initial_sidebar_state="expanded"
)

# Try to import reportlab for PDF generation
try:
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.lib import colors
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_JUSTIFY
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False

# Try to import openpyxl for Excel generation
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

# Set up logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Enhanced Modern CSS with Frame Structure - REPLACE YOUR EXISTING CSS SECTION
st.markdown("""
<style>
    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700;800&display=swap');
    
    /* Global Styles */
    html, body, [class*="css"] {
        font-family: 'Inter', -apple-system, BlinkMacSystemFont, 'Segoe UI', sans-serif;
        background: linear-gradient(135deg, #f8fafc 0%, #e2e8f0 100%);
    }
    
    /* Hide Streamlit Default Elements */
    #MainMenu {visibility: hidden;}
    footer {visibility: hidden;}
    header {visibility: hidden;}
    
    /* Main Application Frame */
    .main-app-frame {
        background: white;
        border-radius: 20px;
        box-shadow: 0 20px 40px rgba(0,0,0,0.1);
        margin: 1rem;
        overflow: hidden;
        border: 1px solid #e2e8f0;
    }
    
    /* Enhanced Header Frame */
    .header-frame {
        background: linear-gradient(135deg, #1e3a8a 0%, #3730a3 50%, #1e40af 100%);
        color: white;
        padding: 2rem 3rem;
        position: relative;
        overflow: hidden;
    }
    
    .header-frame::before {
        content: '';
        position: absolute;
        top: 0;
        left: 0;
        right: 0;
        bottom: 0;
        background: url('data:image/svg+xml,<svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 1000 1000"><polygon fill="%23ffffff08" points="0,1000 1000,0 1000,1000"/></svg>');
        background-size: cover;
        pointer-events: none;
    }
    
    .header-content {
        position: relative;
        z-index: 2;
    }
    
    .header-title {
        font-size: 2.5rem;
        font-weight: 800;
        margin-bottom: 0.5rem;
        text-shadow: 0 2px 4px rgba(0,0,0,0.3);
    }
    
    .header-subtitle {
        font-size: 1.2rem;
        opacity: 0.9;
        margin-bottom: 1rem;
        font-weight: 400;
    }
    
    .header-version {
        display: inline-block;
        background: rgba(255,255,255,0.2);
        padding: 0.5rem 1rem;
        border-radius: 50px;
        font-size: 0.9rem;
        font-weight: 500;
        backdrop-filter: blur(10px);
        border: 1px solid rgba(255,255,255,0.3);
    }
    
    /* Content Frame */
    .content-frame {
        padding: 2rem 3rem;
        min-height: 60vh;
        background: white;
    }
    
    /* Navigation Frame */
    .nav-frame {
        background: #f8fafc;
        border-bottom: 1px solid #e2e8f0;
        padding: 0;
        margin: 0;
    }
    
    /* Sidebar Frame */
    .sidebar-frame {
        background: linear-gradient(180deg, #f8fafc 0%, #f1f5f9 100%);
        border-right: 1px solid #e2e8f0;
        padding: 1.5rem;
        height: 100vh;
        overflow-y: auto;
    }
    
    /* Enhanced Cards */
    .modern-card {
        background: white;
        border-radius: 16px;
        padding: 2rem;
        box-shadow: 0 4px 20px rgba(0,0,0,0.08);
        border: 1px solid #f1f5f9;
        margin-bottom: 1.5rem;
        transition: all 0.3s ease;
        position: relative;
        overflow: hidden;
    }
    
    .modern-card:hover {
        transform: translateY(-2px);
        box-shadow: 0 8px 30px rgba(0,0,0,0.12);
    }
    
    .modern-card::before {
        content: '';
        position: absolute;
        top: 0;
        left: 0;
        right: 0;
        height: 4px;
        background: linear-gradient(90deg, #3b82f6, #8b5cf6, #06b6d4);
    }
    
    /* vROPS Status Cards */
    .vrops-status-card {
        background: linear-gradient(135deg, #ffffff 0%, #f0f9ff 100%);
        border: 2px solid #3b82f6;
        border-radius: 16px;
        padding: 1.5rem;
        margin: 1rem 0;
        position: relative;
        overflow: hidden;
    }
    
    .vrops-connected {
        border-color: #10b981;
        background: linear-gradient(135deg, #ffffff 0%, #ecfdf5 100%);
    }
    
    .vrops-disconnected {
        border-color: #f59e0b;
        background: linear-gradient(135deg, #ffffff 0%, #fffbeb 100%);
    }
    
    .vrops-error {
        border-color: #ef4444;
        background: linear-gradient(135deg, #ffffff 0%, #fef2f2 100%);
    }
    
    /* Metric Cards */
    .metric-card-enhanced {
        background: linear-gradient(135deg, #ffffff 0%, #f8fafc 100%);
        border: 2px solid transparent;
        border-radius: 20px;
        padding: 2rem;
        text-align: center;
        position: relative;
        overflow: hidden;
        transition: all 0.3s ease;
    }
    
    .metric-card-enhanced::before {
        content: '';
        position: absolute;
        top: -2px;
        left: -2px;
        right: -2px;
        bottom: -2px;
        background: linear-gradient(45deg, #3b82f6, #8b5cf6, #06b6d4, #10b981);
        border-radius: 20px;
        z-index: -1;
        animation: gradient-border 3s ease infinite;
        background-size: 400% 400%;
    }
    
    @keyframes gradient-border {
        0%, 100% { background-position: 0% 50%; }
        50% { background-position: 100% 50%; }
    }
    
    .metric-card-enhanced:hover {
        transform: translateY(-5px);
        box-shadow: 0 15px 35px rgba(0,0,0,0.15);
    }
    
    .metric-value {
        font-size: 2.5rem;
        font-weight: 800;
        color: #1e293b;
        margin: 0.5rem 0;
        text-shadow: 0 2px 4px rgba(0,0,0,0.1);
    }
    
    .metric-label {
        font-size: 0.9rem;
        font-weight: 600;
        color: #64748b;
        text-transform: uppercase;
        letter-spacing: 0.5px;
        margin-bottom: 0.5rem;
    }
    
    .metric-description {
        font-size: 0.8rem;
        color: #94a3b8;
        font-weight: 500;
    }
    
    /* Status Indicators */
    .status-indicator {
        display: inline-flex;
        align-items: center;
        padding: 0.5rem 1rem;
        border-radius: 50px;
        font-weight: 600;
        font-size: 0.85rem;
        margin: 0.25rem;
        box-shadow: 0 2px 8px rgba(0,0,0,0.1);
    }
    
    .status-connected {
        background: linear-gradient(135deg, #10b981, #059669);
        color: white;
    }
    
    .status-warning {
        background: linear-gradient(135deg, #f59e0b, #d97706);
        color: white;
    }
    
    .status-error {
        background: linear-gradient(135deg, #ef4444, #dc2626);
        color: white;
    }
    
    /* Enhanced Tabs */
    .stTabs [data-baseweb="tab-list"] {
        gap: 8px;
        background: #f8fafc;
        padding: 0.5rem;
        border-radius: 16px;
        border: 1px solid #e2e8f0;
    }
    
    .stTabs [data-baseweb="tab"] {
        height: 60px;
        background: white;
        border-radius: 12px;
        color: #64748b;
        font-weight: 600;
        padding: 0 24px;
        border: 2px solid transparent;
        transition: all 0.3s ease;
        position: relative;
        overflow: hidden;
    }
    
    .stTabs [data-baseweb="tab"]::before {
        content: '';
        position: absolute;
        top: 0;
        left: 0;
        right: 0;
        bottom: 0;
        background: linear-gradient(135deg, #3b82f6, #8b5cf6);
        opacity: 0;
        transition: opacity 0.3s ease;
    }
    
    .stTabs [aria-selected="true"] {
        background: linear-gradient(135deg, #3b82f6, #8b5cf6) !important;
        color: white !important;
        transform: translateY(-2px);
        box-shadow: 0 8px 25px rgba(59, 130, 246, 0.4);
    }
    
    .stTabs [data-baseweb="tab"]:hover {
        transform: translateY(-2px);
        box-shadow: 0 4px 15px rgba(0,0,0,0.1);
        border-color: #e2e8f0;
    }
    
    /* Sub-tabs styling for Single Workload */
    .stTabs .stTabs [data-baseweb="tab"] {
        height: 50px;
        background-color: #f1f5f9;
        font-size: 14px;
        color: #475569;
        border: 1px solid #e2e8f0;
        margin-right: 4px;
    }
    
    .stTabs .stTabs [aria-selected="true"] {
        background-color: #0f766e !important;
        color: white !important;
        border: 1px solid #0d9488;
        box-shadow: 0 2px 8px rgba(15, 118, 110, 0.3);
    }
    
    .stTabs .stTabs [data-baseweb="tab"]:hover {
        background-color: #e2e8f0;
        color: #334155;
    }
    
    .stTabs .stTabs [aria-selected="true"]:hover {
        background-color: #0d9488 !important;
        color: white !important;
    }
    
    /* Form Elements */
    .stSelectbox > div > div {
        border-radius: 12px;
        border: 2px solid #e2e8f0;
        transition: all 0.3s ease;
    }
    
    .stSelectbox > div > div:focus-within {
        border-color: #3b82f6;
        box-shadow: 0 0 0 3px rgba(59, 130, 246, 0.1);
    }
    
    .stNumberInput > div > div {
        border-radius: 12px;
        border: 2px solid #e2e8f0;
        transition: all 0.3s ease;
    }
    
    .stNumberInput > div > div:focus-within {
        border-color: #3b82f6;
        box-shadow: 0 0 0 3px rgba(59, 130, 246, 0.1);
    }
    
    /* Buttons */
    .stButton > button {
        border-radius: 12px;
        font-weight: 600;
        padding: 0.75rem 2rem;
        border: none;
        transition: all 0.3s ease;
        position: relative;
        overflow: hidden;
    }
    
    .stButton > button[data-testid="baseButton-primary"] {
        background: linear-gradient(135deg, #3b82f6, #8b5cf6);
        color: white;
        box-shadow: 0 4px 15px rgba(59, 130, 246, 0.3);
    }
    
    .stButton > button[data-testid="baseButton-primary"]:hover {
        transform: translateY(-2px);
        box-shadow: 0 8px 25px rgba(59, 130, 246, 0.4);
    }
    
    /* Expandable Sections */
    .streamlit-expanderHeader {
        background: linear-gradient(135deg, #f8fafc, #f1f5f9);
        border-radius: 12px;
        border: 1px solid #e2e8f0;
        font-weight: 600;
        color: #374151;
    }
    
    /* Progress Indicators */
    .progress-ring {
        display: inline-block;
        width: 80px;
        height: 80px;
        margin: 1rem;
    }
    
    /* Alert Boxes */
    .stAlert {
        border-radius: 12px;
        border: none;
        box-shadow: 0 4px 15px rgba(0,0,0,0.08);
    }
    
    /* Footer Frame */
    .footer-frame {
        background: linear-gradient(135deg, #1e293b, #374151);
        color: #e2e8f0;
        padding: 2rem 3rem;
        text-align: center;
        border-top: 1px solid #374151;
    }
    
    /* Animation Classes */
    .fade-in {
        animation: fadeIn 0.6s ease-in;
    }
    
    @keyframes fadeIn {
        from { opacity: 0; transform: translateY(20px); }
        to { opacity: 1; transform: translateY(0); }
    }
    
    .slide-in-left {
        animation: slideInLeft 0.6s ease-out;
    }
    
    @keyframes slideInLeft {
        from { opacity: 0; transform: translateX(-30px); }
        to { opacity: 1; transform: translateX(0); }
    }
    
    /* Loading States */
    .loading-container {
        display: flex;
        justify-content: center;
        align-items: center;
        padding: 3rem;
    }
    
    .loading-spinner {
        width: 50px;
        height: 50px;
        border: 4px solid #f3f4f6;
        border-top: 4px solid #3b82f6;
        border-radius: 50%;
        animation: spin 1s linear infinite;
    }
    
    @keyframes spin {
        0% { transform: rotate(0deg); }
        100% { transform: rotate(360deg); }
    }
    
    /* Responsive Design */
    @media (max-width: 768px) {
        .header-frame {
            padding: 1.5rem 2rem;
        }
        
        .header-title {
            font-size: 2rem;
        }
        
        .content-frame {
            padding: 1.5rem 2rem;
        }
        
        .modern-card {
            padding: 1.5rem;
        }
    }
</style>
""", unsafe_allow_html=True)

class VROPSConnector:
    """VMware vRealize Operations connector for collecting on-premise metrics."""
    
    def __init__(self):
        self.session = None
        self.base_url = None
        self.token = None
        self.connected = False
        
    def connect(self, hostname: str, username: str, password: str, verify_ssl: bool = False) -> Dict[str, Any]:
        """Connect to vROPS instance."""
        try:
            self.base_url = f"https://{hostname}"
            
            # Create session
            self.session = requests.Session()
            if not verify_ssl:
                self.session.verify = False
                
            # Authenticate and get token
            auth_url = f"{self.base_url}/suite-api/api/auth/token/acquire"
            
            auth_data = {
                "username": username,
                "password": password
            }
            
            headers = {
                "Content-Type": "application/json",
                "Accept": "application/json"
            }
            
            response = self.session.post(auth_url, json=auth_data, headers=headers, timeout=30)
            
            if response.status_code == 200:
                self.token = response.json().get('token')
                self.session.headers.update({
                    'Authorization': f'vRealizeOpsToken {self.token}',
                    'Accept': 'application/json'
                })
                self.connected = True
                
                return {
                    'status': 'success',
                    'message': 'Successfully connected to vROPS',
                    'version': self._get_vrops_version()
                }
            else:
                return {
                    'status': 'error',
                    'message': f'Authentication failed: {response.status_code} - {response.text}'
                }
                
        except requests.exceptions.ConnectionError:
            return {
                'status': 'error',
                'message': 'Connection failed: Unable to reach vROPS instance'
            }
        except requests.exceptions.Timeout:
            return {
                'status': 'error',
                'message': 'Connection timeout: vROPS instance took too long to respond'
            }
        except Exception as e:
            return {
                'status': 'error',
                'message': f'Unexpected error: {str(e)}'
            }
    
    def _get_vrops_version(self) -> str:
        """Get vROPS version information."""
        try:
            if not self.connected:
                return "Unknown"
                
            version_url = f"{self.base_url}/suite-api/api/versions"
            response = self.session.get(version_url, timeout=10)
            
            if response.status_code == 200:
                versions = response.json().get('relatedLinks', [])
                if versions:
                    latest_version = versions[-1].get('name', 'Unknown')
                    return latest_version
            return "Connected"
        except:
            return "Connected"
    
    def get_virtual_machines(self, name_filter: str = None) -> Dict[str, Any]:
        """Get list of virtual machines from vROPS."""
        try:
            if not self.connected:
                return {'status': 'error', 'message': 'Not connected to vROPS'}
            
            # Get VMs using resources API
            resources_url = f"{self.base_url}/suite-api/api/resources"
            
            params = {
                'resourceKind': 'VirtualMachine',
                'pageSize': 1000
            }
            
            if name_filter:
                params['name'] = name_filter
            
            response = self.session.get(resources_url, params=params, timeout=30)
            
            if response.status_code == 200:
                data = response.json()
                vms = []
                
                for resource in data.get('resourceList', []):
                    vm_info = {
                        'id': resource.get('identifier'),
                        'name': resource.get('resourceKey', {}).get('name', 'Unknown'),
                        'resourceKind': resource.get('resourceKey', {}).get('resourceKindKey', 'Unknown'),
                        'adapterKind': resource.get('resourceKey', {}).get('adapterKindKey', 'Unknown'),
                        'resourceStatus': resource.get('resourceStatusStates', [{}])[0].get('resourceStatus', 'Unknown') if resource.get('resourceStatusStates') else 'Unknown'
                    }
                    vms.append(vm_info)
                
                return {
                    'status': 'success',
                    'vms': vms,
                    'total_count': len(vms)
                }
            else:
                return {
                    'status': 'error',
                    'message': f'Failed to get VMs: {response.status_code} - {response.text}'
                }
                
        except Exception as e:
            return {
                'status': 'error',
                'message': f'Error getting VMs: {str(e)}'
            }
    
    def get_vm_metrics(self, vm_id: str, days_back: int = 30) -> Dict[str, Any]:
        """Get detailed metrics for a specific VM."""
        try:
            if not self.connected:
                return {'status': 'error', 'message': 'Not connected to vROPS'}
            
            # Calculate time range
            end_time = datetime.now()
            start_time = end_time - timedelta(days=days_back)
            
            # Format timestamps for vROPS API (milliseconds since epoch)
            start_timestamp = int(start_time.timestamp() * 1000)
            end_timestamp = int(end_time.timestamp() * 1000)
            
            # Define metrics to collect
            metric_keys = [
                'cpu|usage_average',
                'cpu|usagemhz_average',
                'mem|usage_average',
                'mem|consumed_average',
                'disk|usage_average',
                'net|usage_average',
                'storage|totalReadLatency_average',
                'storage|totalWriteLatency_average',
                'cpu|ready_summation',
                'mem|swapinRate_average',
                'mem|swapoutRate_average'
            ]
            
            metrics_data = {}
            
            # Get metrics for each key
            for metric_key in metric_keys:
                stats_url = f"{self.base_url}/suite-api/api/resources/{vm_id}/stats"
                
                params = {
                    'statKey': metric_key,
                    'begin': start_timestamp,
                    'end': end_timestamp,
                    'rollUpType': 'AVG',
                    'intervalType': 'HOURS',
                    'intervalQuantifier': 1
                }
                
                try:
                    response = self.session.get(stats_url, params=params, timeout=30)
                    
                    if response.status_code == 200:
                        stat_data = response.json()
                        values = stat_data.get('values', [])
                        
                        if values:
                            # Extract numeric values
                            metric_values = []
                            for value_entry in values:
                                if len(value_entry) >= 2 and value_entry[1] is not None:
                                    try:
                                        metric_values.append(float(value_entry[1]))
                                    except (ValueError, TypeError):
                                        continue
                            
                            if metric_values:
                                metrics_data[metric_key] = {
                                    'values': metric_values,
                                    'average': sum(metric_values) / len(metric_values),
                                    'max': max(metric_values),
                                    'min': min(metric_values),
                                    'latest': metric_values[-1] if metric_values else 0,
                                    'samples': len(metric_values)
                                }
                            else:
                                metrics_data[metric_key] = self._get_empty_metric()
                        else:
                            metrics_data[metric_key] = self._get_empty_metric()
                    else:
                        logger.warning(f"Failed to get metric {metric_key}: {response.status_code}")
                        metrics_data[metric_key] = self._get_empty_metric()
                        
                except Exception as e:
                    logger.warning(f"Error getting metric {metric_key}: {e}")
                    metrics_data[metric_key] = self._get_empty_metric()
            
            # Calculate derived metrics
            processed_metrics = self._process_vm_metrics(metrics_data)
            
            return {
                'status': 'success',
                'vm_id': vm_id,
                'metrics_period_days': days_back,
                'metrics': processed_metrics,
                'raw_metrics': metrics_data
            }
            
        except Exception as e:
            return {
                'status': 'error',
                'message': f'Error getting VM metrics: {str(e)}'
            }
    
    def _get_empty_metric(self) -> Dict[str, Any]:
        """Return empty metric structure."""
        return {
            'values': [],
            'average': 0,
            'max': 0,
            'min': 0,
            'latest': 0,
            'samples': 0
        }
    
    def _process_vm_metrics(self, raw_metrics: Dict[str, Any]) -> Dict[str, Any]:
        """Process raw vROPS metrics into usable format."""
        try:
            processed = {
                'cpu': {
                    'usage_percent_avg': raw_metrics.get('cpu|usage_average', {}).get('average', 0),
                    'usage_percent_max': raw_metrics.get('cpu|usage_average', {}).get('max', 0),
                    'usage_mhz_avg': raw_metrics.get('cpu|usagemhz_average', {}).get('average', 0),
                    'usage_mhz_max': raw_metrics.get('cpu|usagemhz_average', {}).get('max', 0),
                    'ready_percent': raw_metrics.get('cpu|ready_summation', {}).get('average', 0)
                },
                'memory': {
                    'usage_percent_avg': raw_metrics.get('mem|usage_average', {}).get('average', 0),
                    'usage_percent_max': raw_metrics.get('mem|usage_average', {}).get('max', 0),
                    'consumed_mb_avg': raw_metrics.get('mem|consumed_average', {}).get('average', 0),
                    'consumed_mb_max': raw_metrics.get('mem|consumed_average', {}).get('max', 0),
                    'swap_in_rate': raw_metrics.get('mem|swapinRate_average', {}).get('average', 0),
                    'swap_out_rate': raw_metrics.get('mem|swapoutRate_average', {}).get('average', 0)
                },
                'disk': {
                    'usage_percent_avg': raw_metrics.get('disk|usage_average', {}).get('average', 0),
                    'usage_percent_max': raw_metrics.get('disk|usage_average', {}).get('max', 0),
                    'read_latency_ms': raw_metrics.get('storage|totalReadLatency_average', {}).get('average', 0),
                    'write_latency_ms': raw_metrics.get('storage|totalWriteLatency_average', {}).get('average', 0)
                },
                'network': {
                    'usage_kbps_avg': raw_metrics.get('net|usage_average', {}).get('average', 0),
                    'usage_kbps_max': raw_metrics.get('net|usage_average', {}).get('max', 0)
                }
            }
            
            # Calculate performance scores
            processed['performance_scores'] = {
                'cpu_score': self._calculate_cpu_score(processed['cpu']),
                'memory_score': self._calculate_memory_score(processed['memory']),
                'storage_score': self._calculate_storage_score(processed['disk']),
                'overall_score': 0
            }
            
            # Calculate overall score
            scores = processed['performance_scores']
            processed['performance_scores']['overall_score'] = (
                scores['cpu_score'] + scores['memory_score'] + scores['storage_score']
            ) / 3
            
            return processed
            
        except Exception as e:
            logger.error(f"Error processing VM metrics: {e}")
            return {}
    
    def _calculate_cpu_score(self, cpu_metrics: Dict[str, float]) -> float:
        """Calculate CPU performance score (0-100, higher is better)."""
        usage_avg = cpu_metrics.get('usage_percent_avg', 0)
        usage_max = cpu_metrics.get('usage_percent_max', 0)
        ready_percent = cpu_metrics.get('ready_percent', 0)
        
        # Penalty for high utilization
        usage_penalty = max(0, (usage_avg - 70) * 2)  # Penalty starts at 70%
        max_usage_penalty = max(0, (usage_max - 85) * 1.5)  # Penalty starts at 85%
        ready_penalty = ready_percent * 10  # Ready time is bad
        
        base_score = 100
        total_penalty = usage_penalty + max_usage_penalty + ready_penalty
        
        return max(0, base_score - total_penalty)
    
    def _calculate_memory_score(self, memory_metrics: Dict[str, float]) -> float:
        """Calculate memory performance score (0-100, higher is better)."""
        usage_avg = memory_metrics.get('usage_percent_avg', 0)
        usage_max = memory_metrics.get('usage_percent_max', 0)
        swap_in = memory_metrics.get('swap_in_rate', 0)
        swap_out = memory_metrics.get('swap_out_rate', 0)
        
        # Penalty for high utilization
        usage_penalty = max(0, (usage_avg - 80) * 2)  # Penalty starts at 80%
        max_usage_penalty = max(0, (usage_max - 90) * 1.5)  # Penalty starts at 90%
        swap_penalty = (swap_in + swap_out) * 5  # Swapping is bad
        
        base_score = 100
        total_penalty = usage_penalty + max_usage_penalty + swap_penalty
        
        return max(0, base_score - total_penalty)
    
    def _calculate_storage_score(self, disk_metrics: Dict[str, float]) -> float:
        """Calculate storage performance score (0-100, higher is better)."""
        read_latency = disk_metrics.get('read_latency_ms', 0)
        write_latency = disk_metrics.get('write_latency_ms', 0)
        
        # Penalty for high latency
        read_penalty = max(0, (read_latency - 20) * 2)  # Penalty starts at 20ms
        write_penalty = max(0, (write_latency - 20) * 2)  # Penalty starts at 20ms
        
        base_score = 100
        total_penalty = read_penalty + write_penalty
        
        return max(0, base_score - total_penalty)
    
    def get_vm_properties(self, vm_id: str) -> Dict[str, Any]:
        """Get VM configuration properties."""
        try:
            if not self.connected:
                return {'status': 'error', 'message': 'Not connected to vROPS'}
            
            properties_url = f"{self.base_url}/suite-api/api/resources/{vm_id}/properties"
            
            response = self.session.get(properties_url, timeout=30)
            
            if response.status_code == 200:
                properties_data = response.json()
                properties = {}
                
                for prop in properties_data.get('property', []):
                    prop_name = prop.get('name', '')
                    prop_value = prop.get('value', '')
                    properties[prop_name] = prop_value
                
                # Extract key properties
                vm_config = {
                    'cpu_cores': self._extract_numeric_property(properties, ['config|hardware|numCpu', 'cpu|cores', 'numCpu']),
                    'memory_mb': self._extract_numeric_property(properties, ['config|hardware|memoryMB', 'memory|configured', 'memoryMB']),
                    'disk_gb': self._extract_numeric_property(properties, ['config|hardware|diskGB', 'disk|provisioned', 'diskGB']) / 1024 if self._extract_numeric_property(properties, ['config|hardware|diskGB', 'disk|provisioned', 'diskGB']) else 0,
                    'power_state': properties.get('runtime|powerState', 'Unknown'),
                    'guest_os': properties.get('config|guestFullName', properties.get('guestOS', 'Unknown')),
                    'vm_tools_status': properties.get('runtime|toolsStatus', 'Unknown'),
                    'vm_version': properties.get('config|version', 'Unknown')
                }
                
                return {
                    'status': 'success',
                    'vm_id': vm_id,
                    'configuration': vm_config,
                    'all_properties': properties
                }
            else:
                return {
                    'status': 'error',
                    'message': f'Failed to get VM properties: {response.status_code}'
                }
                
        except Exception as e:
            return {
                'status': 'error',
                'message': f'Error getting VM properties: {str(e)}'
            }
    
    def _extract_numeric_property(self, properties: Dict[str, str], possible_keys: List[str]) -> float:
        """Extract numeric value from properties using multiple possible keys."""
        for key in possible_keys:
            if key in properties:
                try:
                    value = properties[key]
                    if isinstance(value, str):
                        # Remove units and convert to float
                        value = value.replace('MB', '').replace('GB', '').replace('GHz', '').replace('MHz', '').strip()
                    return float(value)
                except (ValueError, TypeError):
                    continue
        return 0
    
    def disconnect(self):
        """Disconnect from vROPS."""
        try:
            if self.connected and self.session and self.token:
                logout_url = f"{self.base_url}/suite-api/api/auth/token/release"
                self.session.post(logout_url, timeout=10)
        except:
            pass
        finally:
            self.connected = False
            self.session = None
            self.token = None
            self.base_url = None

class VROPSMetricsProcessor:
    """Process vROPS metrics for AWS migration analysis."""
    
    def __init__(self):
        self.sizing_recommendations = {}
        
    def process_vm_metrics_for_migration(self, vm_metrics: Dict[str, Any], vm_config: Dict[str, Any]) -> Dict[str, Any]:
        """Process vROPS metrics to generate AWS migration recommendations."""
        try:
            if vm_metrics.get('status') != 'success':
                return self._get_fallback_recommendations()
            
            metrics = vm_metrics.get('metrics', {})
            cpu_metrics = metrics.get('cpu', {})
            memory_metrics = metrics.get('memory', {})
            disk_metrics = metrics.get('disk', {})
            network_metrics = metrics.get('network', {})
            
            # Calculate sizing recommendations
            sizing = self._calculate_aws_sizing(cpu_metrics, memory_metrics, disk_metrics, vm_config)
            
            # Generate migration insights
            insights = self._generate_migration_insights(metrics, vm_config)
            
            # Calculate migration complexity based on performance patterns
            complexity = self._calculate_migration_complexity(metrics, vm_config)
            
            return {
                'status': 'success',
                'aws_sizing': sizing,
                'migration_insights': insights,
                'complexity_factors': complexity,
                'performance_summary': {
                    'cpu_efficiency': cpu_metrics.get('usage_percent_avg', 0),
                    'memory_efficiency': memory_metrics.get('usage_percent_avg', 0),
                    'storage_performance': disk_metrics.get('read_latency_ms', 0) + disk_metrics.get('write_latency_ms', 0),
                    'network_utilization': network_metrics.get('usage_kbps_avg', 0),
                    'overall_score': metrics.get('performance_scores', {}).get('overall_score', 50)
                }
            }
            
        except Exception as e:
            logger.error(f"Error processing VM metrics for migration: {e}")
            return self._get_fallback_recommendations()
    
    def _calculate_aws_sizing(self, cpu_metrics: Dict, memory_metrics: Dict, disk_metrics: Dict, vm_config: Dict) -> Dict[str, Any]:
        """Calculate AWS instance sizing based on vROPS metrics."""
        
        # Get current configuration
        current_cpu_cores = vm_config.get('cpu_cores', 2)
        current_memory_mb = vm_config.get('memory_mb', 4096)
        current_disk_gb = vm_config.get('disk_gb', 100)
        
        # Get utilization metrics
        cpu_avg = cpu_metrics.get('usage_percent_avg', 50)
        cpu_max = cpu_metrics.get('usage_percent_max', 70)
        memory_avg = memory_metrics.get('usage_percent_avg', 60)
        memory_max = memory_metrics.get('usage_percent_max', 80)
        
        # Calculate right-sized resources
        # CPU: Size for 70% average utilization with headroom for peaks
        cpu_utilization_factor = max(cpu_avg, cpu_max * 0.8) / 70  # Target 70% utilization
        recommended_vcpus = max(2, math.ceil(current_cpu_cores * cpu_utilization_factor))
        
        # Memory: Size for 80% average utilization with headroom
        memory_utilization_factor = max(memory_avg, memory_max * 0.9) / 80  # Target 80% utilization
        recommended_memory_gb = max(4, math.ceil((current_memory_mb / 1024) * memory_utilization_factor))
        
        # Storage: Add growth buffer
        recommended_storage_gb = max(100, math.ceil(current_disk_gb * 1.3))  # 30% growth buffer
        
        # Estimate IOPS based on performance metrics
        read_latency = disk_metrics.get('read_latency_ms', 10)
        write_latency = disk_metrics.get('write_latency_ms', 10)
        avg_latency = (read_latency + write_latency) / 2
        
        # Estimate IOPS based on latency (inverse relationship)
        if avg_latency < 5:
            estimated_iops = 8000  # High performance
        elif avg_latency < 15:
            estimated_iops = 4000  # Medium performance
        else:
            estimated_iops = 2000  # Standard performance
        
        return {
            'recommended_vcpus': recommended_vcpus,
            'recommended_memory_gb': recommended_memory_gb,
            'recommended_storage_gb': recommended_storage_gb,
            'estimated_iops': estimated_iops,
            'sizing_rationale': {
                'cpu_reasoning': f"Sized for {cpu_avg:.1f}% avg utilization, {cpu_max:.1f}% peak",
                'memory_reasoning': f"Sized for {memory_avg:.1f}% avg utilization, {memory_max:.1f}% peak",
                'storage_reasoning': f"Current {current_disk_gb}GB + 30% growth buffer",
                'performance_tier': self._determine_performance_tier(avg_latency)
            },
            'rightsizing_opportunities': {
                'cpu_oversized': current_cpu_cores > recommended_vcpus * 1.5,
                'cpu_undersized': current_cpu_cores < recommended_vcpus * 0.7,
                'memory_oversized': (current_memory_mb / 1024) > recommended_memory_gb * 1.5,
                'memory_undersized': (current_memory_mb / 1024) < recommended_memory_gb * 0.7
            }
        }
    
    def _determine_performance_tier(self, avg_latency: float) -> str:
        """Determine storage performance tier based on latency."""
        if avg_latency < 5:
            return "High Performance (io2)"
        elif avg_latency < 15:
            return "Balanced (gp3)"
        else:
            return "Standard (gp3)"
    
    def _generate_migration_insights(self, metrics: Dict, vm_config: Dict) -> Dict[str, Any]:
        """Generate migration-specific insights."""
        
        cpu_metrics = metrics.get('cpu', {})
        memory_metrics = metrics.get('memory', {})
        disk_metrics = metrics.get('disk', {})
        performance_scores = metrics.get('performance_scores', {})
        
        insights = {
            'migration_readiness': self._assess_migration_readiness(performance_scores),
            'optimization_opportunities': [],
            'risk_factors': [],
            'recommended_migration_approach': 'lift_and_shift'
        }
        
        # Optimization opportunities
        if cpu_metrics.get('usage_percent_avg', 0) < 30:
            insights['optimization_opportunities'].append("CPU over-provisioned - opportunity for cost savings")
        
        if memory_metrics.get('usage_percent_avg', 0) < 40:
            insights['optimization_opportunities'].append("Memory over-provisioned - opportunity for rightsizing")
        
        if memory_metrics.get('swap_in_rate', 0) > 0 or memory_metrics.get('swap_out_rate', 0) > 0:
            insights['optimization_opportunities'].append("Memory pressure detected - increase memory allocation")
        
        if cpu_metrics.get('ready_percent', 0) > 5:
            insights['risk_factors'].append("CPU ready time high - may indicate CPU contention")
        
        # Storage performance analysis
        avg_latency = (disk_metrics.get('read_latency_ms', 0) + disk_metrics.get('write_latency_ms', 0)) / 2
        if avg_latency > 20:
            insights['risk_factors'].append("High storage latency - consider SSD storage in AWS")
        
        # Migration approach recommendation
        overall_score = performance_scores.get('overall_score', 50)
        if overall_score > 80:
            insights['recommended_migration_approach'] = 'lift_and_shift'
        elif overall_score > 60:
            insights['recommended_migration_approach'] = 'lift_and_optimize'
        else:
            insights['recommended_migration_approach'] = 'refactor_recommended'
        
        return insights
    
    def _assess_migration_readiness(self, performance_scores: Dict) -> str:
        """Assess VM readiness for migration."""
        overall_score = performance_scores.get('overall_score', 50)
        
        if overall_score >= 80:
            return "High - VM performing well, ready for migration"
        elif overall_score >= 60:
            return "Medium - Some optimization needed before migration"
        elif overall_score >= 40:
            return "Low - Performance issues should be addressed"
        else:
            return "Critical - Significant performance problems require attention"
    
    def _calculate_migration_complexity(self, metrics: Dict, vm_config: Dict) -> Dict[str, Any]:
        """Calculate migration complexity factors based on performance data."""
        
        complexity_factors = {
            'performance_complexity': 0,
            'resource_complexity': 0,
            'storage_complexity': 0,
            'overall_complexity': 0
        }
        
        cpu_metrics = metrics.get('cpu', {})
        memory_metrics = metrics.get('memory', {})
        disk_metrics = metrics.get('disk', {})
        
        # Performance complexity
        cpu_ready = cpu_metrics.get('ready_percent', 0)
        memory_swap = memory_metrics.get('swap_in_rate', 0) + memory_metrics.get('swap_out_rate', 0)
        
        perf_complexity = 0
        if cpu_ready > 10:
            perf_complexity += 30
        elif cpu_ready > 5:
            perf_complexity += 15
        
        if memory_swap > 0:
            perf_complexity += 20
        
        complexity_factors['performance_complexity'] = min(perf_complexity, 100)
        
        # Resource complexity (based on sizing)
        cpu_cores = vm_config.get('cpu_cores', 2)
        memory_gb = vm_config.get('memory_mb', 4096) / 1024
        
        resource_complexity = 0
        if cpu_cores > 16:
            resource_complexity += 20
        if memory_gb > 64:
            resource_complexity += 20
        
        complexity_factors['resource_complexity'] = resource_complexity
        
        # Storage complexity
        avg_latency = (disk_metrics.get('read_latency_ms', 0) + disk_metrics.get('write_latency_ms', 0)) / 2
        storage_complexity = 0
        
        if avg_latency > 50:
            storage_complexity += 30
        elif avg_latency > 20:
            storage_complexity += 15
        
        complexity_factors['storage_complexity'] = storage_complexity
        
        # Overall complexity
        complexity_factors['overall_complexity'] = (
            complexity_factors['performance_complexity'] * 0.4 +
            complexity_factors['resource_complexity'] * 0.3 +
            complexity_factors['storage_complexity'] * 0.3
        )
        
        return complexity_factors
    
    def _get_fallback_recommendations(self) -> Dict[str, Any]:
        """Return fallback recommendations when vROPS data is not available."""
        return {
            'status': 'fallback',
            'aws_sizing': {
                'recommended_vcpus': 2,
                'recommended_memory_gb': 8,
                'recommended_storage_gb': 100,
                'estimated_iops': 3000,
                'sizing_rationale': {
                    'cpu_reasoning': "Default sizing - no vROPS data available",
                    'memory_reasoning': "Default sizing - no vROPS data available",
                    'storage_reasoning': "Default sizing - no vROPS data available",
                    'performance_tier': "Standard (gp3)"
                },
                'rightsizing_opportunities': {
                    'cpu_oversized': False,
                    'cpu_undersized': False,
                    'memory_oversized': False,
                    'memory_undersized': False
                }
            },
            'migration_insights': {
                'migration_readiness': "Unknown - no performance data available",
                'optimization_opportunities': ["Collect performance data for accurate sizing"],
                'risk_factors': ["No baseline performance data"],
                'recommended_migration_approach': 'lift_and_shift'
            },
            'complexity_factors': {
                'performance_complexity': 50,
                'resource_complexity': 50,
                'storage_complexity': 50,
                'overall_complexity': 50
            },
            'performance_summary': {
                'cpu_efficiency': 0,
                'memory_efficiency': 0,
                'storage_performance': 0,
                'network_utilization': 0,
                'overall_score': 50
            }
        }

class ClaudeAIMigrationAnalyzer:
    """Real Claude AI powered migration complexity analyzer using Anthropic API."""
    
    def __init__(self):
        self.complexity_factors = {
            'application_architecture': {'weight': 0.25},
            'technical_stack': {'weight': 0.20},
            'operational_complexity': {'weight': 0.20},
            'business_impact': {'weight': 0.20},
            'organizational_readiness': {'weight': 0.15}
        }

    def analyze_workload_complexity(self, workload_inputs: Dict, environment: str, vrops_data: Dict = None) -> Dict[str, Any]:
        """Analyze migration complexity using real Claude AI API with optional vROPS data."""
        
        try:
            # Get Claude API key from Streamlit secrets or environment
            api_key = self._get_claude_api_key()
            
            if not api_key:
                logger.warning("Claude API key not found, using fallback analysis")
                return self._get_fallback_analysis()
            
            # Initialize Claude client
            client = anthropic.Anthropic(api_key=api_key)
            
            # Prepare the prompt for Claude including vROPS data
            analysis_prompt = self._create_analysis_prompt(workload_inputs, environment, vrops_data)
            
            # Make API call to Claude
            response = client.messages.create(
                model="claude-3-5-sonnet-20241022",
                max_tokens=3000,
                temperature=0.1,
                system="You are an expert AWS migration architect with deep experience in VMware environments and vRealize Operations. Analyze the provided workload information and vROPS performance data to provide detailed migration recommendations in JSON format.",
                messages=[
                    {
                        "role": "user",
                        "content": analysis_prompt
                    }
                ]
            )
            
            # Parse Claude's response
            analysis_result = self._parse_claude_response(response.content[0].text)
            
            # Enhance analysis with vROPS insights if available
            if vrops_data and vrops_data.get('status') == 'success':
                analysis_result = self._enhance_with_vrops_insights(analysis_result, vrops_data)
            
            return analysis_result
            
        except Exception as e:
            logger.error(f"Error in Claude AI analysis: {e}")
            return self._get_fallback_analysis()

    def _get_claude_api_key(self) -> Optional[str]:
        """Get Claude API key from Streamlit secrets or environment variables."""
        try:
            # Try Streamlit secrets first
            if hasattr(st, 'secrets') and 'ANTHROPIC_API_KEY' in st.secrets:
                return str(st.secrets['ANTHROPIC_API_KEY'])
            
            # Try environment variable
            import os
            return os.getenv('ANTHROPIC_API_KEY')
            
        except Exception:
            return None

    def _create_analysis_prompt(self, workload_inputs: Dict, environment: str, vrops_data: Dict = None) -> str:
        """Create a detailed prompt for Claude to analyze the workload including vROPS data."""
        
        vrops_section = ""
        if vrops_data and vrops_data.get('status') == 'success':
            vrops_metrics = vrops_data.get('performance_summary', {})
            aws_sizing = vrops_data.get('aws_sizing', {})
            migration_insights = vrops_data.get('migration_insights', {})
            
            vrops_section = f"""
        
        **vROPS Performance Data (Real Production Metrics):**
        - CPU Efficiency: {vrops_metrics.get('cpu_efficiency', 0):.1f}%
        - Memory Efficiency: {vrops_metrics.get('memory_efficiency', 0):.1f}%
        - Storage Performance Score: {vrops_metrics.get('storage_performance', 0):.1f}ms
        - Network Utilization: {vrops_metrics.get('network_utilization', 0):.1f} Kbps
        - Overall Performance Score: {vrops_metrics.get('overall_score', 50):.1f}/100
        
        **vROPS-based AWS Sizing Recommendations:**
        - Recommended vCPUs: {aws_sizing.get('recommended_vcpus', 2)}
        - Recommended Memory: {aws_sizing.get('recommended_memory_gb', 8)} GB
        - Recommended Storage: {aws_sizing.get('recommended_storage_gb', 100)} GB
        - Estimated IOPS: {aws_sizing.get('estimated_iops', 3000)}
        - Performance Tier: {aws_sizing.get('sizing_rationale', {}).get('performance_tier', 'Standard')}
        
        **Migration Readiness Assessment:**
        - Readiness Level: {migration_insights.get('migration_readiness', 'Unknown')}
        - Recommended Approach: {migration_insights.get('recommended_migration_approach', 'lift_and_shift')}
        - Optimization Opportunities: {len(migration_insights.get('optimization_opportunities', []))} identified
        - Risk Factors: {len(migration_insights.get('risk_factors', []))} identified
        """
        
        prompt = f"""
        Please analyze the following workload for AWS migration and provide a comprehensive assessment in JSON format.

        **Workload Information:**
        - Name: {workload_inputs.get('workload_name', 'Unknown')}
        - Type: {workload_inputs.get('workload_type', 'web_application')}
        - Operating System: {workload_inputs.get('operating_system', 'linux')}
        - Environment: {environment}
        - Region: {workload_inputs.get('region', 'us-east-1')}

        **Current Infrastructure:**
        - CPU Cores: {workload_inputs.get('on_prem_cores', 8)}
        - Peak CPU Usage: {workload_inputs.get('peak_cpu_percent', 70)}%
        - RAM: {workload_inputs.get('on_prem_ram_gb', 32)} GB
        - Peak RAM Usage: {workload_inputs.get('peak_ram_percent', 80)}%
        - Storage: {workload_inputs.get('storage_current_gb', 500)} GB
        - Peak IOPS: {workload_inputs.get('peak_iops', 5000)}
        - Peak Throughput: {workload_inputs.get('peak_throughput_mbps', 250)} MB/s
        - Infrastructure Age: {workload_inputs.get('infrastructure_age_years', 3)} years
        - Business Criticality: {workload_inputs.get('business_criticality', 'medium')}
        {vrops_section}

        Please provide your analysis in the following JSON structure:
        
        {{
            "complexity_score": <number 0-100>,
            "complexity_level": "<LOW|MEDIUM|HIGH|CRITICAL>",
            "complexity_color": "<low|medium|high|critical>",
            "migration_strategy": {{
                "approach": "<migration approach>",
                "methodology": "<migration methodology>",
                "timeline": "<estimated timeline>",
                "risk_level": "<risk assessment>"
            }},
            "migration_steps": [
                {{
                    "phase": "<phase name>",
                    "duration": "<duration>",
                    "tasks": ["<task1>", "<task2>", "<task3>"],
                    "deliverables": ["<deliverable1>", "<deliverable2>"]
                }}
            ],
            "risk_factors": [
                {{
                    "category": "<risk category>",
                    "risk": "<risk description>",
                    "probability": "<Low|Medium|High>",
                    "impact": "<Low|Medium|High|Critical>",
                    "mitigation": "<mitigation strategy>"
                }}
            ],
            "estimated_timeline": {{
                "min_weeks": <number>,
                "max_weeks": <number>,
                "confidence": "<Low|Medium|High>"
            }},
            "recommendations": [
                "<recommendation1>",
                "<recommendation2>",
                "<recommendation3>"
            ],
            "success_factors": [
                "<factor1>",
                "<factor2>",
                "<factor3>"
            ],
            "vrops_insights": {{
                "performance_impact": "<impact assessment>",
                "sizing_confidence": "<confidence level>",
                "optimization_potential": "<optimization assessment>"
            }}
        }}

        Consider the environment type ({environment}) when assessing complexity, risk, and timeline. Production environments should have higher complexity scores and longer timelines due to stricter requirements.

        If vROPS data is available, use it to:
        1. Assess performance-based migration complexity
        2. Validate sizing recommendations
        3. Identify performance optimization opportunities
        4. Assess migration readiness based on current performance

        Base your complexity score on:
        - Technical complexity (30%)
        - Performance factors from vROPS (25%)
        - Operational requirements (20%)
        - Business impact (15%)
        - Migration risk (10%)

        Provide specific, actionable recommendations based on AWS best practices and vROPS performance insights.
        """
        
        return prompt

    def _enhance_with_vrops_insights(self, analysis_result: Dict[str, Any], vrops_data: Dict) -> Dict[str, Any]:
        """Enhance Claude analysis with vROPS-specific insights."""
        try:
            migration_insights = vrops_data.get('migration_insights', {})
            complexity_factors = vrops_data.get('complexity_factors', {})
            performance_summary = vrops_data.get('performance_summary', {})
            
            # Adjust complexity score based on performance data
            performance_complexity = complexity_factors.get('overall_complexity', 50)
            original_complexity = analysis_result.get('complexity_score', 50)
            
            # Weighted combination of Claude analysis and vROPS performance data
            enhanced_complexity = (original_complexity * 0.7) + (performance_complexity * 0.3)
            analysis_result['complexity_score'] = min(100, enhanced_complexity)
            
            # Add vROPS-specific recommendations
            vrops_recommendations = []
            
            for opportunity in migration_insights.get('optimization_opportunities', []):
                vrops_recommendations.append(f"Performance Optimization: {opportunity}")
            
            for risk in migration_insights.get('risk_factors', []):
                vrops_recommendations.append(f"Performance Risk: {risk}")
            
            # Merge recommendations
            original_recs = analysis_result.get('recommendations', [])
            analysis_result['recommendations'] = original_recs + vrops_recommendations[:3]  # Limit to top 3
            
            # Add vROPS insights section
            analysis_result['vrops_insights'] = {
                'performance_impact': f"Overall performance score: {performance_summary.get('overall_score', 50):.1f}/100",
                'sizing_confidence': "High - based on real performance data" if performance_summary.get('overall_score', 0) > 70 else "Medium - performance optimization recommended",
                'optimization_potential': migration_insights.get('migration_readiness', 'Unknown'),
                'recommended_approach': migration_insights.get('recommended_migration_approach', 'lift_and_shift')
            }
            
            return analysis_result
            
        except Exception as e:
            logger.error(f"Error enhancing analysis with vROPS insights: {e}")
            return analysis_result

    def _parse_claude_response(self, response_text: str) -> Dict[str, Any]:
        """Parse Claude's JSON response and validate it."""
        try:
            # Try to extract JSON from the response
            import re
            
            # Look for JSON block in the response
            json_match = re.search(r'\{.*\}', response_text, re.DOTALL)
            
            if json_match:
                json_str = json_match.group(0)
                parsed_response = json.loads(json_str)
                
                # Validate required fields
                required_fields = ['complexity_score', 'complexity_level', 'migration_strategy']
                
                for field in required_fields:
                    if field not in parsed_response:
                        logger.warning(f"Missing required field: {field}")
                        return self._get_fallback_analysis()
                
                # Ensure complexity_color is set
                if 'complexity_color' not in parsed_response:
                    level = parsed_response.get('complexity_level', 'MEDIUM').lower()
                    parsed_response['complexity_color'] = level.lower()
                
                return parsed_response
            
            else:
                logger.warning("No JSON found in Claude response")
                return self._get_fallback_analysis()
                
        except json.JSONDecodeError as e:
            logger.error(f"Failed to parse Claude response as JSON: {e}")
            return self._get_fallback_analysis()
        except Exception as e:
            logger.error(f"Error parsing Claude response: {e}")
            return self._get_fallback_analysis()

    def _get_fallback_analysis(self) -> Dict[str, Any]:
        """Fallback analysis when Claude API is not available."""
        return {
            'complexity_score': 50,
            'complexity_level': 'MEDIUM',
            'complexity_color': 'medium',
            'migration_strategy': {
                'approach': 'Standard Migration with Optimization',
                'methodology': 'Lift and shift with selective modernization',
                'timeline': '6-10 weeks',
                'risk_level': 'Medium'
            },
            'migration_steps': [
                {
                    'phase': 'Discovery & Assessment',
                    'duration': '1-2 weeks',
                    'tasks': [
                        'Infrastructure inventory and dependency mapping',
                        'Performance baseline establishment',
                        'Security and compliance assessment'
                    ],
                    'deliverables': ['Migration plan', 'Risk assessment', 'Architecture design']
                },
                {
                    'phase': 'Migration Execution',
                    'duration': '3-6 weeks',
                    'tasks': [
                        'Environment setup and configuration',
                        'Application migration and testing',
                        'Performance optimization'
                    ],
                    'deliverables': ['Migrated workload', 'Test results', 'Documentation']
                },
                {
                    'phase': 'Validation & Go-Live',
                    'duration': '1-2 weeks',
                    'tasks': [
                        'End-to-end testing',
                        'Production cutover',
                        'Post-migration optimization'
                    ],
                    'deliverables': ['Production workload', 'Support procedures']
                }
            ],
            'risk_factors': [
                {
                    'category': 'Technical',
                    'risk': 'Application compatibility issues',
                    'probability': 'Medium',
                    'impact': 'High',
                    'mitigation': 'Comprehensive testing in staging environment'
                },
                {
                    'category': 'Operational',
                    'risk': 'Extended downtime during cutover',
                    'probability': 'Low',
                    'impact': 'High',
                    'mitigation': 'Blue-green deployment strategy'
                }
            ],
            'estimated_timeline': {'min_weeks': 6, 'max_weeks': 10, 'confidence': 'Medium'},
            'recommendations': [
                'Implement comprehensive backup and rollback procedures',
                'Conduct thorough security and compliance validation',
                'Plan for adequate testing and validation phases',
                'Establish clear success criteria and KPIs'
            ],
            'success_factors': [
                'Strong project leadership and governance',
                'Clear communication with all stakeholders',
                'Adequate resource allocation and timeline',
                'Comprehensive testing strategy'
            ],
            'vrops_insights': {
                'performance_impact': 'No vROPS data available',
                'sizing_confidence': 'Low - using estimates',
                'optimization_potential': 'Unknown without performance data'
            }
        }

# REPLACE the existing AWSCostCalculator class (around line 1000-1500) with this improved version:

class AWSCostCalculator:
    """Enhanced AWS service cost calculator with real API integration and better error handling."""
    
def _get_ec2_pricing_with_os(self, instance_type: str, operating_system: str = 'linux') -> dict:
    """Get EC2 pricing with OS-specific adjustments."""
    
    # Get base Linux pricing
    base_pricing = self._get_ec2_pricing(instance_type)
    
    # Windows licensing adds approximately 20-40% to the cost
    if operating_system.lower() == 'windows':
        windows_multiplier = 1.3  # 30% increase for Windows licensing
        
        for pricing_model in base_pricing:
            if pricing_model not in ['source', 'last_updated']:
                base_pricing[pricing_model] = base_pricing[pricing_model] * windows_multiplier
        
        # Update metadata
        base_pricing['source'] = base_pricing.get('source', 'fallback') + '_windows'
    
    return base_pricing

# Update your _calculate_compute_costs method to use OS-specific pricing
def _calculate_compute_costs(self, env: str, compute_recs: Dict, requirements: Dict, operating_system: str = 'linux') -> Dict[str, Any]:
    """Calculate compute-related costs with OS-specific pricing."""
    
    instance_type = compute_recs['primary_instance']['type']
    instance_count = self._get_instance_count(env)
    
    # EC2 instance costs with OS-specific pricing
    instance_pricing = self._get_ec2_pricing_with_os(instance_type, operating_system)
    
    # Determine pricing model
    pricing_model = 'on_demand'
    if env in ['PROD', 'PREPROD']:
        pricing_model = 'ri_1y'
    
    monthly_instance_cost = instance_pricing[pricing_model] * instance_count * 730  # hours per month
    
    
    
    def __init__(self, region='us-east-1'):
        self.region = region
        self.pricing_client = None
        self.aws_connected = False
        
        # Try to initialize AWS client with multiple credential sources
        try:
            # Option 1: Try Streamlit secrets first
            if hasattr(st, 'secrets') and 'aws' in st.secrets:
                self.pricing_client = boto3.client(
                    'pricing',
                    region_name='us-east-1',  # Pricing API only available in us-east-1
                    aws_access_key_id=st.secrets['aws']['access_key_id'],
                    aws_secret_access_key=st.secrets['aws']['secret_access_key']
                )
                logger.info("Using AWS credentials from Streamlit secrets")
            else:
                # Option 2: Use default credential chain (AWS CLI, environment variables, IAM roles)
                self.pricing_client = boto3.client('pricing', region_name='us-east-1')
                logger.info("Using AWS credentials from default credential chain")
            
            # Test the connection with a minimal API call
            test_response = self.pricing_client.get_products(ServiceCode='AmazonEC2', MaxResults=1)
            self.aws_connected = True
            logger.info("✅ AWS Pricing API connected successfully")
            
        except Exception as e:
            logger.warning(f"⚠️ AWS Pricing API connection failed: {e}")
            logger.warning("🔄 Will use fallback pricing data")
            self.aws_connected = False
        
        # Enhanced pricing data with more instance types
        self.pricing = {
            'compute': {
                'ec2_instances': {},
                'elastic_ip': 0.005  # per hour
            },
            'network': {
                'alb': 0.0225,  # per ALB per hour
                'nlb': 0.0225,  # per NLB per hour
                'nat_gateway': 0.045,  # per NAT per hour
                'cloudfront': {'per_gb': 0.085},
                'route53': {'hosted_zone': 0.50},
                'data_transfer': {'out_to_internet_up_to_10tb': 0.09},
                'vpn_gateway': 0.05  # per hour
            },
            'storage': {
                'ebs': {
                    'gp3': 0.08,  # per GB-month
                    'io2': 0.125,  # per GB-month
                    'io2_iops': 0.065,  # per provisioned IOPS
                    'snapshots': 0.05  # per GB-month
                },
                's3': {
                    'standard': 0.023,  # per GB-month
                    'standard_ia': 0.0125,  # per GB-month
                    'glacier': 0.004  # per GB-month
                }
            },
            'database': {
                'aurora_mysql': {
                    'db.r6g.large': 0.274,  # per hour
                    'db.r6g.xlarge': 0.548
                },
                'rds_mysql': {
                    'db.r6g.large': 0.252,  # per hour
                    'db.r6g.xlarge': 0.504
                },
                'storage': {
                    'aurora_storage': 0.10,  # per GB-month
                    'aurora_io': 0.20,  # per million requests
                    'rds_gp2': 0.115  # per GB-month
                },
                'backup_storage': 0.095,  # per GB-month
                'rds_proxy': 0.015  # per vCPU-hour
            },
            'security': {
                'secrets_manager': 0.40,  # per secret per month
                'kms': 1.00,  # per key per month
                'config': 0.003,  # per configuration item
                'cloudtrail': 0.10,  # per 100,000 events
                'guardduty': 0.10,  # per GB analyzed
                'security_hub': 0.0010,  # per security check
                'waf': 0.60  # per web ACL per month
            },
            'monitoring': {
                'cloudwatch': {
                    'metrics': 0.30,  # per custom metric
                    'dashboards': 3.00,  # per dashboard
                    'alarms': 0.10,  # per alarm
                    'logs_ingestion': 0.50,  # per GB
                    'logs_storage': 0.03  # per GB-month
                },
                'xray': 0.000005,  # per trace
                'synthetics': 0.0012  # per canary run
            }
        }
        return{
        'ec2_instances': {
            'cost': monthly_instance_cost,
            'details': f"{instance_count}x {instance_type} ({pricing_model}) - {operating_system.title()}",
            'breakdown': {
                'instance_type': instance_type,
                'instance_count': instance_count,
                'pricing_model': pricing_model,
                'operating_system': operating_system,
                'unit_cost': instance_pricing[pricing_model],
                'monthly_hours': 730,
                'os_multiplier': 1.3 if operating_system.lower() == 'windows' else 1.0
            }
        },
    }
        
    def get_real_ec2_pricing(self, instance_type: str) -> Optional[Dict[str, float]]:
        """Get real EC2 pricing from AWS API."""
        if not self.aws_connected:
            return None
            
        try:
            filters = [
                {'Type': 'TERM_MATCH', 'Field': 'instanceType', 'Value': instance_type},
                {'Type': 'TERM_MATCH', 'Field': 'location', 'Value': self._get_region_name(self.region)},
                {'Type': 'TERM_MATCH', 'Field': 'operatingSystem', 'Value': 'Linux'},
                {'Type': 'TERM_MATCH', 'Field': 'preInstalledSw', 'Value': 'NA'},
                {'Type': 'TERM_MATCH', 'Field': 'tenancy', 'Value': 'Shared'},
            ]
            
            response = self.pricing_client.get_products(
                ServiceCode='AmazonEC2',
                Filters=filters,
                MaxResults=10
            )
            
            if response['PriceList']:
                # Parse pricing data
                for price_item in response['PriceList']:
                    data = json.loads(price_item)
                    terms = data.get('terms', {})
                    
                    # Get On-Demand pricing
                    on_demand = terms.get('OnDemand', {})
                    for term_data in on_demand.values():
                        price_dimensions = term_data.get('priceDimensions', {})
                        for dim_data in price_dimensions.values():
                            usd_price = dim_data.get('pricePerUnit', {}).get('USD')
                            if usd_price:
                                return {
                                    'on_demand': float(usd_price),
                                    'source': 'aws_api',
                                    'last_updated': datetime.now().isoformat()
                                }
            
            return None
            
        except Exception as e:
            logger.error(f"Error getting real pricing for {instance_type}: {e}")
            return None

    def _get_ec2_pricing(self, instance_type: str) -> dict:
        """Get EC2 pricing with real AWS API and enhanced fallback."""
        
        # Try AWS API first
        real_pricing = self.get_real_ec2_pricing(instance_type)
        
        if real_pricing:
            # Estimate other pricing models based on typical AWS discounts
            on_demand_price = real_pricing['on_demand']
            return {
                'on_demand': on_demand_price,
                'ri_1y': on_demand_price * 0.65,  # ~35% discount
                'ri_3y': on_demand_price * 0.55,  # ~45% discount
                'spot': on_demand_price * 0.30,   # ~70% discount
                'source': 'aws_api',
                'last_updated': real_pricing['last_updated']
            }
        
        # Enhanced fallback pricing with more instance types
        fallback_prices = {
            # General Purpose - M6i instances
            'm6i.large': {'on_demand': 0.0864, 'ri_1y': 0.0605, 'ri_3y': 0.0432, 'spot': 0.0259},
            'm6i.xlarge': {'on_demand': 0.1728, 'ri_1y': 0.1210, 'ri_3y': 0.0864, 'spot': 0.0518},
            'm6i.2xlarge': {'on_demand': 0.3456, 'ri_1y': 0.2419, 'ri_3y': 0.1728, 'spot': 0.1037},
            'm6i.4xlarge': {'on_demand': 0.6912, 'ri_1y': 0.4838, 'ri_3y': 0.3456, 'spot': 0.2074},
            'm6i.8xlarge': {'on_demand': 1.3824, 'ri_1y': 0.9677, 'ri_3y': 0.6912, 'spot': 0.4147},
            
            # Memory Optimized - R6i instances  
            'r6i.large': {'on_demand': 0.1008, 'ri_1y': 0.0706, 'ri_3y': 0.0504, 'spot': 0.0302},
            'r6i.xlarge': {'on_demand': 0.2016, 'ri_1y': 0.1411, 'ri_3y': 0.1008, 'spot': 0.0605},
            'r6i.2xlarge': {'on_demand': 0.4032, 'ri_1y': 0.2822, 'ri_3y': 0.2016, 'spot': 0.1210},
            'r6i.4xlarge': {'on_demand': 0.8064, 'ri_1y': 0.5645, 'ri_3y': 0.4032, 'spot': 0.2419},
            
            # Compute Optimized - C6i instances
            'c6i.large': {'on_demand': 0.0765, 'ri_1y': 0.0536, 'ri_3y': 0.0383, 'spot': 0.0230},
            'c6i.xlarge': {'on_demand': 0.1530, 'ri_1y': 0.1071, 'ri_3y': 0.0765, 'spot': 0.0459},
            'c6i.2xlarge': {'on_demand': 0.3060, 'ri_1y': 0.2142, 'ri_3y': 0.1530, 'spot': 0.0918},
            'c6i.4xlarge': {'on_demand': 0.6120, 'ri_1y': 0.4284, 'ri_3y': 0.3060, 'spot': 0.1836},
            
            # Burstable - T3 instances
            't3.micro': {'on_demand': 0.0104, 'ri_1y': 0.0062, 'ri_3y': 0.0041, 'spot': 0.0031},
            't3.small': {'on_demand': 0.0208, 'ri_1y': 0.0125, 'ri_3y': 0.0083, 'spot': 0.0062},
            't3.medium': {'on_demand': 0.0416, 'ri_1y': 0.0250, 'ri_3y': 0.0166, 'spot': 0.0125},
            't3.large': {'on_demand': 0.0832, 'ri_1y': 0.0499, 'ri_3y': 0.0333, 'spot': 0.0250},
        }
        
        pricing = fallback_prices.get(instance_type, {
            'on_demand': 0.1, 'ri_1y': 0.07, 'ri_3y': 0.05, 'spot': 0.03
        })
        
        # Add metadata
        pricing.update({
            'source': 'fallback',
            'last_updated': datetime.now().isoformat()
        })
        
        return pricing

    # Keep all the existing methods but update references to use the new pricing structure
    # ... (rest of the existing methods remain the same)

    def _get_region_name(self, region_code: str) -> str:
        """Map AWS region code to full name."""
        region_map = {
            'us-east-1': 'US East (N. Virginia)',
            'us-east-2': 'US East (Ohio)',
            'us-west-1': 'US West (N. California)',
            'us-west-2': 'US West (Oregon)',
            'eu-west-1': 'EU (Ireland)',
            'eu-west-2': 'EU (London)',
            'ap-southeast-1': 'Asia Pacific (Singapore)',
            'ap-southeast-2': 'Asia Pacific (Sydney)',
        }
        return region_map.get(region_code, region_code)
    
    def calculate_service_costs(self, env: str, tech_recs: Dict, requirements: Dict) -> Dict[str, Any]:
        """Calculate detailed costs for all AWS services."""
        
        costs = {
            'compute': self._calculate_compute_costs(env, tech_recs['compute'], requirements),
            'network': self._calculate_network_costs(env, tech_recs['network'], requirements),
            'storage': self._calculate_storage_costs(env, tech_recs['storage'], requirements),
            'database': self._calculate_database_costs(env, tech_recs['database'], requirements),
            'security': self._calculate_security_costs(env, tech_recs['security'], requirements),
            'monitoring': self._calculate_monitoring_costs(env, tech_recs['monitoring'], requirements)
        }
        
        # Calculate totals
        total_monthly = sum(category['total'] for category in costs.values())
        total_annual = total_monthly * 12
        
        costs['summary'] = {
            'total_monthly': total_monthly,
            'total_annual': total_annual,
            'largest_cost_category': max(costs.keys(), key=lambda k: costs[k]['total'] if k != 'summary' else 0)
        }
        
        return costs
    
    def _calculate_compute_costs(self, env: str, compute_recs: Dict, requirements: Dict) -> Dict[str, Any]:
        """Calculate compute-related costs."""
        
        instance_type = compute_recs['primary_instance']['type']
        instance_count = self._get_instance_count(env)
        
        # EC2 instance costs
        instance_pricing = self.pricing['compute']['ec2_instances'].get(instance_type, {
        'on_demand': 0.1, 'ri_1y': 0.07, 'ri_3y': 0.05, 'spot': 0.03
    })
        
        # Determine pricing model
        pricing_model = 'on_demand'
        if env in ['PROD', 'PREPROD']:
            pricing_model = 'ri_1y'
        
        monthly_instance_cost = instance_pricing[pricing_model] * instance_count
        
        # Auto Scaling (no additional cost, but affects instance count)
        scaling_info = compute_recs.get('scaling', {})
        max_instances = scaling_info.get('max_instances', instance_count)
        
        # Elastic IP costs (if needed)
        eip_cost = self.pricing['compute']['elastic_ip'] * instance_count if env in ['PROD', 'PREPROD'] else 0
        
        total_compute = monthly_instance_cost + eip_cost
        
        return {
            'ec2_instances': {
                'cost': monthly_instance_cost,
                'details': f"{instance_count}x {instance_type} ({pricing_model})",
                'breakdown': {
                    'instance_type': instance_type,
                    'instance_count': instance_count,
                    'pricing_model': pricing_model,
                    'unit_cost': instance_pricing[pricing_model],
                    'max_instances_scaling': max_instances
                }
            },
            'elastic_ip': {
                'cost': eip_cost,
                'details': f"{instance_count} Elastic IPs" if eip_cost > 0 else "No Elastic IPs"
            },
            'auto_scaling': {
                'cost': 0,
                'details': "No additional cost - scales existing instances"
            },
            'total': total_compute,
            'optimization_notes': self._get_compute_optimization_notes(env, instance_type, pricing_model)
        }
    
    def _calculate_network_costs(self, env: str, network_recs: Dict, requirements: Dict) -> Dict[str, Any]:
        """Calculate network-related costs."""
        
        # Load Balancer costs
        lb_cost = 0
        lb_type = network_recs.get('load_balancer', '')
        if 'ALB' in lb_type:
            lb_cost = self.pricing['network']['alb']
        elif 'NLB' in lb_type:
            lb_cost = self.pricing['network']['nlb']
        
        # NAT Gateway costs
        nat_cost = 0
        if network_recs.get('nat_gateway') == 'Required':
            nat_count = 2 if env in ['PROD', 'PREPROD'] else 1
            nat_cost = self.pricing['network']['nat_gateway'] * nat_count
        
        # CloudFront costs (estimated)
        cdn_cost = 0
        if network_recs.get('cdn') == 'CloudFront':
            estimated_gb_month = self._estimate_cdn_usage(env)
            cdn_cost = estimated_gb_month * self.pricing['network']['cloudfront']['per_gb']
        
        # Route 53 costs
        dns_cost = self.pricing['network']['route53']['hosted_zone']
        if 'health checks' in network_recs.get('dns', ''):
            dns_cost += 0.50 * 2
        
        # Data transfer costs (estimated)
        data_transfer_cost = self._estimate_data_transfer_costs(env)
        
        # VPN costs
        vpn_cost = 0
        if 'VPN' in network_recs.get('vpn', ''):
            vpn_cost = self.pricing['network']['vpn_gateway']
        
        total_network = lb_cost + nat_cost + cdn_cost + dns_cost + data_transfer_cost + vpn_cost
        
        return {
            'load_balancer': {
                'cost': lb_cost,
                'details': f"{lb_type}" if lb_cost > 0 else "No load balancer"
            },
            'nat_gateway': {
                'cost': nat_cost,
                'details': f"{int(nat_cost / self.pricing['network']['nat_gateway'])} NAT Gateways" if nat_cost > 0 else "No NAT Gateway"
            },
            'cloudfront_cdn': {
                'cost': cdn_cost,
                'details': f"~{self._estimate_cdn_usage(env)} GB/month" if cdn_cost > 0 else "No CDN"
            },
            'route53_dns': {
                'cost': dns_cost,
                'details': "Hosted zone + health checks" if dns_cost > 1 else "Basic hosted zone"
            },
            'data_transfer': {
                'cost': data_transfer_cost,
                'details': f"Estimated data transfer costs"
            },
            'vpn_gateway': {
                'cost': vpn_cost,
                'details': "Site-to-site VPN" if vpn_cost > 0 else "No VPN"
            },
            'total': total_network,
            'optimization_notes': self._get_network_optimization_notes(env)
        }
    
    def _calculate_storage_costs(self, env: str, storage_recs: Dict, requirements: Dict) -> Dict[str, Any]:
        """Calculate storage-related costs."""
        
        # Primary EBS storage
        storage_gb = requirements.get('storage_GB', 100)
        
        # EBS volume costs
        primary_storage_type = storage_recs.get('primary_storage', 'gp3')
        if 'gp3' in primary_storage_type:
            ebs_cost = storage_gb * self.pricing['storage']['ebs']['gp3']
        elif 'io2' in primary_storage_type:
            ebs_cost = storage_gb * self.pricing['storage']['ebs']['io2']
            # Add IOPS costs for io2
            iops = self._extract_iops_from_recommendation(storage_recs.get('iops_recommendation', ''))
            if iops > 3000:
                additional_iops = iops - 3000
                ebs_cost += additional_iops * self.pricing['storage']['ebs']['io2_iops']
        else:
            ebs_cost = storage_gb * self.pricing['storage']['ebs']['gp3']
        
        # Snapshot costs (for backup)
        snapshot_frequency = self._get_snapshot_frequency(storage_recs.get('backup_strategy', ''))
        snapshot_retention_gb = storage_gb * snapshot_frequency * 0.3
        snapshot_cost = snapshot_retention_gb * self.pricing['storage']['ebs']['snapshots']
        
        # S3 costs for long-term backup/archival
        s3_cost = 0
        if env in ['PROD', 'PREPROD']:
            s3_storage_gb = storage_gb * 2
            s3_cost = s3_storage_gb * self.pricing['storage']['s3']['standard_ia']
        
        total_storage = ebs_cost + snapshot_cost + s3_cost
        
        return {
            'ebs_primary': {
                'cost': ebs_cost,
                'details': f"{storage_gb} GB {primary_storage_type}",
                'breakdown': {
                    'volume_cost': storage_gb * (self.pricing['storage']['ebs']['io2'] if 'io2' in primary_storage_type else self.pricing['storage']['ebs']['gp3']),
                    'iops_cost': ebs_cost - (storage_gb * (self.pricing['storage']['ebs']['io2'] if 'io2' in primary_storage_type else self.pricing['storage']['ebs']['gp3']))
                }
            },
            'ebs_snapshots': {
                'cost': snapshot_cost,
                'details': f"~{snapshot_retention_gb:.0f} GB snapshot storage"
            },
            's3_backup': {
                'cost': s3_cost,
                'details': f"~{s3_storage_gb if s3_cost > 0 else 0} GB S3 IA storage" if s3_cost > 0 else "No S3 backup"
            },
            'total': total_storage,
            'optimization_notes': self._get_storage_optimization_notes(env, storage_recs)
        }
    
    def _calculate_database_costs(self, env: str, db_recs: Dict, requirements: Dict) -> Dict[str, Any]:
        """Calculate database-related costs."""
        
        # RDS instance costs
        instance_class = db_recs.get('instance_class', 'db.r6g.large')
        engine = db_recs.get('engine', 'RDS MySQL')
        
        # Determine pricing based on engine
        if 'Aurora' in engine:
            db_instance_cost = self.pricing['database']['aurora_mysql'].get(instance_class, self.pricing['database']['aurora_mysql']['db.r6g.large'])
        else:
            db_instance_cost = self.pricing['database']['rds_mysql'].get(instance_class, self.pricing['database']['rds_mysql']['db.r6g.large'])
        
        # Multi-AZ costs (double the instance cost)
        if db_recs.get('multi_az'):
            db_instance_cost *= 2
        
        # Storage costs
        db_storage_gb = max(requirements.get('storage_GB', 100), 20)
        
        if 'Aurora' in engine:
            storage_cost = db_storage_gb * self.pricing['database']['storage']['aurora_storage']
            io_cost = 1000000 * self.pricing['database']['storage']['aurora_io']
            storage_cost += io_cost
        else:
            storage_cost = db_storage_gb * self.pricing['database']['storage']['rds_gp2']
        
        # Backup storage costs
        backup_retention = self._extract_backup_days(db_recs.get('backup_retention', '7 days'))
        backup_storage_gb = db_storage_gb * (backup_retention / 30) * 0.2
        backup_cost = backup_storage_gb * self.pricing['database']['backup_storage']
        
        # Read replica costs
        read_replica_cost = 0
        read_replicas = self._extract_read_replica_count(db_recs.get('read_replicas', 'No read replicas'))
        if read_replicas > 0:
            read_replica_cost = db_instance_cost * read_replicas * 0.8
        
        # RDS Proxy costs (if used)
        proxy_cost = 0
        if 'RDS Proxy' in db_recs.get('connection_pooling', ''):
            vcpus = requirements.get('vCPUs', 2)
            proxy_cost = vcpus * self.pricing['database']['rds_proxy'] * 730
        
        total_database = db_instance_cost + storage_cost + backup_cost + read_replica_cost + proxy_cost
        
        return {
            'db_instance': {
                'cost': db_instance_cost,
                'details': f"{instance_class} ({engine}) {'Multi-AZ' if db_recs.get('multi_az') else 'Single-AZ'}"
            },
            'db_storage': {
                'cost': storage_cost,
                'details': f"{db_storage_gb} GB {'Aurora' if 'Aurora' in engine else 'RDS'} storage"
            },
            'backup_storage': {
                'cost': backup_cost,
                'details': f"~{backup_storage_gb:.0f} GB backup retention ({backup_retention} days)"
            },
            'read_replicas': {
                'cost': read_replica_cost,
                'details': f"{read_replicas} read replicas" if read_replicas > 0 else "No read replicas"
            },
            'rds_proxy': {
                'cost': proxy_cost,
                'details': f"RDS Proxy for {vcpus} vCPUs" if proxy_cost > 0 else "No RDS Proxy"
            },
            'total': total_database,
            'optimization_notes': self._get_database_optimization_notes(env, db_recs)
        }
    
    def _calculate_security_costs(self, env: str, security_recs: Dict, requirements: Dict) -> Dict[str, Any]:
        """Calculate security-related costs."""
        
        # Secrets Manager
        secrets_count = self._estimate_secrets_count(env)
        secrets_cost = secrets_count * self.pricing['security']['secrets_manager']
        
        # KMS keys
        kms_keys = self._estimate_kms_keys(env, security_recs)
        kms_cost = kms_keys * self.pricing['security']['kms']
        
        # Config
        config_items = self._estimate_config_items(env)
        config_cost = config_items * self.pricing['security']['config']
        
        # CloudTrail
        cloudtrail_cost = self.pricing['security']['cloudtrail']
        
        # GuardDuty (production environments)
        guardduty_cost = 0
        if env in ['PROD', 'PREPROD']:
            guardduty_cost = self.pricing['security']['guardduty']
        
        # Security Hub
        security_hub_cost = 0
        if env in ['PROD', 'PREPROD']:
            findings_per_month = 1000
            security_hub_cost = findings_per_month * self.pricing['security']['security_hub']
        
        # WAF (if web application)
        waf_cost = 0
        if env in ['PROD', 'PREPROD']:
            waf_cost = self.pricing['security']['waf']
        
        total_security = secrets_cost + kms_cost + config_cost + cloudtrail_cost + guardduty_cost + security_hub_cost + waf_cost
        
        return {
            'secrets_manager': {
                'cost': secrets_cost,
                'details': f"{secrets_count} secrets"
            },
            'kms': {
                'cost': kms_cost,
                'details': f"{kms_keys} customer managed keys"
            },
            'config': {
                'cost': config_cost,
                'details': f"~{config_items} configuration items"
            },
            'cloudtrail': {
                'cost': cloudtrail_cost,
                'details': "Management events trail"
            },
            'guardduty': {
                'cost': guardduty_cost,
                'details': "Threat detection" if guardduty_cost > 0 else "Not enabled"
            },
            'security_hub': {
                'cost': security_hub_cost,
                'details': f"~{1000 if security_hub_cost > 0 else 0} findings/month" if security_hub_cost > 0 else "Not enabled"
            },
            'waf': {
                'cost': waf_cost,
                'details': "Web Application Firewall" if waf_cost > 0 else "Not enabled"
            },
            'total': total_security,
            'optimization_notes': self._get_security_optimization_notes(env)
        }
    
    def _calculate_monitoring_costs(self, env: str, monitoring_recs: Dict, requirements: Dict) -> Dict[str, Any]:
        """Calculate monitoring-related costs."""
        
        # CloudWatch metrics
        custom_metrics = self._estimate_custom_metrics(env)
        metrics_cost = custom_metrics * self.pricing['monitoring']['cloudwatch']['metrics']
        
        # CloudWatch dashboards
        dashboards = self._estimate_dashboards(env)
        dashboard_cost = dashboards * self.pricing['monitoring']['cloudwatch']['dashboards']
        
        # CloudWatch alarms
        alarms = self._estimate_alarms(env)
        alarm_cost = alarms * self.pricing['monitoring']['cloudwatch']['alarms']
        
        # CloudWatch logs
        log_ingestion_gb = self._estimate_log_volume(env)
        log_cost = log_ingestion_gb * self.pricing['monitoring']['cloudwatch']['logs_ingestion']
        log_storage_cost = log_ingestion_gb * self.pricing['monitoring']['cloudwatch']['logs_storage']
        
        # X-Ray (if enabled)
        xray_cost = 0
        if monitoring_recs.get('apm') == 'X-Ray':
            trace_volume = 1
            xray_cost = trace_volume * self.pricing['monitoring']['xray']
        
        # Synthetics (if enabled)
        synthetics_cost = 0
        if monitoring_recs.get('synthetic_monitoring') == 'Required':
            canary_runs = 30 * 24 * 12
            synthetics_cost = canary_runs * self.pricing['monitoring']['synthetics']
        
        total_monitoring = metrics_cost + dashboard_cost + alarm_cost + log_cost + log_storage_cost + xray_cost + synthetics_cost
        
        return {
            'cloudwatch_metrics': {
                'cost': metrics_cost,
                'details': f"{custom_metrics} custom metrics"
            },
            'cloudwatch_dashboards': {
                'cost': dashboard_cost,
                'details': f"{dashboards} dashboards"
            },
            'cloudwatch_alarms': {
                'cost': alarm_cost,
                'details': f"{alarms} alarms"
            },
            'cloudwatch_logs': {
                'cost': log_cost + log_storage_cost,
                'details': f"~{log_ingestion_gb} GB/month ingestion + storage",
                'breakdown': {
                    'ingestion': log_cost,
                    'storage': log_storage_cost
                }
            },
            'xray': {
                'cost': xray_cost,
                'details': "Application tracing" if xray_cost > 0 else "Not enabled"
            },
            'synthetics': {
                'cost': synthetics_cost,
                'details': f"~{canary_runs if synthetics_cost > 0 else 0} canary runs/month" if synthetics_cost > 0 else "Not enabled"
            },
            'total': total_monitoring,
            'optimization_notes': self._get_monitoring_optimization_notes(env)
        }
    
    # Helper methods for cost calculations
    def _get_instance_count(self, env: str) -> int:
        """Get instance count based on environment."""
        counts = {'DEV': 1, 'QA': 1, 'UAT': 2, 'PREPROD': 2, 'PROD': 3}
        return counts.get(env, 2)
    
    def _estimate_cdn_usage(self, env: str) -> float:
        """Estimate CDN data transfer in GB."""
        usage = {'DEV': 0, 'QA': 0, 'UAT': 50, 'PREPROD': 100, 'PROD': 500}
        return usage.get(env, 100)
    
    def _estimate_data_transfer_costs(self, env: str) -> float:
        """Estimate data transfer costs."""
        transfer_gb = {'DEV': 10, 'QA': 20, 'UAT': 50, 'PREPROD': 100, 'PROD': 500}
        gb = transfer_gb.get(env, 50)
        return gb * self.pricing['network']['data_transfer']['out_to_internet_up_to_10tb']
    
    def _extract_iops_from_recommendation(self, iops_rec: str) -> int:
        """Extract IOPS number from recommendation string."""
        import re
        match = re.search(r'(\d+)', iops_rec)
        return int(match.group(1)) if match else 3000
    
    def _get_snapshot_frequency(self, backup_strategy: str) -> float:
        """Get snapshot frequency multiplier."""
        if 'Daily' in backup_strategy:
            return 30
        elif 'Twice daily' in backup_strategy:
            return 60
        elif 'Continuous' in backup_strategy:
            return 365
        return 30
    
    def _extract_backup_days(self, retention: str) -> int:
        """Extract backup retention days."""
        import re
        match = re.search(r'(\d+)', retention)
        return int(match.group(1)) if match else 7
    
    def _extract_read_replica_count(self, replica_config: str) -> int:
        """Extract read replica count."""
        if 'No read replicas' in replica_config:
            return 0
        elif '1 read replica' in replica_config:
            return 1
        elif '1-2 read replicas' in replica_config:
            return 2
        elif '2-3 read replicas' in replica_config:
            return 3
        elif '3+ read replicas' in replica_config:
            return 3
        return 0
    
    def _estimate_secrets_count(self, env: str) -> int:
        """Estimate number of secrets."""
        counts = {'DEV': 3, 'QA': 5, 'UAT': 8, 'PREPROD': 12, 'PROD': 15}
        return counts.get(env, 8)
    
    def _estimate_kms_keys(self, env: str, security_recs: Dict) -> int:
        """Estimate KMS keys needed."""
        base_keys = {'DEV': 1, 'QA': 2, 'UAT': 3, 'PREPROD': 4, 'PROD': 5}
        return base_keys.get(env, 3)
    
    def _estimate_config_items(self, env: str) -> int:
        """Estimate AWS Config items."""
        counts = {'DEV': 10, 'QA': 20, 'UAT': 50, 'PREPROD': 100, 'PROD': 200}
        return counts.get(env, 50)
    
    def _estimate_custom_metrics(self, env: str) -> int:
        """Estimate custom CloudWatch metrics."""
        counts = {'DEV': 5, 'QA': 10, 'UAT': 25, 'PREPROD': 50, 'PROD': 100}
        return counts.get(env, 25)
    
    def _estimate_dashboards(self, env: str) -> int:
        """Estimate CloudWatch dashboards."""
        counts = {'DEV': 1, 'QA': 1, 'UAT': 2, 'PREPROD': 3, 'PROD': 5}
        return counts.get(env, 2)
    
    def _estimate_alarms(self, env: str) -> int:
        """Estimate CloudWatch alarms."""
        counts = {'DEV': 5, 'QA': 10, 'UAT': 20, 'PREPROD': 40, 'PROD': 80}
        return counts.get(env, 20)
    
    def _estimate_log_volume(self, env: str) -> float:
        """Estimate log volume in GB per month."""
        volumes = {'DEV': 1, 'QA': 2, 'UAT': 5, 'PREPROD': 15, 'PROD': 50}
        return volumes.get(env, 10)
    
    # Optimization notes methods
    def _get_compute_optimization_notes(self, env: str, instance_type: str, pricing_model: str) -> List[str]:
        """Get compute optimization recommendations."""
        notes = []
        if pricing_model == 'on_demand' and env in ['PROD', 'PREPROD']:
            notes.append("💡 Consider Reserved Instances for 20-40% cost savings")
        if 'large' in instance_type and env == 'DEV':
            notes.append("💡 Consider smaller instances for development environment")
        notes.append("💡 Monitor CPU utilization to right-size instances")
        return notes
    
    def _get_network_optimization_notes(self, env: str) -> List[str]:
        """Get network optimization recommendations."""
        notes = [
            "💡 Use CloudFront to reduce data transfer costs",
            "💡 Implement VPC endpoints to avoid NAT Gateway costs for AWS services"
        ]
        if env in ['DEV', 'QA']:
            notes.append("💡 Single NAT Gateway sufficient for non-production")
        return notes
    
    def _get_storage_optimization_notes(self, env: str, storage_recs: Dict) -> List[str]:
        """Get storage optimization recommendations."""
        notes = [
            "💡 Use gp3 volumes for better price-performance ratio",
            "💡 Implement lifecycle policies for S3 backup storage"
        ]
        if env in ['PROD', 'PREPROD']:
            notes.append("💡 Consider EBS snapshots cross-region replication")
        return notes
    
    def _get_database_optimization_notes(self, env: str, db_recs: Dict) -> List[str]:
        """Get database optimization recommendations."""
        notes = [
            "💡 Use Aurora for better price-performance at scale",
            "💡 Monitor database utilization for right-sizing"
        ]
        if env not in ['PROD']:
            notes.append("💡 Single-AZ deployment sufficient for non-production")
        return notes
    
    def _get_security_optimization_notes(self, env: str) -> List[str]:
        """Get security optimization recommendations."""
        notes = [
            "💡 Use AWS managed keys where possible to reduce costs",
            "💡 Implement proper log retention policies"
        ]
        if env in ['DEV', 'QA']:
            notes.append("💡 Reduce security tooling for development environments")
        return notes
    
    def _get_monitoring_optimization_notes(self, env: str) -> List[str]:
        """Get monitoring optimization recommendations."""
        notes = [
            "💡 Use metric filters to reduce custom metric costs",
            "💡 Implement log retention policies to control storage costs"
        ]
        if env in ['DEV', 'QA']:
            notes.append("💡 Reduce monitoring frequency for development environments")
        return notes
        
class EnhancedEnvironmentAnalyzer:
    """Enhanced environment analyzer with detailed complexity explanations."""
    
    def __init__(self):
        self.environments = ['DEV', 'QA', 'UAT', 'PREPROD', 'PROD']
        self.cost_calculator = AWSCostCalculator()
        
    def get_detailed_complexity_explanation(self, env: str, env_results: Dict) -> Dict[str, Any]:
        """Get detailed explanation of environment complexity."""
        
        claude_analysis = env_results.get('claude_analysis', {})
        complexity_score = claude_analysis.get('complexity_score', 50)
        requirements = env_results.get('requirements', {})
        
        # Calculate specific complexity factors
        factors = {
            'Resource Intensity': self._calculate_resource_intensity(requirements),
            'Migration Risk': self._calculate_migration_risk(env, claude_analysis),
            'Operational Complexity': self._calculate_operational_complexity(env),
            'Compliance Requirements': self._calculate_compliance_complexity(env),
            'Integration Dependencies': self._calculate_integration_complexity(env)
        }
        
        return {
            'overall_score': complexity_score,
            'complexity_level': claude_analysis.get('complexity_level', 'MEDIUM'),
            'factors': factors,
            'detailed_reasons': self._generate_detailed_reasons(env, factors, complexity_score),
            'specific_challenges': self._identify_specific_challenges(env, factors),
            'mitigation_strategies': self._generate_mitigation_strategies(env, factors)
        }
    
    def _calculate_resource_intensity(self, requirements: Dict) -> Dict[str, Any]:
        """Calculate resource intensity factor."""
        vcpus = requirements.get('vCPUs', 2)
        ram_gb = requirements.get('RAM_GB', 8)
        storage_gb = requirements.get('storage_GB', 100)
        
        cpu_intensity = min(vcpus / 2, 10) * 10
        memory_intensity = min(ram_gb / 8, 10) * 10
        storage_intensity = min(storage_gb / 100, 10) * 10
        
        overall_score = (cpu_intensity + memory_intensity + storage_intensity) / 3
        
        return {
            'score': overall_score,
            'cpu_intensity': cpu_intensity,
            'memory_intensity': memory_intensity,
            'storage_intensity': storage_intensity,
            'description': self._get_resource_description(overall_score)
        }
    
    def _calculate_migration_risk(self, env: str, claude_analysis: Dict) -> Dict[str, Any]:
        """Calculate migration risk factor."""
        base_risks = {'DEV': 20, 'QA': 30, 'UAT': 50, 'PREPROD': 70, 'PROD': 90}
        base_score = base_risks.get(env, 50)
        
        complexity_multiplier = claude_analysis.get('complexity_score', 50) / 50
        final_score = min(base_score * complexity_multiplier, 100)
        
        return {
            'score': final_score,
            'base_risk': base_score,
            'risk_level': self._get_risk_level(final_score),
            'description': self._get_migration_risk_description(env, final_score)
        }
    
    def _calculate_operational_complexity(self, env: str) -> Dict[str, Any]:
        """Calculate operational complexity."""
        complexity_factors = {
            'DEV': {
                'score': 25,
                'factors': ['Minimal SLA requirements', 'Simple monitoring', 'Basic security'],
                'description': 'Low operational complexity - development environment'
            },
            'QA': {
                'score': 35,
                'factors': ['Automated testing integration', 'Test data management'],
                'description': 'Low-medium complexity - automated testing requirements'
            },
            'UAT': {
                'score': 55,
                'factors': ['User access management', 'Performance validation'],
                'description': 'Medium complexity - user acceptance validation'
            },
            'PREPROD': {
                'score': 75,
                'factors': ['Production-like configuration', 'Advanced monitoring'],
                'description': 'High complexity - production simulation requirements'
            },
            'PROD': {
                'score': 90,
                'factors': ['24/7 availability', 'Advanced monitoring & alerting', 'Disaster recovery'],
                'description': 'Very high complexity - business-critical operations'
            }
        }
        return complexity_factors.get(env, {'score': 50, 'factors': [], 'description': 'Medium complexity'})
    
    def _calculate_compliance_complexity(self, env: str) -> Dict[str, Any]:
        """Calculate compliance complexity."""
        compliance_scores = {'DEV': 10, 'QA': 20, 'UAT': 40, 'PREPROD': 70, 'PROD': 95}
        score = compliance_scores.get(env, 50)
        
        return {
            'score': score,
            'level': self._get_compliance_level(score),
            'requirements': self._get_compliance_requirements(env),
            'description': f'{env} environment compliance requirements'
        }
    
    def _calculate_integration_complexity(self, env: str) -> Dict[str, Any]:
        """Calculate integration complexity."""
        integration_scores = {'DEV': 30, 'QA': 45, 'UAT': 60, 'PREPROD': 80, 'PROD': 95}
        score = integration_scores.get(env, 50)
        
        return {
            'score': score,
            'complexity_level': self._get_integration_level(score),
            'integration_points': self._get_integration_points(env),
            'description': f'{env} environment integration requirements'
        }
    
    def _generate_detailed_reasons(self, env: str, factors: Dict, overall_score: float) -> List[str]:
        """Generate detailed reasons for complexity score."""
        reasons = []
        
        resource_factor = factors['Resource Intensity']
        if resource_factor['score'] > 70:
            reasons.append(f"High resource requirements: {resource_factor['description']}")
        elif resource_factor['score'] > 40:
            reasons.append(f"Moderate resource requirements: {resource_factor['description']}")
        else:
            reasons.append(f"Light resource requirements: {resource_factor['description']}")
        
        migration_factor = factors['Migration Risk']
        if migration_factor['score'] > 70:
            reasons.append(f"High migration risk due to {env} environment criticality")
        
        ops_factor = factors['Operational Complexity']
        if ops_factor['score'] > 60:
            reasons.append(f"Complex operational requirements: {', '.join(ops_factor['factors'][:2])}")
        
        compliance_factor = factors['Compliance Requirements']
        if compliance_factor['score'] > 50:
            reasons.append(f"Significant compliance requirements for {env} environment")
        
        return reasons
    
    def _identify_specific_challenges(self, env: str, factors: Dict) -> List[str]:
        """Identify specific challenges for the environment."""
        env_challenges = {
            'DEV': [
                'Frequent code deployments and updates',
                'Developer access and permissions management',
                'Integration with development tools and CI/CD'
            ],
            'QA': [
                'Automated testing integration',
                'Test data management and refresh',
                'Performance and load testing capabilities'
            ],
            'UAT': [
                'User acceptance testing coordination',
                'Business stakeholder access management',
                'Production-like data simulation'
            ],
            'PREPROD': [
                'Production parity maintenance',
                'Disaster recovery testing',
                'Performance validation under load'
            ],
            'PROD': [
                'Zero-downtime deployment requirements',
                'Business continuity and disaster recovery',
                'Advanced monitoring and alerting',
                'Compliance and audit requirements'
            ]
        }
        return env_challenges.get(env, ['Standard migration challenges'])
    
    def _generate_mitigation_strategies(self, env: str, factors: Dict) -> List[str]:
        """Generate mitigation strategies."""
        env_strategies = {
            'DEV': [
                'Implement Infrastructure as Code for consistent deployments',
                'Use containerization for development environment isolation',
                'Establish automated backup and restore procedures'
            ],
            'QA': [
                'Implement automated testing pipelines',
                'Use data masking for sensitive test data',
                'Establish performance baselines and monitoring'
            ],
            'UAT': [
                'Create detailed user acceptance testing plans',
                'Implement role-based access controls',
                'Establish change management procedures'
            ],
            'PREPROD': [
                'Maintain production parity through automation',
                'Implement comprehensive disaster recovery testing',
                'Use blue-green deployment strategies'
            ],
            'PROD': [
                'Implement zero-downtime deployment strategies',
                'Establish comprehensive monitoring and alerting',
                'Create detailed disaster recovery plans',
                'Implement advanced security and compliance controls'
            ]
        }
        return env_strategies.get(env, ['Follow standard migration best practices'])
    
    def get_technical_recommendations(self, env: str, env_results: Dict) -> Dict[str, Any]:
        """Get comprehensive technical recommendations."""
        
        requirements = env_results.get('requirements', {})
        cost_breakdown = env_results.get('cost_breakdown', {})
        selected_instance = cost_breakdown.get('selected_instance', {})
        
        return {
            'compute': self._get_compute_recommendations(env, selected_instance, requirements),
            'network': self._get_network_recommendations(env, requirements),
            'storage': self._get_storage_recommendations(env, requirements),
            'database': self._get_database_recommendations(env, requirements),
            'security': self._get_security_recommendations(env),
            'monitoring': self._get_monitoring_recommendations(env),
            'backup': self._get_backup_recommendations(env),
            'scaling': self._get_scaling_recommendations(env, requirements)
        }
    
    def _get_compute_recommendations(self, env: str, selected_instance: Dict, requirements: Dict) -> Dict[str, Any]:
        """Get compute recommendations."""
        instance_type = selected_instance.get('type', 'N/A')
        vcpus = selected_instance.get('vCPU', requirements.get('vCPUs', 2))
        ram_gb = selected_instance.get('RAM', requirements.get('RAM_GB', 8))
        
        return {
            'primary_instance': {
                'type': instance_type,
                'vcpus': vcpus,
                'memory_gb': ram_gb,
                'rationale': f'Selected {instance_type} for {env} environment based on performance requirements'
            },
            'alternative_instances': [
                {'type': 'm6a.large', 'rationale': 'AMD-based cost optimization'},
                {'type': 'c6i.xlarge', 'rationale': 'Compute-optimized alternative'}
            ],
            'placement_strategy': self._get_placement_strategy(env),
            'auto_scaling': self._get_auto_scaling_config(env),
            'pricing_optimization': self._get_pricing_optimization(env)
        }
    
    def _get_network_recommendations(self, env: str, requirements: Dict) -> Dict[str, Any]:
        """Get network recommendations."""
        network_configs = {
            'DEV': {
                'vpc_design': 'Single AZ, basic VPC',
                'subnets': 'Public and private subnets',
                'security_groups': 'Development-focused security groups',
                'load_balancer': 'Application Load Balancer (if needed)',
                'cdn': 'Optional',
                'dns': 'Route 53 basic',
                'nat_gateway': 'Optional',
                'vpn': 'Site-to-site VPN for secure access'
            },
            'QA': {
                'vpc_design': 'Single AZ with testing isolation',
                'subnets': 'Isolated testing subnets',
                'security_groups': 'Testing-specific security groups',
                'load_balancer': 'ALB for load testing',
                'cdn': 'Optional',
                'dns': 'Route 53 basic',
                'nat_gateway': 'Optional',
                'vpn': 'Site-to-site VPN for secure access'
            },
            'UAT': {
                'vpc_design': 'Multi-AZ for availability testing',
                'subnets': 'Production-like subnet design',
                'security_groups': 'Production-like security',
                'load_balancer': 'ALB with SSL termination',
                'cdn': 'Optional',
                'dns': 'Route 53 basic',
                'nat_gateway': 'Optional',
                'vpn': 'Site-to-site VPN for secure access'
            },
            'PREPROD': {
                'vpc_design': 'Full multi-AZ production design',
                'subnets': 'Production mirror subnet design',
                'security_groups': 'Production security groups',
                'load_balancer': 'ALB/NLB with full features',
                'cdn': 'CloudFront',
                'dns': 'Route 53 with health checks',
                'nat_gateway': 'Required',
                'vpn': 'Site-to-site VPN for secure access'
            },
            'PROD': {
                'vpc_design': 'Multi-AZ with disaster recovery',
                'subnets': 'Highly available subnet design',
                'security_groups': 'Strict production security',
                'load_balancer': 'ALB/NLB with advanced features',
                'cdn': 'CloudFront',
                'dns': 'Route 53 with health checks',
                'nat_gateway': 'Required',
                'vpn': 'Site-to-site VPN for secure access'
            }
        }
        return network_configs.get(env, network_configs['UAT'])
    
    def _get_storage_recommendations(self, env: str, requirements: Dict) -> Dict[str, Any]:
        """Get storage recommendations."""
        storage_gb = requirements.get('storage_GB', 100)
        
        storage_configs = {
            'DEV': {
                'primary_storage': 'gp3 (General Purpose SSD)',
                'backup_strategy': 'Daily snapshots, 7-day retention',
                'encryption': 'EBS encryption enabled',
                'performance': 'Standard performance'
            },
            'QA': {
                'primary_storage': 'gp3 (General Purpose SSD)',
                'backup_strategy': 'Daily snapshots, 14-day retention',
                'encryption': 'EBS encryption enabled',
                'performance': 'Enhanced performance for testing'
            },
            'UAT': {
                'primary_storage': 'gp3 with provisioned IOPS',
                'backup_strategy': 'Twice daily snapshots, 30-day retention',
                'encryption': 'EBS encryption with customer keys',
                'performance': 'Production-like performance'
            },
            'PREPROD': {
                'primary_storage': 'io2 (Provisioned IOPS SSD)',
                'backup_strategy': 'Continuous backup, 90-day retention',
                'encryption': 'EBS encryption with customer keys',
                'performance': 'High performance, production parity'
            },
            'PROD': {
                'primary_storage': 'io2 (Provisioned IOPS SSD)',
                'backup_strategy': 'Continuous backup, cross-region replication',
                'encryption': 'EBS encryption with customer managed keys',
                'performance': 'Maximum performance optimization'
            }
        }
        
        config = storage_configs.get(env, storage_configs['UAT'])
        config.update({
            'recommended_size': f"{storage_gb * self._get_storage_multiplier(env)} GB",
            'iops_recommendation': self._get_iops_recommendation(env, storage_gb),
            'throughput_recommendation': self._get_throughput_recommendation(env),
            'lifecycle_policy': self._get_lifecycle_policy(env)
        })
        return config
    
    def _get_database_recommendations(self, env: str, requirements: Dict) -> Dict[str, Any]:
        """Get database recommendations."""
        db_configs = {
            'DEV': {
                'engine': 'RDS MySQL/PostgreSQL',
                'instance_class': 'db.t4g.micro or db.t4g.small',
                'multi_az': False,
                'backup_retention': '7 days',
                'encryption': 'Enabled',
                'monitoring': 'Basic CloudWatch'
            },
            'QA': {
                'engine': 'RDS MySQL/PostgreSQL',
                'instance_class': 'db.t4g.small or db.t4g.medium',
                'multi_az': False,
                'backup_retention': '14 days',
                'encryption': 'Enabled',
                'monitoring': 'Enhanced monitoring'
            },
            'UAT': {
                'engine': 'RDS MySQL/PostgreSQL',
                'instance_class': 'db.r6g.large',
                'multi_az': True,
                'backup_retention': '30 days',
                'encryption': 'Enabled with customer keys',
                'monitoring': 'Performance Insights'
            },
            'PREPROD': {
                'engine': 'RDS MySQL/PostgreSQL/Aurora',
                'instance_class': 'db.r6g.xlarge',
                'multi_az': True,
                'backup_retention': '35 days',
                'encryption': 'Enabled with customer managed keys',
                'monitoring': 'Performance Insights + Enhanced monitoring'
            },
            'PROD': {
                'engine': 'Aurora MySQL/PostgreSQL',
                'instance_class': 'db.r6g.xlarge or higher',
                'multi_az': True,
                'backup_retention': '35 days with cross-region backup',
                'encryption': 'Enabled with customer managed keys',
                'monitoring': 'Full Performance Insights + Enhanced monitoring'
            }
        }
        
        config = db_configs.get(env, db_configs['UAT'])
        config.update({
            'read_replicas': self._get_read_replica_config(env),
            'connection_pooling': 'RDS Proxy' if env in ['PREPROD', 'PROD'] else 'Application-level',
            'maintenance_window': self._get_maintenance_window(env)
        })
        return config
    
    def _get_security_recommendations(self, env: str) -> Dict[str, Any]:
        """Get security recommendations."""
        security_configs = {
            'DEV': {
                'iam_roles': 'Developer access with resource restrictions',
                'encryption': 'Standard EBS and S3 encryption',
                'network_security': 'Basic security groups and NACLs',
                'compliance': 'Basic security standards',
                'monitoring': 'CloudTrail for audit logging',
                'secrets_management': 'AWS Secrets Manager',
                'certificate_management': 'ACM with auto-renewal'
            },
            'PROD': {
                'iam_roles': 'Least privilege production access',
                'encryption': 'Customer-managed keys with rotation',
                'network_security': 'Maximum security controls',
                'compliance': 'Full regulatory compliance',
                'monitoring': 'Complete security and compliance monitoring',
                'secrets_management': 'AWS Secrets Manager',
                'certificate_management': 'ACM with auto-renewal'
            }
        }
        return security_configs.get(env, security_configs['DEV'])
    
    def _get_monitoring_recommendations(self, env: str) -> Dict[str, Any]:
        """Get monitoring recommendations."""
        monitoring_configs = {
            'DEV': {
                'cloudwatch': 'Basic metrics and logs',
                'alerting': 'Development team notifications',
                'dashboards': 'Basic development dashboard',
                'log_retention': '30 days',
                'apm': 'Optional',
                'synthetic_monitoring': 'Not required',
                'cost_monitoring': 'Cost Explorer + Budgets',
                'health_checks': 'Basic'
            },
            'PROD': {
                'cloudwatch': 'Premium monitoring with custom metrics',
                'alerting': '24/7 operations with escalation',
                'dashboards': 'Executive and operations dashboards',
                'log_retention': 'Long-term retention (3+ years)',
                'apm': 'X-Ray',
                'synthetic_monitoring': 'Required',
                'cost_monitoring': 'Cost Explorer + Budgets',
                'health_checks': 'Route 53 health checks'
            }
        }
        return monitoring_configs.get(env, monitoring_configs['DEV'])
    
    def _get_backup_recommendations(self, env: str) -> Dict[str, Any]:
        """Get backup recommendations."""
        backup_configs = {
            'DEV': {'frequency': 'Daily', 'retention': '7 days', 'cross_region': False, 'testing': 'Monthly'},
            'QA': {'frequency': 'Daily', 'retention': '14 days', 'cross_region': False, 'testing': 'Bi-weekly'},
            'UAT': {'frequency': 'Twice daily', 'retention': '30 days', 'cross_region': False, 'testing': 'Weekly'},
            'PREPROD': {'frequency': 'Every 6 hours', 'retention': '90 days', 'cross_region': True, 'testing': 'Weekly'},
            'PROD': {'frequency': 'Continuous (point-in-time recovery)', 'retention': '7 years (compliance)', 'cross_region': True, 'testing': 'Daily'}
        }
        return backup_configs.get(env, backup_configs['UAT'])
    
    def _get_scaling_recommendations(self, env: str, requirements: Dict) -> Dict[str, Any]:
        """Get scaling recommendations."""
        scaling_configs = {
            'DEV': {'auto_scaling': 'Basic scaling for cost optimization', 'min_instances': 1, 'max_instances': 2, 'target_utilization': '70%'},
            'QA': {'auto_scaling': 'Load testing optimized scaling', 'min_instances': 1, 'max_instances': 4, 'target_utilization': '60%'},
            'UAT': {'auto_scaling': 'User load optimized scaling', 'min_instances': 2, 'max_instances': 6, 'target_utilization': '65%'},
            'PREPROD': {'auto_scaling': 'Production-like scaling', 'min_instances': 2, 'max_instances': 10, 'target_utilization': '70%'},
            'PROD': {'auto_scaling': 'High availability scaling', 'min_instances': 3, 'max_instances': 20, 'target_utilization': '75%'}
        }
        return scaling_configs.get(env, scaling_configs['UAT'])
    
    # Helper methods
    def _get_resource_description(self, score: float) -> str:
        if score > 70:
            return "High resource intensity requiring powerful instances"
        elif score > 40:
            return "Moderate resource requirements"
        else:
            return "Light resource requirements suitable for smaller instances"
    
    def _get_risk_level(self, score: float) -> str:
        if score > 80: return "Very High"
        elif score > 60: return "High"
        elif score > 40: return "Medium"
        elif score > 20: return "Low"
        else: return "Very Low"
    
    def _get_migration_risk_description(self, env: str, score: float) -> str:
        risk_descriptions = {
            'DEV': "Development environment - can be rebuilt if needed",
            'QA': "Testing environment - important for development process",
            'UAT': "User acceptance environment - critical for business validation",
            'PREPROD': "Pre-production environment - high business impact",
            'PROD': "Production environment - maximum business criticality"
        }
        return risk_descriptions.get(env, "Standard environment risk")
    
    def _get_compliance_level(self, score: float) -> str:
        if score > 80: return "Full Compliance"
        elif score > 60: return "High Compliance"
        elif score > 40: return "Medium Compliance"
        else: return "Basic Compliance"
    
    def _get_compliance_requirements(self, env: str) -> List[str]:
        requirements = {
            'DEV': ['Basic security standards', 'Data protection'],
            'QA': ['Testing data compliance', 'Security standards'],
            'UAT': ['User data protection', 'Business compliance'],
            'PREPROD': ['Production-like compliance', 'Security validation'],
            'PROD': ['Full regulatory compliance', 'Audit requirements', 'Data sovereignty']
        }
        return requirements.get(env, ['Standard compliance'])
    
    def _get_integration_level(self, score: float) -> str:
        if score > 80: return "Complex"
        elif score > 60: return "Moderate"
        else: return "Simple"
    
    def _get_integration_points(self, env: str) -> List[str]:
        integrations = {
            'DEV': ['CI/CD systems', 'Development tools', 'Version control'],
            'QA': ['Testing frameworks', 'Test data systems', 'Reporting tools'],
            'UAT': ['Business applications', 'User directories', 'Approval systems'],
            'PREPROD': ['Production integrations', 'Monitoring systems', 'External APIs'],
            'PROD': ['All business systems', 'External partners', 'Real-time integrations']
        }
        return integrations.get(env, ['Standard integrations'])
    
    def _get_placement_strategy(self, env: str) -> str:
        strategies = {
            'DEV': 'Single AZ placement for cost optimization',
            'QA': 'Single AZ with availability considerations',
            'UAT': 'Multi-AZ for availability testing',
            'PREPROD': 'Multi-AZ production-like placement',
            'PROD': 'Multi-AZ with placement groups for performance'
        }
        return strategies.get(env, 'Standard placement')
    
    def _get_auto_scaling_config(self, env: str) -> str:
        configs = {
            'DEV': 'Cost-optimized scaling',
            'QA': 'Load testing optimized',
            'UAT': 'User load optimized',
            'PREPROD': 'Production-like scaling',
            'PROD': 'High availability scaling'
        }
        return configs.get(env, 'Standard scaling')
    
    def _get_pricing_optimization(self, env: str) -> str:
        optimizations = {
            'DEV': 'Spot instances recommended for cost savings',
            'QA': 'Mix of on-demand and spot instances',
            'UAT': 'On-demand with some reserved instances',
            'PREPROD': 'Reserved instances for predictable costs',
            'PROD': 'Reserved instances with savings plans'
        }
        return optimizations.get(env, 'Standard pricing')
    
    def _get_storage_multiplier(self, env: str) -> float:
        multipliers = {'DEV': 1.0, 'QA': 1.2, 'UAT': 1.3, 'PREPROD': 1.4, 'PROD': 1.5}
        return multipliers.get(env, 1.2)
    
    def _get_iops_recommendation(self, env: str, storage_gb: int) -> str:
        if env in ['PREPROD', 'PROD']:
            return f"{max(storage_gb * 3, 1000)} IOPS (Provisioned)"
        else:
            return f"{storage_gb * 3} IOPS (Baseline gp3)"
    
    def _get_throughput_recommendation(self, env: str) -> str:
        throughput = {
            'DEV': '125 MB/s (gp3 baseline)',
            'QA': '250 MB/s (enhanced)',
            'UAT': '500 MB/s (production-like)',
            'PREPROD': '1000 MB/s (high performance)',
            'PROD': '1000+ MB/s (maximum performance)'
        }
        return throughput.get(env, '250 MB/s')
    
    def _get_lifecycle_policy(self, env: str) -> str:
        policies = {
            'DEV': 'Move to IA after 30 days, archive after 90 days',
            'QA': 'Move to IA after 60 days, archive after 180 days',
            'UAT': 'Move to IA after 90 days, archive after 365 days',
            'PREPROD': 'Move to IA after 180 days, archive after 2 years',
            'PROD': 'Move to IA after 365 days, archive after 3 years'
        }
        return policies.get(env, 'Standard lifecycle policy')
    
    def _get_read_replica_config(self, env: str) -> str:
        configs = {
            'DEV': 'No read replicas needed',
            'QA': '1 read replica for testing',
            'UAT': '1-2 read replicas for user load',
            'PREPROD': '2-3 read replicas for production testing',
            'PROD': '3+ read replicas across AZs'
        }
        return configs.get(env, '1 read replica')
    
    def _get_maintenance_window(self, env: str) -> str:
        windows = {
            'DEV': 'Any time (flexible)',
            'QA': 'Off-hours testing window',
            'UAT': 'Business off-hours',
            'PREPROD': 'Coordinated with production',
            'PROD': 'Strict maintenance windows'
        }
        return windows.get(env, 'Standard maintenance window')
    
class EnhancedEnterpriseEC2Calculator:
    """Enhanced calculator with comprehensive instance types and environment support plus vROPS integration."""
    
    def __init__(self):
        try:
            self.claude_analyzer = ClaudeAIMigrationAnalyzer()
            self.vrops_processor = VROPSMetricsProcessor()
            
            # Comprehensive instance types
            self.INSTANCE_TYPES = [
                {
                    "type": "m6i.large", "vCPU": 2, "RAM": 8, "max_ebs_bandwidth": 4750,
                    "network": "Up to 12.5 Gbps", "family": "general", "processor": "Intel Xeon Ice Lake",
                    "architecture": "x86_64", "storage": "EBS Only", "network_performance": "Up to 12.5 Gigabit",
                    "ebs_optimized": True, "enhanced_networking": True, "placement_group": True
                },
                {
                    "type": "m6i.xlarge", "vCPU": 4, "RAM": 16, "max_ebs_bandwidth": 9500,
                    "network": "Up to 12.5 Gbps", "family": "general", "processor": "Intel Xeon Ice Lake",
                    "architecture": "x86_64", "storage": "EBS Only", "network_performance": "Up to 12.5 Gigabit",
                    "ebs_optimized": True, "enhanced_networking": True, "placement_group": True
                },
                {
                    "type": "m6i.2xlarge", "vCPU": 8, "RAM": 32, "max_ebs_bandwidth": 19000,
                    "network": "Up to 12.5 Gbps", "family": "general", "processor": "Intel Xeon Ice Lake",
                    "architecture": "x86_64", "storage": "EBS Only", "network_performance": "Up to 12.5 Gigabit",
                    "ebs_optimized": True, "enhanced_networking": True, "placement_group": True
                },
                {
                    "type": "m6i.4xlarge", "vCPU": 16, "RAM": 64, "max_ebs_bandwidth": 38000,
                    "network": "Up to 12.5 Gbps", "family": "general", "processor": "Intel Xeon Ice Lake",
                    "architecture": "x86_64", "storage": "EBS Only", "network_performance": "Up to 12.5 Gigabit",
                    "ebs_optimized": True, "enhanced_networking": True, "placement_group": True
                },
                {
                    "type": "r6i.large", "vCPU": 2, "RAM": 16, "max_ebs_bandwidth": 4750,
                    "network": "Up to 12.5 Gbps", "family": "memory", "processor": "Intel Xeon Ice Lake",
                    "architecture": "x86_64", "storage": "EBS Only", "network_performance": "Up to 12.5 Gigabit",
                    "ebs_optimized": True, "enhanced_networking": True, "placement_group": True
                },
                {
                    "type": "r6i.xlarge", "vCPU": 4, "RAM": 32, "max_ebs_bandwidth": 9500,
                    "network": "Up to 12.5 Gbps", "family": "memory", "processor": "Intel Xeon Ice Lake",
                    "architecture": "x86_64", "storage": "EBS Only", "network_performance": "Up to 12.5 Gigabit",
                    "ebs_optimized": True, "enhanced_networking": True, "placement_group": True
                }
            ]
            
            # Environment multipliers
            self.ENV_MULTIPLIERS = {
                "PROD": {"cpu_ram": 1.0, "storage": 1.0, "description": "Production environment"},
                "PREPROD": {"cpu_ram": 0.9, "storage": 0.9, "description": "Pre-production environment"},
                "UAT": {"cpu_ram": 0.7, "storage": 0.7, "description": "User acceptance testing"},
                "QA": {"cpu_ram": 0.6, "storage": 0.6, "description": "Quality assurance"},
                "DEV": {"cpu_ram": 0.4, "storage": 0.4, "description": "Development environment"}
            }
            
            # Default inputs
            self.inputs = {
                "workload_name": "Sample Enterprise Workload",
                "workload_type": "web_application",
                "operating_system": "linux",
                "region": "us-east-1",
                "on_prem_cores": 8,
                "peak_cpu_percent": 70,
                "on_prem_ram_gb": 32,
                "peak_ram_percent": 80,
                "storage_current_gb": 500,
                "storage_growth_rate": 0.15,
                "peak_iops": 5000,
                "peak_throughput_mbps": 250,
                "infrastructure_age_years": 3,
                "business_criticality": "medium"
            }
            
            logger.info("Enhanced calculator initialized successfully")
        except Exception as e:
            logger.error(f"Error initializing calculator: {e}")
            raise

    def calculate_enhanced_requirements(self, env: str, vrops_data: Dict = None) -> Dict[str, Any]:
        """Calculate requirements with Claude AI analysis and optional vROPS data."""
        
        try:
            # If vROPS data is available, use it to override/enhance inputs
            if vrops_data and vrops_data.get('status') == 'success':
                self._enhance_inputs_with_vrops(vrops_data)
            
            # Standard requirements calculation
            requirements = self._calculate_standard_requirements(env)
            
            # Pass operating system to cost calculations
            operating_system = self.inputs.get('operating_system', 'linux')
            
             # Update cost calculation to include OS
        if 'cost_breakdown' in requirements:
            # Recalculate with OS-specific pricing
            vcpus = requirements['requirements']['vCPUs']
            ram_gb = requirements['requirements']['RAM_GB'] 
            storage_gb = requirements['requirements']['storage_GB']
            
            # Get OS-specific cost breakdown
            selected_instance = self._select_best_instance(vcpus, ram_gb)
            instance_type = selected_instance['type']
            
            os_pricing = self._get_ec2_pricing_with_os(instance_type, operating_system)
            
            # Update the cost breakdown
            requirements['cost_breakdown']['os_pricing'] = os_pricing
            requirements['cost_breakdown']['operating_system'] = operating_system
            
            # Recalculate total costs with OS pricing
            monthly_instance_cost = {}
            for pricing_model, hourly_rate in os_pricing.items():
                if pricing_model not in ['source', 'last_updated', 'os_multiplier']:
                    monthly_instance_cost[pricing_model] = hourly_rate * 730
            
            storage_monthly = storage_gb * 0.08
            network_monthly = 50
            
            total_costs = {}
            for pricing_model, instance_cost in monthly_instance_cost.items():
                total_costs[pricing_model] = instance_cost + storage_monthly + network_monthly
            
            requirements['cost_breakdown']['total_costs'] = total_costs
            requirements['cost_breakdown']['instance_costs'] = monthly_instance_cost
        
        # Claude AI migration analysis with vROPS data and OS info
        claude_analysis = self.claude_analyzer.analyze_workload_complexity(self.inputs, env, vrops_data)
        
        # Enhanced results
        enhanced_results = {
            **requirements,
            'claude_analysis': claude_analysis,
            'environment': env,
            'operating_system': operating_system,
            'vrops_data': vrops_data if vrops_data and vrops_data.get('status') == 'success' else None
        }
        
        return enhanced_results
    except Exception as e:
        logger.error(f"Error in enhanced requirements calculation: {e}")
        return self._get_fallback_requirements(env)

    def _enhance_inputs_with_vrops(self, vrops_data: Dict):
        """Enhance workload inputs with vROPS sizing recommendations."""
        try:
            aws_sizing = vrops_data.get('aws_sizing', {})
            
            # Update CPU and memory based on vROPS recommendations
            if aws_sizing.get('recommended_vcpus'):
                self.inputs['on_prem_cores'] = aws_sizing['recommended_vcpus']
            
            if aws_sizing.get('recommended_memory_gb'):
                self.inputs['on_prem_ram_gb'] = aws_sizing['recommended_memory_gb']
            
            if aws_sizing.get('recommended_storage_gb'):
                self.inputs['storage_current_gb'] = aws_sizing['recommended_storage_gb']
            
            if aws_sizing.get('estimated_iops'):
                self.inputs['peak_iops'] = aws_sizing['estimated_iops']
            
            # Update utilization metrics from performance data
            performance_summary = vrops_data.get('performance_summary', {})
            if performance_summary.get('cpu_efficiency'):
                self.inputs['peak_cpu_percent'] = performance_summary['cpu_efficiency']
            
            if performance_summary.get('memory_efficiency'):
                self.inputs['peak_ram_percent'] = performance_summary['memory_efficiency']
                
        except Exception as e:
            logger.error(f"Error enhancing inputs with vROPS data: {e}")

    def _calculate_standard_requirements(self, env: str) -> Dict[str, Any]:
        """Calculate standard infrastructure requirements."""
        try:
            env_mult = self.ENV_MULTIPLIERS[env]
            
            required_vcpus = max(math.ceil(self.inputs["on_prem_cores"] * 1.2 * env_mult["cpu_ram"]), 2)
            required_ram = max(math.ceil(self.inputs["on_prem_ram_gb"] * 1.3 * env_mult["cpu_ram"]), 4)
            required_storage = math.ceil(self.inputs["storage_current_gb"] * 1.2 * env_mult["storage"])
            
            return {
                "requirements": {
                    "vCPUs": required_vcpus,
                    "RAM_GB": required_ram,
                    "storage_GB": required_storage,
                    "multi_az": env in ["PROD", "PREPROD"]
                },
                "cost_breakdown": self._calculate_basic_costs(required_vcpus, required_ram, required_storage, env),
                "tco_analysis": self._calculate_tco(required_vcpus, required_ram, env)
            }
        except Exception as e:
            logger.error(f"Error calculating standard requirements: {e}")
            return self._get_fallback_requirements(env)

    def _calculate_basic_costs(self, vcpus: int, ram_gb: int, storage_gb: int, env: str) -> Dict[str, Any]:
        """Calculate basic costs with realistic EC2 pricing."""
        try:
            selected_instance = self._select_best_instance(vcpus, ram_gb)
            instance_type = selected_instance['type']
            
            pricing = self._get_fallback_pricing(instance_type)
            
            monthly_instance_cost = {
                'on_demand': pricing['on_demand'] * 730,
                'ri_1y_no_upfront': pricing['ri_1y_no_upfront'] * 730,
                'ri_3y_no_upfront': pricing['ri_3y_no_upfront'] * 730,
                'spot': pricing['spot'] * 730
            }
            
            storage_cost_per_gb = 0.08
            monthly_storage_cost = storage_gb * storage_cost_per_gb
            monthly_network_cost = 50
            
            total_costs = {}
            for pricing_model, instance_cost in monthly_instance_cost.items():
                total_costs[pricing_model] = instance_cost + monthly_storage_cost + monthly_network_cost
            
            return {
                "total_costs": total_costs,
                "instance_costs": monthly_instance_cost,
                "storage_costs": {"primary_storage": monthly_storage_cost},
                "network_costs": {"data_transfer": monthly_network_cost},
                "selected_instance": selected_instance
            }
        except Exception as e:
            logger.error(f"Error calculating basic costs: {e}")
            return {
                "total_costs": {"on_demand": 1000, "ri_1y_no_upfront": 700},
                "selected_instance": {'type': 'm6i.large', 'vCPU': 2, 'RAM': 8}
            }

    def _select_best_instance(self, required_vcpus: int, required_ram_gb: int) -> Dict[str, Any]:
        """Select the best matching instance type."""
        try:
            best_instance = None
            best_score = 0
            
            for instance in self.INSTANCE_TYPES:
                if instance['vCPU'] >= required_vcpus and instance['RAM'] >= required_ram_gb:
                    cpu_efficiency = required_vcpus / instance['vCPU']
                    ram_efficiency = required_ram_gb / instance['RAM']
                    overall_efficiency = (cpu_efficiency + ram_efficiency) / 2
                    
                    if overall_efficiency > best_score:
                        best_score = overall_efficiency
                        best_instance = instance.copy()
                        best_instance['efficiency_score'] = overall_efficiency
            
            if best_instance is None:
                best_instance = {
                    'type': 'm6i.large',
                    'vCPU': 2,
                    'RAM': 8,
                    'family': 'general',
                    'efficiency_score': 0.5
                }
            
            return best_instance
        except Exception as e:
            logger.error(f"Error selecting best instance: {e}")
            return {'type': 'm6i.large', 'vCPU': 2, 'RAM': 8, 'family': 'general'}

    def _calculate_tco(self, vcpus: int, ram_gb: int, env: str) -> Dict[str, Any]:
        """Calculate TCO analysis."""
        try:
            selected_instance = self._select_best_instance(vcpus, ram_gb)
            pricing = self._get_fallback_pricing(selected_instance['type'])
            
            on_demand_monthly = pricing['on_demand'] * 730
            ri_1y_monthly = pricing['ri_1y_no_upfront'] * 730
            ri_3y_monthly = pricing['ri_3y_no_upfront'] * 730
            
            storage_monthly = max(self.inputs.get('storage_current_gb', 500), 100) * 0.08
            network_monthly = 50
            
            total_on_demand = on_demand_monthly + storage_monthly + network_monthly
            total_ri_1y = ri_1y_monthly + storage_monthly + network_monthly
            total_ri_3y = ri_3y_monthly + storage_monthly + network_monthly
            
            costs = {
                'on_demand': total_on_demand,
                'ri_1y_no_upfront': total_ri_1y,
                'ri_3y_no_upfront': total_ri_3y
            }
            
            best_option = min(costs.keys(), key=lambda k: costs[k])
            best_cost = costs[best_option]
            savings = total_on_demand - best_cost
            
            return {
                "monthly_cost": best_cost,
                "monthly_savings": savings,
                "best_pricing_option": best_option,
                "roi_3_years": (savings * 36 / total_on_demand) * 100 if total_on_demand > 0 else 0
            }
        except Exception as e:
            logger.error(f"Error calculating TCO: {e}")
            return {"monthly_cost": 1000, "monthly_savings": 200, "best_pricing_option": "ri_1y_no_upfront"}

    def _get_fallback_pricing(self, instance_type: str) -> Dict[str, float]:
        """Fallback pricing data."""
        fallback_prices = {
            'm6i.large': {'on_demand': 0.0864, 'ri_1y_no_upfront': 0.0605, 'ri_3y_no_upfront': 0.0432, 'spot': 0.0259},
            'm6i.xlarge': {'on_demand': 0.1728, 'ri_1y_no_upfront': 0.1210, 'ri_3y_no_upfront': 0.0864, 'spot': 0.0518},
            'm6i.2xlarge': {'on_demand': 0.3456, 'ri_1y_no_upfront': 0.2419, 'ri_3y_no_upfront': 0.1728, 'spot': 0.1037},
            'm6i.4xlarge': {'on_demand': 0.6912, 'ri_1y_no_upfront': 0.4838, 'ri_3y_no_upfront': 0.3456, 'spot': 0.2074},
            'r6i.large': {'on_demand': 0.1008, 'ri_1y_no_upfront': 0.0706, 'ri_3y_no_upfront': 0.0504, 'spot': 0.0302},
            'r6i.xlarge': {'on_demand': 0.2016, 'ri_1y_no_upfront': 0.1411, 'ri_3y_no_upfront': 0.1008, 'spot': 0.0605}
        }
        return fallback_prices.get(instance_type, {'on_demand': 0.1, 'ri_1y_no_upfront': 0.07, 'ri_3y_no_upfront': 0.05, 'spot': 0.03})

    def _get_fallback_requirements(self, env: str) -> Dict[str, Any]:
        """Fallback requirements."""
        return {
            'requirements': {'vCPUs': 2, 'RAM_GB': 8, 'storage_GB': 100},
            'cost_breakdown': {'total_costs': {'on_demand': 500}},
            'tco_analysis': {'monthly_cost': 500, 'monthly_savings': 150},
            'claude_analysis': self.claude_analyzer._get_fallback_analysis(),
            'environment': env,
            'vrops_data': None
        }
    
class EnvironmentHeatMapGenerator:
    """Generate environment heat maps for workload analysis."""
    
    def __init__(self):
        self.environments = ['DEV', 'QA', 'UAT', 'PREPROD', 'PROD']
        self.metrics = ['Cost', 'Complexity', 'Risk', 'Timeline', 'Resources']

    def generate_heat_map_data(self, workload_results: Dict) -> pd.DataFrame:
        """Generate heat map data for environments."""
        try:
            heat_data = []
            
            for env in self.environments:
                env_results = workload_results.get(env, {})
                
                cost_score = self._calculate_cost_score(env_results)
                complexity_score = self._calculate_complexity_score(env_results)
                risk_score = self._calculate_risk_score(env_results)
                timeline_score = self._calculate_timeline_score(env_results)
                resource_score = self._calculate_resource_score(env_results)
                
                heat_data.append({
                    'Environment': env,
                    'Cost': cost_score,
                    'Complexity': complexity_score,
                    'Risk': risk_score,
                    'Timeline': timeline_score,
                    'Resources': resource_score
                })
            
            return pd.DataFrame(heat_data)
        except Exception as e:
            logger.error(f"Error generating heat map data: {e}")
            return pd.DataFrame()

    def _calculate_cost_score(self, env_results: Dict) -> float:
        """Calculate cost score for heat map."""
        try:
            if not env_results:
                return 50
            
            cost_breakdown = env_results.get('cost_breakdown', {})
            total_costs = cost_breakdown.get('total_costs', {})
            monthly_cost = total_costs.get('on_demand', 1000)
            
            if monthly_cost < 500: return 20
            elif monthly_cost < 1500: return 50
            elif monthly_cost < 3000: return 75
            else: return 95
        except Exception:
            return 50

    def _calculate_complexity_score(self, env_results: Dict) -> float:
        """Calculate complexity score for heat map."""
        try:
            claude_analysis = env_results.get('claude_analysis', {})
            return claude_analysis.get('complexity_score', 50)
        except Exception:
            return 50

    def _calculate_risk_score(self, env_results: Dict) -> float:
        """Calculate risk score for heat map."""
        try:
            claude_analysis = env_results.get('claude_analysis', {})
            complexity_score = claude_analysis.get('complexity_score', 50)
            return min(complexity_score * 1.2, 100)
        except Exception:
            return 50

    def _calculate_timeline_score(self, env_results: Dict) -> float:
        """Calculate timeline score for heat map."""
        try:
            claude_analysis = env_results.get('claude_analysis', {})
            timeline = claude_analysis.get('estimated_timeline', {})
            max_weeks = timeline.get('max_weeks', 8)
            
            if max_weeks < 4: return 20
            elif max_weeks < 8: return 40
            elif max_weeks < 16: return 70
            else: return 90
        except Exception:
            return 50

    def _calculate_resource_score(self, env_results: Dict) -> float:
        """Calculate resource score for heat map."""
        try:
            requirements = env_results.get('requirements', {})
            vcpus = requirements.get('vCPUs', 2)
            ram_gb = requirements.get('RAM_GB', 8)
            
            resource_intensity = (vcpus * 10) + (ram_gb * 2)
            
            if resource_intensity < 50: return 25
            elif resource_intensity < 150: return 50
            elif resource_intensity < 300: return 75
            else: return 95
        except Exception:
            return 50

    def create_heat_map_visualization(self, heat_data: pd.DataFrame) -> go.Figure:
        """Create Plotly heat map visualization."""
        try:
            if heat_data.empty:
                fig = go.Figure()
                fig.add_annotation(text="No data available for heat map", 
                                 xref="paper", yref="paper", x=0.5, y=0.5, showarrow=False)
                return fig
            
            environments = heat_data['Environment'].tolist()
            metrics = ['Cost', 'Complexity', 'Risk', 'Timeline', 'Resources']
            
            z_data = []
            for metric in metrics:
                if metric in heat_data.columns:
                    z_data.append(heat_data[metric].tolist())
                else:
                    z_data.append([50] * len(environments))
            
            fig = go.Figure(data=go.Heatmap(
                z=z_data,
                x=environments,
                y=metrics,
                colorscale='RdYlGn_r',
                zmin=0,
                zmax=100,
                colorbar=dict(
                    title="Impact Level",
                    tickvals=[0, 25, 50, 75, 100],
                    ticktext=["Very Low", "Low", "Medium", "High", "Very High"]
                ),
                text=[[f"{val:.0f}" for val in row] for row in z_data],
                texttemplate="%{text}",
                textfont={"size": 12},
                hoverongaps=False
            ))
            
            fig.update_layout(
                title="Environment Impact Heat Map",
                xaxis_title="Environment",
                yaxis_title="Impact Metrics",
                width=800,
                height=400
            )
            
            return fig
        except Exception as e:
            logger.error(f"Error creating heat map visualization: {e}")
            fig = go.Figure()
            fig.add_annotation(text=f"Error creating heat map: {str(e)}", 
                             xref="paper", yref="paper", x=0.5, y=0.5, showarrow=False)
            return fig    
    
class BulkWorkloadAnalyzer:
    """Handle bulk workload analysis from uploaded files."""
    
    def __init__(self):
        self.claude_analyzer = ClaudeAIMigrationAnalyzer()
        self.calculator = EnhancedEnterpriseEC2Calculator()
        
    def process_bulk_upload(self, uploaded_file, file_type: str) -> Dict[str, Any]:
        """Process bulk upload file and return analysis results."""
        try:
            if file_type == 'csv':
                return self._process_csv_file(uploaded_file)
            elif file_type == 'excel':
                return self._process_excel_file(uploaded_file)
            else:
                raise ValueError(f"Unsupported file type: {file_type}")
                
        except Exception as e:
            logger.error(f"Error processing bulk upload: {e}")
            return {'error': str(e), 'workloads': []}
    
    def _process_csv_file(self, uploaded_file) -> Dict[str, Any]:
        """Process CSV file."""
        try:
            # Read CSV content
            content = uploaded_file.read().decode('utf-8')
            csv_data = list(csv.DictReader(StringIO(content)))
            
            return self._analyze_workloads(csv_data)
            
        except Exception as e:
            return {'error': f"CSV processing error: {str(e)}", 'workloads': []}
    
    def _process_excel_file(self, uploaded_file) -> Dict[str, Any]:
        """Process Excel file."""
        try:
            if not OPENPYXL_AVAILABLE:
                return {'error': 'openpyxl not available for Excel processing', 'workloads': []}
                
            # Read Excel content
            df = pd.read_excel(uploaded_file, engine='openpyxl')
            workloads_data = df.to_dict('records')
            
            return self._analyze_workloads(workloads_data)
            
        except Exception as e:
            return {'error': f"Excel processing error: {str(e)}", 'workloads': []}
    
    def _analyze_workloads(self, workloads_data: List[Dict]) -> Dict[str, Any]:
        """Analyze multiple workloads."""
        results = {
            'total_workloads': len(workloads_data),
            'successful_analyses': 0,
            'failed_analyses': 0,
            'workloads': [],
            'summary': {}
        }
        
        for i, workload_data in enumerate(workloads_data):
            try:
                # Validate and normalize workload data
                normalized_workload = self._normalize_workload_data(workload_data)
                
                # Analyze for all environments
                workload_results = {}
                for env in ['DEV', 'QA', 'UAT', 'PREPROD', 'PROD']:
                    env_analysis = self._analyze_single_workload(normalized_workload, env)
                    workload_results[env] = env_analysis
                
                results['workloads'].append({
                    'index': i + 1,
                    'workload_name': normalized_workload.get('workload_name', f'Workload {i+1}'),
                    'status': 'success',
                    'analysis': workload_results
                })
                
                results['successful_analyses'] += 1
                
            except Exception as e:
                results['workloads'].append({
                    'index': i + 1,
                    'workload_name': workload_data.get('workload_name', f'Workload {i+1}'),
                    'status': 'failed',
                    'error': str(e)
                })
                results['failed_analyses'] += 1
        
        # Generate summary
        results['summary'] = self._generate_bulk_summary(results['workloads'])
        
        return results
    
    def _normalize_workload_data(self, workload_data: Dict) -> Dict:
        """Normalize and validate workload data."""
        # Define field mappings (CSV column -> internal field)
        field_mappings = {
            'workload_name': 'workload_name',
            'name': 'workload_name',
            'application_name': 'workload_name',
            'workload_type': 'workload_type',
            'type': 'workload_type',
            'application_type': 'workload_type',
            'operating_system': 'operating_system',
            'os': 'operating_system',
            'cpu_cores': 'on_prem_cores',
            'cores': 'on_prem_cores',
            'on_prem_cores': 'on_prem_cores',
            'peak_cpu_percent': 'peak_cpu_percent',
            'peak_cpu': 'peak_cpu_percent',
            'cpu_utilization': 'peak_cpu_percent',
            'ram_gb': 'on_prem_ram_gb',
            'memory_gb': 'on_prem_ram_gb',
            'on_prem_ram_gb': 'on_prem_ram_gb',
            'peak_ram_percent': 'peak_ram_percent',
            'peak_ram': 'peak_ram_percent',
            'memory_utilization': 'peak_ram_percent',
            'storage_gb': 'storage_current_gb',
            'storage_current_gb': 'storage_current_gb',
            'disk_gb': 'storage_current_gb',
            'peak_iops': 'peak_iops',
            'iops': 'peak_iops',
            'peak_throughput_mbps': 'peak_throughput_mbps',
            'throughput_mbps': 'peak_throughput_mbps',
            'infrastructure_age_years': 'infrastructure_age_years',
            'age_years': 'infrastructure_age_years',
            'business_criticality': 'business_criticality',
            'criticality': 'business_criticality',
            'region': 'region'
        }
        
        normalized = {}
        
        # Normalize field names (case-insensitive)
        for csv_field, value in workload_data.items():
            csv_field_lower = csv_field.lower().strip()
            
            if csv_field_lower in field_mappings:
                internal_field = field_mappings[csv_field_lower]
                normalized[internal_field] = value
        
        # Set defaults for missing fields
        defaults = {
            'workload_name': 'Unknown Workload',
            'workload_type': 'web_application',
            'operating_system': 'linux',
            'region': 'us-east-1',
            'on_prem_cores': 2,
            'peak_cpu_percent': 70,
            'on_prem_ram_gb': 8,
            'peak_ram_percent': 80,
            'storage_current_gb': 100,
            'peak_iops': 3000,
            'peak_throughput_mbps': 100,
            'infrastructure_age_years': 3,
            'business_criticality': 'medium'
        }
        
        for field, default_value in defaults.items():
            if field not in normalized or normalized[field] is None or normalized[field] == '':
                normalized[field] = default_value
            else:
                # Convert numeric fields
                if field in ['on_prem_cores', 'peak_cpu_percent', 'on_prem_ram_gb', 
                           'peak_ram_percent', 'storage_current_gb', 'peak_iops', 
                           'peak_throughput_mbps', 'infrastructure_age_years']:
                    try:
                        normalized[field] = float(normalized[field])
                    except (ValueError, TypeError):
                        normalized[field] = default_value
        
        return normalized
    
    def _analyze_single_workload(self, workload_inputs: Dict, environment: str) -> Dict[str, Any]:
        """Analyze a single workload for a specific environment."""
        
        # Update calculator inputs
        self.calculator.inputs.update(workload_inputs)
        
        # Calculate enhanced requirements
        return self.calculator.calculate_enhanced_requirements(environment)
    
    def _generate_bulk_summary(self, workloads: List[Dict]) -> Dict[str, Any]:
        """Generate summary statistics from bulk analysis."""
        successful_workloads = [w for w in workloads if w['status'] == 'success']
        
        if not successful_workloads:
            return {'error': 'No successful analyses to summarize'}
        
        # Aggregate statistics
        total_monthly_costs = []
        complexity_scores = []
        instance_types = {}
        
        for workload in successful_workloads:
            try:
                prod_analysis = workload['analysis']['PROD']
                
                # Cost data
                tco = prod_analysis.get('tco_analysis', {})
                monthly_cost = tco.get('monthly_cost', 0)
                if monthly_cost > 0:
                    total_monthly_costs.append(monthly_cost)
                
                # Complexity data
                claude_analysis = prod_analysis.get('claude_analysis', {})
                complexity = claude_analysis.get('complexity_score', 0)
                if complexity > 0:
                    complexity_scores.append(complexity)
                
                # Instance types
                cost_breakdown = prod_analysis.get('cost_breakdown', {})
                selected_instance = cost_breakdown.get('selected_instance', {})
                instance_type = selected_instance.get('type', 'Unknown')
                instance_types[instance_type] = instance_types.get(instance_type, 0) + 1
                
            except Exception as e:
                logger.warning(f"Error processing workload summary: {e}")
                continue
        
        # Calculate summary statistics
        summary = {
            'total_workloads_analyzed': len(successful_workloads),
            'total_monthly_cost': sum(total_monthly_costs),
            'total_annual_cost': sum(total_monthly_costs) * 12,
            'average_monthly_cost': sum(total_monthly_costs) / len(total_monthly_costs) if total_monthly_costs else 0,
            'average_complexity_score': sum(complexity_scores) / len(complexity_scores) if complexity_scores else 0,
            'most_common_instance_type': max(instance_types.items(), key=lambda x: x[1])[0] if instance_types else 'N/A',
            'instance_type_distribution': instance_types,
            'cost_range': {
                'min': min(total_monthly_costs) if total_monthly_costs else 0,
                'max': max(total_monthly_costs) if total_monthly_costs else 0
            },
            'complexity_range': {
                'min': min(complexity_scores) if complexity_scores else 0,
                'max': max(complexity_scores) if complexity_scores else 0
            }
        }
        
        return summary

# Enhanced Streamlit Functions with vROPS Integration
def initialize_enhanced_session_state():
    """Initialize enhanced session state with vROPS connector."""
    try:
        if 'enhanced_calculator' not in st.session_state:
            st.session_state.enhanced_calculator = EnhancedEnterpriseEC2Calculator()
        if 'enhanced_results' not in st.session_state:
            st.session_state.enhanced_results = None
        if 'bulk_results' not in st.session_state:
            st.session_state.bulk_results = None
        if 'vrops_connector' not in st.session_state:
            st.session_state.vrops_connector = VROPSConnector()
        if 'vrops_connection_status' not in st.session_state:
            st.session_state.vrops_connection_status = {'status': 'disconnected'}
        if 'vrops_vms' not in st.session_state:
            st.session_state.vrops_vms = []
        if 'selected_vm_metrics' not in st.session_state:
            st.session_state.selected_vm_metrics = None
            
        logger.info("Session state initialized successfully")
        
    except Exception as e:
        st.error(f"Error initializing session state: {str(e)}")
        logger.error(f"Error initializing session state: {e}")
        st.session_state.enhanced_calculator = None
        st.session_state.enhanced_results = None

def render_vrops_connection_tab():
    """Render vROPS connection and VM selection tab."""
    
    st.markdown("### 🔗 vRealize Operations Connection")
    
    # Connection Status
    connection_status = st.session_state.vrops_connection_status
    
    if connection_status['status'] == 'connected':
        st.markdown(f"""
        <div class="vrops-status-card vrops-connected">
            <h4 style="margin: 0; color: #065f46;">🟢 Connected to vROPS</h4>
            <p style="margin: 0.5rem 0 0 0; color: #047857;">
                <strong>Status:</strong> {connection_status.get('message', 'Successfully connected')}<br>
                <strong>Version:</strong> {connection_status.get('version', 'Unknown')}
            </p>
        </div>
        """, unsafe_allow_html=True)
        
        if st.button("🔌 Disconnect from vROPS", key="vrops_disconnect"):
            st.session_state.vrops_connector.disconnect()
            st.session_state.vrops_connection_status = {'status': 'disconnected'}
            st.session_state.vrops_vms = []
            st.session_state.selected_vm_metrics = None
            st.rerun()
    
    elif connection_status['status'] == 'error':
        st.markdown(f"""
        <div class="vrops-status-card vrops-error">
            <h4 style="margin: 0; color: #dc2626;">🔴 vROPS Connection Error</h4>
            <p style="margin: 0.5rem 0 0 0; color: #b91c1c;">
                <strong>Error:</strong> {connection_status.get('message', 'Connection failed')}
            </p>
        </div>
        """, unsafe_allow_html=True)
    
    else:
        st.markdown(f"""
        <div class="vrops-status-card vrops-disconnected">
            <h4 style="margin: 0; color: #d97706;">🟡 Not Connected to vROPS</h4>
            <p style="margin: 0.5rem 0 0 0; color: #92400e;">
                Connect to vRealize Operations to import real performance metrics for accurate AWS sizing recommendations.
            </p>
        </div>
        """, unsafe_allow_html=True)
    
    # Connection Form
    if connection_status['status'] != 'connected':
        st.markdown("#### vROPS Connection Settings")
        
        with st.form("vrops_connection_form"):
            col1, col2 = st.columns(2)
            
            with col1:
                vrops_hostname = st.text_input(
                    "vROPS Hostname/IP",
                    placeholder="vrops.company.com",
                    help="Enter the hostname or IP address of your vROPS instance"
                )
                
                vrops_username = st.text_input(
                    "Username",
                    placeholder="admin@local",
                    help="vROPS username with read access to VMs"
                )
            
            with col2:
                vrops_password = st.text_input(
                    "Password",
                    type="password",
                    help="vROPS user password"
                )
                
                verify_ssl = st.checkbox(
                    "Verify SSL Certificate",
                    value=False,
                    help="Uncheck for self-signed certificates"
                )
            
            connect_submitted = st.form_submit_button("🔗 Connect to vROPS", type="primary")
            
            if connect_submitted:
                if vrops_hostname and vrops_username and vrops_password:
                    with st.spinner("🔄 Connecting to vROPS..."):
                        connection_result = st.session_state.vrops_connector.connect(
                            vrops_hostname, vrops_username, vrops_password, verify_ssl
                        )
                        st.session_state.vrops_connection_status = connection_result
                        
                        if connection_result['status'] == 'connected':
                            st.success("✅ Successfully connected to vROPS!")
                            st.rerun()
                        else:
                            st.error(f"❌ Connection failed: {connection_result['message']}")
                else:
                    st.error("Please fill in all connection fields")
    
    # VM Selection (only if connected)
    if connection_status['status'] == 'connected':
        st.markdown("---")
        st.markdown("#### 🖥️ Virtual Machine Selection")
        
        # VM Search and List
        col1, col2 = st.columns([3, 1])
        
        with col1:
            vm_filter = st.text_input(
                "Filter VMs by name",
                placeholder="Enter VM name or partial name",
                help="Leave empty to see all VMs"
            )
        
        with col2:
            if st.button("🔍 Search VMs", key="search_vms"):
                with st.spinner("🔄 Retrieving VMs from vROPS..."):
                    vm_result = st.session_state.vrops_connector.get_virtual_machines(vm_filter)
                    
                    if vm_result['status'] == 'success':
                        st.session_state.vrops_vms = vm_result['vms']
                        st.success(f"✅ Found {vm_result['total_count']} VMs")
                    else:
                        st.error(f"❌ Error retrieving VMs: {vm_result['message']}")
                        st.session_state.vrops_vms = []
        
        # Display VM List
        if st.session_state.vrops_vms:
            st.markdown(f"**Found {len(st.session_state.vrops_vms)} Virtual Machines:**")
            
            # Create VM selection
            vm_options = {}
            for vm in st.session_state.vrops_vms:
                vm_name = vm['name']
                vm_status = vm.get('resourceStatus', 'Unknown')
                display_name = f"{vm_name} ({vm_status})"
                vm_options[display_name] = vm
            
            selected_vm_display = st.selectbox(
                "Select VM for analysis:",
                options=list(vm_options.keys()),
                help="Choose a VM to import performance metrics"
            )
            
            if selected_vm_display:
                selected_vm = vm_options[selected_vm_display]
                
                col1, col2 = st.columns([2, 1])
                
                with col1:
                    metrics_period = st.slider(
                        "Performance data collection period (days)",
                        min_value=7,
                        max_value=90,
                        value=30,
                        help="Number of days of historical performance data to analyze"
                    )
                
                with col2:
                    if st.button("📊 Import VM Metrics", type="primary", key="import_vm_metrics"):
                        import_vm_metrics(selected_vm, metrics_period)
        
        # Display imported metrics summary
        if st.session_state.selected_vm_metrics:
            render_vm_metrics_summary()

def import_vm_metrics(selected_vm: Dict, metrics_period: int):
    """Import VM metrics from vROPS."""
    
    with st.spinner(f"🔄 Importing performance metrics for {selected_vm['name']}..."):
        try:
            # Get VM metrics
            vm_metrics = st.session_state.vrops_connector.get_vm_metrics(
                selected_vm['id'], 
                metrics_period
            )
            
            # Get VM properties
            vm_properties = st.session_state.vrops_connector.get_vm_properties(
                selected_vm['id']
            )
            
            if vm_metrics['status'] == 'success' and vm_properties['status'] == 'success':
                # Process metrics for migration analysis
                vrops_processor = VROPSMetricsProcessor()
                processed_metrics = vrops_processor.process_vm_metrics_for_migration(
                    vm_metrics, 
                    vm_properties['configuration']
                )
                
                # Store in session state
                st.session_state.selected_vm_metrics = {
                    'vm_info': selected_vm,
                    'vm_config': vm_properties['configuration'],
                    'performance_data': vm_metrics,
                    'processed_metrics': processed_metrics,
                    'collection_period': metrics_period
                }
                
                # Update workload inputs with vROPS data
                calculator = st.session_state.enhanced_calculator
                if processed_metrics.get('status') == 'success':
                    aws_sizing = processed_metrics.get('aws_sizing', {})
                    
                    # Update workload name
                    calculator.inputs['workload_name'] = selected_vm['name']
                    
                    # Update sizing based on vROPS recommendations
                    if aws_sizing.get('recommended_vcpus'):
                        calculator.inputs['on_prem_cores'] = aws_sizing['recommended_vcpus']
                    
                    if aws_sizing.get('recommended_memory_gb'):
                        calculator.inputs['on_prem_ram_gb'] = aws_sizing['recommended_memory_gb']
                    
                    if aws_sizing.get('recommended_storage_gb'):
                        calculator.inputs['storage_current_gb'] = aws_sizing['recommended_storage_gb']
                    
                    if aws_sizing.get('estimated_iops'):
                        calculator.inputs['peak_iops'] = aws_sizing['estimated_iops']
                    
                    # Update utilization from actual performance data
                    perf_summary = processed_metrics.get('performance_summary', {})
                    if perf_summary.get('cpu_efficiency'):
                        calculator.inputs['peak_cpu_percent'] = perf_summary['cpu_efficiency']
                    
                    if perf_summary.get('memory_efficiency'):
                        calculator.inputs['peak_ram_percent'] = perf_summary['memory_efficiency']
                
                st.success(f"✅ Successfully imported metrics for {selected_vm['name']}!")
                st.info("💡 Workload configuration has been updated with vROPS performance data. Go to the Configuration tab to review and run analysis.")
                
            else:
                error_msg = vm_metrics.get('message', '') or vm_properties.get('message', '')
                st.error(f"❌ Error importing metrics: {error_msg}")
                
        except Exception as e:
            st.error(f"❌ Error importing VM metrics: {str(e)}")
            logger.error(f"Error importing VM metrics: {e}")

def render_vm_metrics_summary():
    """Render summary of imported VM metrics."""
    
    vm_metrics = st.session_state.selected_vm_metrics
    
    if not vm_metrics:
        return
    
    st.markdown("---")
    st.markdown("#### 📊 Imported vROPS Performance Data")
    
    vm_info = vm_metrics['vm_info']
    vm_config = vm_metrics['vm_config']
    processed_metrics = vm_metrics['processed_metrics']
    
    # VM Information
    st.markdown(f"**VM:** {vm_info['name']}")
    
    col1, col2, col3, col4 = st.columns(4)
    
    with col1:
        st.metric(
            "CPU Cores",
            f"{vm_config.get('cpu_cores', 'N/A')}",
            help="Current VM CPU allocation"
        )
    
    with col2:
        st.metric(
            "Memory (GB)",
            f"{vm_config.get('memory_mb', 0) / 1024:.1f}" if vm_config.get('memory_mb') else "N/A",
            help="Current VM memory allocation"
        )
    
    with col3:
        st.metric(
            "Storage (GB)",
            f"{vm_config.get('disk_gb', 'N/A')}",
            help="Current VM storage allocation"
        )
    
    with col4:
        perf_summary = processed_metrics.get('performance_summary', {})
        overall_score = perf_summary.get('overall_score', 0)
        st.metric(
            "Performance Score",
            f"{overall_score:.0f}/100",
            help="Overall performance score from vROPS data"
        )
    
    # Performance Summary
    if processed_metrics.get('status') == 'success':
        st.markdown("**Performance Insights:**")
        
        perf_summary = processed_metrics.get('performance_summary', {})
        migration_insights = processed_metrics.get('migration_insights', {})
        
        col1, col2 = st.columns(2)
        
        with col1:
            st.markdown("**Resource Utilization:**")
            st.markdown(f"• **CPU Efficiency:** {perf_summary.get('cpu_efficiency', 0):.1f}%")
            st.markdown(f"• **Memory Efficiency:** {perf_summary.get('memory_efficiency', 0):.1f}%")
            st.markdown(f"• **Storage Performance:** {perf_summary.get('storage_performance', 0):.1f}ms latency")
            st.markdown(f"• **Network Utilization:** {perf_summary.get('network_utilization', 0):.1f} Kbps")
        
        with col2:
            st.markdown("**Migration Assessment:**")
            st.markdown(f"• **Readiness:** {migration_insights.get('migration_readiness', 'Unknown')}")
            st.markdown(f"• **Recommended Approach:** {migration_insights.get('recommended_migration_approach', 'Unknown')}")
            
            optimization_count = len(migration_insights.get('optimization_opportunities', []))
            risk_count = len(migration_insights.get('risk_factors', []))
            st.markdown(f"• **Optimization Opportunities:** {optimization_count}")
            st.markdown(f"• **Risk Factors:** {risk_count}")
        
        # AWS Sizing Recommendations
        aws_sizing = processed_metrics.get('aws_sizing', {})
        if aws_sizing:
            st.markdown("**vROPS-based AWS Sizing Recommendations:**")
            
            sizing_col1, sizing_col2, sizing_col3, sizing_col4 = st.columns(4)
            
            with sizing_col1:
                st.metric(
                    "Recommended vCPUs",
                    aws_sizing.get('recommended_vcpus', 'N/A')
                )
            
            with sizing_col2:
                st.metric(
                    "Recommended Memory (GB)",
                    aws_sizing.get('recommended_memory_gb', 'N/A')
                )
            
            with sizing_col3:
                st.metric(
                    "Recommended Storage (GB)",
                    aws_sizing.get('recommended_storage_gb', 'N/A')
                )
            
            with sizing_col4:
                st.metric(
                    "Estimated IOPS",
                    aws_sizing.get('estimated_iops', 'N/A')
                )
    
    # Clear imported data button
    if st.button("🗑️ Clear Imported Data", key="clear_vrops_data"):
        st.session_state.selected_vm_metrics = None
        # Reset calculator inputs to defaults
        st.session_state.enhanced_calculator.inputs = {
            "workload_name": "Sample Enterprise Workload",
            "workload_type": "web_application",
            "operating_system": "linux",
            "region": "us-east-1",
            "on_prem_cores": 8,
            "peak_cpu_percent": 70,
            "on_prem_ram_gb": 32,
            "peak_ram_percent": 80,
            "storage_current_gb": 500,
            "storage_growth_rate": 0.15,
            "peak_iops": 5000,
            "peak_throughput_mbps": 250,
            "infrastructure_age_years": 3,
            "business_criticality": "medium"
        }
        st.success("✅ Imported vROPS data cleared. Configuration reset to defaults.")
        st.rerun()

def render_enhanced_configuration():
    """Render enhanced configuration with change detection and auto-refresh."""
    
    st.markdown("### ⚙️ Enhanced Enterprise Workload Configuration")
    
    # Check if calculator exists
    if 'enhanced_calculator' not in st.session_state or st.session_state.enhanced_calculator is None:
        st.error("⚠️ Calculator not initialized. Please refresh the page.")
        return
        
    calculator = st.session_state.enhanced_calculator
    
    # Store original inputs to detect changes
    if 'original_inputs' not in st.session_state:
        st.session_state.original_inputs = calculator.inputs.copy()
    
    # vROPS Data Status
    if st.session_state.selected_vm_metrics:
        vm_info = st.session_state.selected_vm_metrics['vm_info']
        st.markdown(f"""
        <div class="vrops-status-card vrops-connected">
            <h4 style="margin: 0; color: #065f46;">📊 vROPS Data Imported</h4>
            <p style="margin: 0.5rem 0 0 0; color: #047857;">
                <strong>VM:</strong> {vm_info['name']}<br>
                <strong>Status:</strong> Configuration updated with real performance metrics
            </p>
        </div>
        """, unsafe_allow_html=True)
    else:
        st.info("💡 Connect to vROPS in the 'vROPS Connection' tab to import real performance metrics for more accurate sizing.")
    
    # Detect if configuration has changed
    config_changed = st.session_state.original_inputs != calculator.inputs
    
    if config_changed and st.session_state.enhanced_results:
        st.warning("⚠️ Configuration has changed since last analysis. Click 'Run Enhanced Analysis' to see updated results.")
        if st.button("🔄 Clear Outdated Results", key="clear_outdated"):
            st.session_state.enhanced_results = None
            st.rerun()
    
    # Basic workload information
    with st.expander("📋 Workload Information", expanded=True):
        col1, col2 = st.columns(2)
        
        with col1:
            calculator.inputs["workload_name"] = st.text_input(
                "Workload Name",
                value=calculator.inputs["workload_name"],
                help="Descriptive name for this workload",
                key="workload_name_input"
            )
            
            workload_types = {
                'web_application': 'Web Application (Frontend, CDN)',
                'application_server': 'Application Server (APIs, Middleware)',
                'database_server': 'Database Server (RDBMS, NoSQL)',
                'file_server': 'File Server (Storage, Backup)',
                'compute_intensive': 'Compute Intensive (HPC, Analytics)',
                'analytics_workload': 'Analytics Workload (BI, Data Processing)'
            }
            
            calculator.inputs["workload_type"] = st.selectbox(
                "Workload Type",
                list(workload_types.keys()),
                index=list(workload_types.keys()).index(calculator.inputs["workload_type"]),
                format_func=lambda x: workload_types[x],
                help="Select the primary workload pattern",
                key="workload_type_input"
            )
        
        with col2:
            calculator.inputs["region"] = st.selectbox(
                "Primary AWS Region",
                ["us-east-1", "us-west-2", "eu-west-1", "ap-southeast-1"],
                index=["us-east-1", "us-west-2", "eu-west-1", "ap-southeast-1"].index(calculator.inputs["region"]),
                help="Primary AWS region for deployment",
                key="region_input"
            )
            
            calculator.inputs["operating_system"] = st.selectbox(
                "Operating System",
                ["linux", "windows"],
                index=["linux", "windows"].index(calculator.inputs["operating_system"]),
                format_func=lambda x: "Linux (Amazon Linux, Ubuntu, RHEL)" if x == "linux" else "Windows Server",
                key="os_input"
            )
    
    # Infrastructure metrics with change detection
    with st.expander("🖥️ Current Infrastructure Metrics", expanded=True):
        col1, col2, col3 = st.columns(3)
        
        with col1:
            st.markdown("**Compute Resources**")
            calculator.inputs["on_prem_cores"] = st.number_input(
                "CPU Cores", 
                min_value=1, 
                max_value=128, 
                value=int(calculator.inputs["on_prem_cores"]),
                key="cpu_cores_input"
            )
            calculator.inputs["peak_cpu_percent"] = st.slider(
                "Peak CPU %", 
                0, 
                100, 
                int(calculator.inputs["peak_cpu_percent"]),
                key="peak_cpu_input"
            )
        
        with col2:
            st.markdown("**Memory Resources**")
            calculator.inputs["on_prem_ram_gb"] = st.number_input(
                "RAM (GB)", 
                min_value=1, 
                max_value=1024, 
                value=int(calculator.inputs["on_prem_ram_gb"]),
                key="ram_gb_input"
            )
            calculator.inputs["peak_ram_percent"] = st.slider(
                "Peak RAM %", 
                0, 
                100, 
                int(calculator.inputs["peak_ram_percent"]),
                key="peak_ram_input"
            )
        
        with col3:
            st.markdown("**Storage & I/O**")
            calculator.inputs["storage_current_gb"] = st.number_input(
                "Storage (GB)", 
                min_value=1, 
                value=int(calculator.inputs["storage_current_gb"]),
                key="storage_gb_input"
            )
            calculator.inputs["peak_iops"] = st.number_input(
                "Peak IOPS", 
                min_value=1, 
                value=int(calculator.inputs["peak_iops"]),
                key="peak_iops_input"
            )
    
    # Business Context
    with st.expander("🏢 Business Context", expanded=False):
        col1, col2 = st.columns(2)
        
        with col1:
            calculator.inputs["business_criticality"] = st.selectbox(
                "Business Criticality",
                ["low", "medium", "high", "critical"],
                index=["low", "medium", "high", "critical"].index(calculator.inputs["business_criticality"]),
                help="Business impact level of this workload",
                key="criticality_input"
            )
        
        with col2:
            calculator.inputs["infrastructure_age_years"] = st.number_input(
                "Infrastructure Age (Years)",
                min_value=0,
                max_value=15,
                value=int(calculator.inputs["infrastructure_age_years"]),
                help="Age of current infrastructure",
                key="infra_age_input"
            )
    
    # Check for changes after all inputs
    current_inputs = calculator.inputs.copy()
    inputs_changed = st.session_state.original_inputs != current_inputs
    
    # Analysis buttons
    st.markdown("---")
    
    # Auto-refresh option
    col1, col2, col3 = st.columns([2, 1, 1])
    
    with col1:
        if st.button("🚀 Run Enhanced Analysis", type="primary", key="main_enhanced_analysis_button"):
            # Update original inputs to current state
            st.session_state.original_inputs = current_inputs.copy()
            run_enhanced_analysis()
    
    with col2:
        auto_refresh = st.checkbox("Auto-refresh", value=False, help="Automatically refresh when configuration changes")
    
    with col3:
        if inputs_changed and st.button("🔄 Reset Config", help="Reset to last analyzed configuration"):
            calculator.inputs.update(st.session_state.original_inputs)
            st.rerun()
    
    # Auto-refresh logic
    if auto_refresh and inputs_changed:
        st.session_state.original_inputs = current_inputs.copy()
        with st.spinner("🔄 Auto-refreshing analysis..."):
            run_enhanced_analysis()
        st.rerun()
    
    # Status indicators
    if inputs_changed:
        st.info("💡 Configuration has changed. Click 'Run Enhanced Analysis' to see updated results.")
    
    # Success message with navigation hint
    if st.session_state.enhanced_results and not inputs_changed:
        st.success("✅ Analysis completed! Results are up-to-date with current configuration.")
        st.info("💡 Visit the 'Results', 'Heat Map', and 'Technical Reports' tabs for detailed analysis.")
    elif st.session_state.enhanced_results and inputs_changed:
        st.warning("⚠️ Results shown are based on previous configuration. Re-run analysis to see current results.")
        
def run_enhanced_analysis():
    """Run enhanced analysis with improved error handling and status updates."""
    
    with st.spinner("🔄 Running enhanced analysis with Claude AI and vROPS data..."):
        try:
            calculator = st.session_state.enhanced_calculator
            
            if calculator is None:
                st.error("Calculator not available. Please refresh the page.")
                return
            
            # Get vROPS data if available
            vrops_data = None
            if st.session_state.selected_vm_metrics:
                vrops_data = st.session_state.selected_vm_metrics['processed_metrics']
            
            # Show progress
            progress_bar = st.progress(0)
            status_text = st.empty()
            
            # Calculate for all environments
            results = {}
            environments = list(calculator.ENV_MULTIPLIERS.keys())
            
            for i, env in enumerate(environments):
                status_text.text(f"Analyzing {env} environment...")
                progress_bar.progress((i + 1) / len(environments))
                
                results[env] = calculator.calculate_enhanced_requirements(env, vrops_data)
            
            status_text.text("Generating heat map and finalizing results...")
            
            # Generate heat map data
            heat_map_generator = EnvironmentHeatMapGenerator()
            heat_map_data = heat_map_generator.generate_heat_map_data(results)
            heat_map_fig = heat_map_generator.create_heat_map_visualization(heat_map_data)
            
            # Store results
            st.session_state.enhanced_results = {
                'inputs': calculator.inputs.copy(),
                'recommendations': results,
                'heat_map_data': heat_map_data,
                'heat_map_fig': heat_map_fig,
                'vrops_enhanced': vrops_data is not None,
                'timestamp': datetime.now()
            }
            
            # Clear progress indicators
            progress_bar.empty()
            status_text.empty()
            
            success_message = "✅ Enhanced analysis completed successfully!"
            if vrops_data:
                success_message += " 📊 Analysis enhanced with real vROPS performance data!"
            
            st.success(success_message)
            
            # Show quick summary
            prod_results = results.get('PROD', {})
            claude_analysis = prod_results.get('claude_analysis', {})
            tco_analysis = prod_results.get('tco_analysis', {})
            
            col1, col2, col3 = st.columns(3)
            with col1:
                complexity = claude_analysis.get('complexity_score', 50)
                st.metric("Complexity", f"{complexity:.0f}/100")
            with col2:
                cost = tco_analysis.get('monthly_cost', 0)
                st.metric("Monthly Cost", f"${cost:,.0f}")
            with col3:
                timeline = claude_analysis.get('estimated_timeline', {}).get('max_weeks', 8)
                st.metric("Timeline", f"{timeline} weeks")
            
        except Exception as e:
            st.error(f"❌ Error during enhanced analysis: {str(e)}")
            logger.error(f"Error in enhanced analysis: {e}")
            # Show detailed error for debugging
            st.error(f"Debug info: {type(e).__name__}: {str(e)}")
            
            # Clear progress indicators on error
            if 'progress_bar' in locals():
                progress_bar.empty()
            if 'status_text' in locals():
                status_text.empty()
        
def show_pricing_source_indicator(pricing_data: Dict[str, Any]):
    """Show pricing source indicator in the main interface."""
    
    source = pricing_data.get('source', 'unknown')
    last_updated = pricing_data.get('last_updated', 'Unknown')
    
    if source == 'aws_api':
        st.success(f"💰 Real-time AWS pricing (Updated: {last_updated[:16]})")
    elif source == 'fallback':
        st.info(f"💰 Fallback pricing data (Generated: {last_updated[:16]})")
        st.caption("💡 Configure AWS credentials for real-time pricing")
    else:
        st.warning("💰 Unknown pricing source")

def run_enhanced_analysis():
    """Run enhanced analysis with optional vROPS data."""
    
    with st.spinner("🔄 Running enhanced analysis with Claude AI and vROPS data..."):
        try:
            calculator = st.session_state.enhanced_calculator
            
            if calculator is None:
                st.error("Calculator not available. Please refresh the page.")
                return
            
            # Get vROPS data if available
            vrops_data = None
            if st.session_state.selected_vm_metrics:
                vrops_data = st.session_state.selected_vm_metrics['processed_metrics']
            
            # Calculate for all environments
            results = {}
            for env in calculator.ENV_MULTIPLIERS.keys():
                results[env] = calculator.calculate_enhanced_requirements(env, vrops_data)
            
            # Generate heat map data
            heat_map_generator = EnvironmentHeatMapGenerator()
            heat_map_data = heat_map_generator.generate_heat_map_data(results)
            heat_map_fig = heat_map_generator.create_heat_map_visualization(heat_map_data)
            
            # Store results
            st.session_state.enhanced_results = {
                'inputs': calculator.inputs.copy(),
                'recommendations': results,
                'heat_map_data': heat_map_data,
                'heat_map_fig': heat_map_fig,
                'vrops_enhanced': vrops_data is not None
            }
            
            success_message = "✅ Enhanced analysis completed successfully!"
            if vrops_data:
                success_message += " 📊 Analysis enhanced with real vROPS performance data!"
            
            st.success(success_message)
            
        except Exception as e:
            st.error(f"❌ Error during enhanced analysis: {str(e)}")
            logger.error(f"Error in enhanced analysis: {e}")

def render_enhanced_results():
    """Render enhanced analysis results with vROPS insights."""
    
    if 'enhanced_results' not in st.session_state or st.session_state.enhanced_results is None:
        st.markdown("""
        <div style="text-align: center; padding: 3rem; background: #f8fafc; border-radius: 12px; border: 2px dashed #cbd5e1;">
            <h3 style="color: #64748b; margin-bottom: 1rem;">🔍 No Analysis Results</h3>
            <p style="color: #64748b; margin-bottom: 1.5rem;">Configure your workload in the "Configuration" tab and click "Run Enhanced Analysis" to see detailed results here.</p>
            <div style="font-size: 0.9rem; color: #94a3b8;">
                💡 <strong>Tip:</strong> Connect to vROPS first for enhanced accuracy with real performance data
            </div>
        </div>
        """, unsafe_allow_html=True)
        return
    
    try:
        results = st.session_state.enhanced_results
        #st.markdown("### 📊 Enhanced Analysis Results")
        
        # Check if configuration has changed since analysis
        calculator = st.session_state.enhanced_calculator
        current_inputs = calculator.inputs.copy() if calculator else {}
        analysis_inputs = results.get('inputs', {})
        config_changed = current_inputs != analysis_inputs
        
        # Staleness warning
        if config_changed:
            st.markdown("""
            <div style="background: #fef3c7; border: 2px solid #f59e0b; border-radius: 8px; padding: 1rem; margin-bottom: 1rem;">
                <h4 style="color: #92400e; margin: 0;">⚠️ Results May Be Outdated</h4>
                <p style="color: #92400e; margin: 0.5rem 0 0 0;">
                    Configuration has changed since this analysis was run. 
                    <strong>Re-run the analysis</strong> to see updated results.
                </p>
            </div>
            """, unsafe_allow_html=True)
            
            col1, col2 = st.columns([1, 4])
            with col1:
                if st.button("🔄 Re-run Analysis", type="primary", key="rerun_from_results"):
                    run_enhanced_analysis()
                    st.rerun()
        else:
            # Fresh results indicator
            analysis_time = results.get('timestamp', datetime.now())
            time_ago = datetime.now() - analysis_time
            if time_ago.total_seconds() < 60:
                time_str = "just now"
            elif time_ago.total_seconds() < 3600:
                time_str = f"{int(time_ago.total_seconds() / 60)} minutes ago"
            else:
                time_str = f"{int(time_ago.total_seconds() / 3600)} hours ago"
            
            st.markdown(f"""
            <div style="background: #dcfce7; border: 2px solid #16a34a; border-radius: 8px; padding: 0.75rem; margin-bottom: 1rem;">
                <div style="color: #166534; font-weight: 600;">✅ Results are up-to-date (analyzed {time_str})</div>
            </div>
            """, unsafe_allow_html=True)
        
        # vROPS Enhancement Indicator
        if results.get('vrops_enhanced'):
            st.markdown("""
            <div class="vrops-status-card vrops-connected">
                <h4 style="margin: 0; color: #065f46;">📊 vROPS Enhanced Analysis</h4>
                <p style="margin: 0.5rem 0 0 0; color: #047857;">
                    This analysis has been enhanced with real performance data from vRealize Operations for more accurate sizing and migration planning.
                </p>
            </div>
            """, unsafe_allow_html=True)
        
        recommendations = results.get('recommendations', {})
        if not recommendations or 'PROD' not in recommendations:
            st.warning("⚠️ Analysis results incomplete. Please run the analysis again.")
            return
        
        prod_results = recommendations['PROD']
        claude_analysis = prod_results.get('claude_analysis', {})
        tco_analysis = prod_results.get('tco_analysis', {})
        
        # Summary metrics
        col1, col2, col3, col4 = st.columns(4)
        
        with col1:
            complexity_score = claude_analysis.get('complexity_score', 50)
            complexity_level = claude_analysis.get('complexity_level', 'MEDIUM')
            # Color coding for complexity
            if complexity_score >= 80:
                complexity_color = "#dc2626"  # Red
            elif complexity_score >= 60:
                complexity_color = "#f59e0b"  # Orange
            elif complexity_score >= 40:
                complexity_color = "#3b82f6"  # Blue
            else:
                complexity_color = "#10b981"  # Green
            
            
            
            st.markdown(f"""
            <div class="metric-card">
                <div style="font-size: 0.875rem; font-weight: 600; color: #6b7280; margin-bottom: 0.5rem;">🤖 Migration Complexity</div>
                <div style="font-size: 2rem; font-weight: 700; color: #1f2937; margin-bottom: 0.25rem;">{complexity_score:.0f}/100</div>
                <div style="font-size: 0.75rem; color: #9ca3af;">{complexity_level}</div>
            </div>
            """, unsafe_allow_html=True)
        
        with col2:
            monthly_cost = tco_analysis.get('monthly_cost', 0)
            
            st.markdown(f"""
            <div class="metric-card">
                <div style="font-size: 0.875rem; font-weight: 600; color: #6b7280; margin-bottom: 0.5rem;">☁️ AWS Monthly Cost</div>
                <div style="font-size: 2rem; font-weight: 700; color: #1f2937; margin-bottom: 0.25rem;">${monthly_cost:,.0f}</div>
                <div style="font-size: 0.75rem; color: #9ca3af;">Optimized Pricing</div>
            </div>
            """, unsafe_allow_html=True)
        
        with col3:
            timeline = claude_analysis.get('estimated_timeline', {})
            max_weeks = timeline.get('max_weeks', 8)
            
            st.markdown(f"""
            <div class="metric-card">
                <div style="font-size: 0.875rem; font-weight: 600; color: #6b7280; margin-bottom: 0.5rem;">⏱️ Migration Timeline</div>
                <div style="font-size: 2rem; font-weight: 700; color: #1f2937; margin-bottom: 0.25rem;">{max_weeks}</div>
                <div style="font-size: 0.75rem; color: #9ca3af;">Weeks (Estimated)</div>
            </div>
            """, unsafe_allow_html=True)
        
        with col4:
            st.markdown(f"""
            <div class="metric-card">
                <div style="font-size: 0.875rem; font-weight: 600; color: #6b7280; margin-bottom: 0.5rem;">🖥️ Instance Type</div>
                <div style="font-size: 1.5rem; font-weight: 700; color: #1f2937; margin-bottom: 0.25rem;">
                    {prod_results.get('cost_breakdown', {}).get('selected_instance', {}).get('type', 'N/A')}
                </div>
                <div style="font-size: 0.75rem; color: #9ca3af;">Recommended</div>
            </div>
            """, unsafe_allow_html=True)
            
        # Show configuration that was analyzed
        if config_changed:
            with st.expander("⚙️ Configuration Used for This Analysis", expanded=False):
                config_display = analysis_inputs.copy()
                # Remove sensitive or irrelevant fields
                display_config = {
                    'Workload Name': config_display.get('workload_name', 'N/A'),
                    'Workload Type': config_display.get('workload_type', 'N/A'),
                    'Operating System': config_display.get('operating_system', 'N/A'),
                    'CPU Cores': config_display.get('on_prem_cores', 'N/A'),
                    'RAM (GB)': config_display.get('on_prem_ram_gb', 'N/A'),
                    'Storage (GB)': config_display.get('storage_current_gb', 'N/A'),
                    'Peak CPU %': config_display.get('peak_cpu_percent', 'N/A'),
                    'Peak RAM %': config_display.get('peak_ram_percent', 'N/A'),
                    'Business Criticality': config_display.get('business_criticality', 'N/A'),
                    'AWS Region': config_display.get('region', 'N/A')
                }
                
                config_df = pd.DataFrame(list(display_config.items()), columns=['Setting', 'Value'])
                st.dataframe(config_df, use_container_width=True, hide_index=True)    
        
        # Claude AI Analysis with vROPS insights
        st.markdown("### 🤖 Claude AI Migration Analysis")
        
        if claude_analysis:
            col1, col2 = st.columns(2)
            
            with col1:
                st.markdown("**Migration Strategy**")
                strategy = claude_analysis.get('migration_strategy', {})
                if strategy:
                    st.markdown(f"**Approach:** {strategy.get('approach', 'N/A')}")
                    st.markdown(f"**Methodology:** {strategy.get('methodology', 'N/A')}")
                    st.markdown(f"**Timeline:** {strategy.get('timeline', 'N/A')}")
                    st.markdown(f"**Risk Level:** {strategy.get('risk_level', 'N/A')}")
            
            with col2:
                st.markdown("**Migration Steps**")
                migration_steps = claude_analysis.get('migration_steps', [])
                
                for i, step in enumerate(migration_steps[:3], 1):
                    if isinstance(step, dict):
                        with st.expander(f"Phase {i}: {step.get('phase', 'N/A')}", expanded=False):
                            st.markdown(f"**Duration:** {step.get('duration', 'N/A')}")
                            
                            tasks = step.get('tasks', [])
                            if tasks:
                                st.markdown("**Key Tasks:**")
                                for task in tasks[:3]:
                                    st.markdown(f"• {task}")
        
        # vROPS Insights (if available)
        vrops_insights = claude_analysis.get('vrops_insights', {})
        if vrops_insights and results.get('vrops_enhanced'):
            st.markdown("### 📊 vRealize Operations Insights")
            
            col1, col2 = st.columns(2)
            
            with col1:
                st.markdown("**Performance Assessment**")
                st.markdown(f"• **Performance Impact:** {vrops_insights.get('performance_impact', 'N/A')}")
                st.markdown(f"• **Sizing Confidence:** {vrops_insights.get('sizing_confidence', 'N/A')}")
                st.markdown(f"• **Optimization Potential:** {vrops_insights.get('optimization_potential', 'N/A')}")
            
            with col2:
                st.markdown("**Migration Approach**")
                st.markdown(f"• **Recommended Approach:** {vrops_insights.get('recommended_approach', 'N/A')}")
                
                # Show vROPS data source
                if st.session_state.selected_vm_metrics:
                    vm_info = st.session_state.selected_vm_metrics['vm_info']
                    collection_period = st.session_state.selected_vm_metrics['collection_period']
                    st.markdown(f"• **Data Source:** {vm_info['name']} ({collection_period} days)")
        
        # Cost Analysis
        st.markdown("### 💰 Cost Analysis")
        
        cost_breakdown = prod_results.get('cost_breakdown', {})
        total_costs = cost_breakdown.get('total_costs', {})
        
        # ADD THIS: Show pricing source information
        selected_instance = cost_breakdown.get('selected_instance', {})
        if selected_instance:
            calculator = AWSCostCalculator()
            instance_pricing = calculator._get_ec2_pricing(selected_instance.get('type', 'm6i.large'))
            show_pricing_source_indicator(instance_pricing)
        
        if total_costs:
            col1, col2 = st.columns(2)
            
            with col1:
                st.markdown("**Instance Pricing Comparison**")
                cost_data = []
                for pricing_model, cost in total_costs.items():
                    cost_data.append({
                        'Pricing Model': pricing_model.replace('_', ' ').title(),
                        'Monthly Cost': f"${cost:,.2f}",
                        'Annual Cost': f"${cost*12:,.2f}"
                    })
                
                df_costs = pd.DataFrame(cost_data)
                st.dataframe(df_costs, use_container_width=True, hide_index=True)
            
            with col2:
                st.markdown("**AWS Service Cost Breakdown (PROD)**")
                
                # Calculate detailed service costs if available
                try:
                    analyzer = EnhancedEnvironmentAnalyzer()
                    tech_recs = analyzer.get_technical_recommendations('PROD', prod_results)
                    cost_calculator = AWSCostCalculator()
                    service_costs = cost_calculator.calculate_service_costs(
                        'PROD', tech_recs, prod_results.get('requirements', {}))
                    
                    service_cost_data = []
                    categories = ['compute', 'network', 'storage', 'database', 'security', 'monitoring']
                    for cat in categories:
                        if cat in service_costs:
                            service_cost_data.append({
                                'Service Category': cat.title(),
                                'Monthly Cost': f"${service_costs[cat]['total']:.2f}"
                            })
                    
                    if service_cost_data:
                        df_service_costs = pd.DataFrame(service_cost_data)
                        st.dataframe(df_service_costs, use_container_width=True, hide_index=True)
                        
                        # ADD THIS: Show if using real AWS pricing
                    if selected_instance:
                        source = instance_pricing.get('source', 'unknown')
                        if source == 'aws_api':
                            st.success("✅ Real-time AWS pricing")
                        else:
                            st.info("ℹ️ Using fallback pricing")                     
                                               
                        
                        # Show total from service breakdown
                        total_services = sum(service_costs[cat]['total'] for cat in categories if cat in service_costs)
                        st.markdown(f"**Total Monthly AWS Services Cost: ${total_services:.2f}**")
                    else:
                        # Fallback to basic cost display
                        basic_cost_data = [
                            {'Cost Component': 'Instance Costs', 'Monthly Cost': f"${cost_breakdown.get('instance_costs', {}).get('on_demand', 0):.2f}"},
                            {'Cost Component': 'Storage Costs', 'Monthly Cost': f"${cost_breakdown.get('storage_costs', {}).get('primary_storage', 0):.2f}"},
                            {'Cost Component': 'Network Costs', 'Monthly Cost': f"${cost_breakdown.get('network_costs', {}).get('data_transfer', 0):.2f}"}
                        ]
                        
                        df_basic_costs = pd.DataFrame(basic_cost_data)
                        st.dataframe(df_basic_costs, use_container_width=True, hide_index=True)
                
                except Exception as e:
                    logger.error(f"Error calculating detailed service costs: {e}")
                    # Fallback to basic cost display
                    basic_cost_data = [
                        {'Cost Component': 'Instance Costs', 'Monthly Cost': f"${cost_breakdown.get('instance_costs', {}).get('on_demand', 0):.2f}"},
                        {'Cost Component': 'Storage Costs', 'Monthly Cost': f"${cost_breakdown.get('storage_costs', {}).get('primary_storage', 0):.2f}"},
                        {'Cost Component': 'Network Costs', 'Monthly Cost': f"${cost_breakdown.get('network_costs', {}).get('data_transfer', 0):.2f}"}
                    ]
                    
                    df_basic_costs = pd.DataFrame(basic_cost_data)
                    st.dataframe(df_basic_costs, use_container_width=True, hide_index=True)
        
    except Exception as e:
        st.error(f"❌ Error displaying results: {str(e)}")
        logger.error(f"Error in render_enhanced_results: {e}")
        
        # Show debug information in development
        if st.checkbox("Show debug information", key="debug_results"):
            st.code(f"Error details: {type(e).__name__}: {str(e)}")
            if 'enhanced_results' in st.session_state:
                st.json(st.session_state.enhanced_results)

def render_enhanced_environment_heatmap_tab():
    """Render enhanced environment heat map tab with detailed explanations."""
    
    st.markdown("### 🌡️ Environment Impact Analysis with Detailed Explanations")
    
    if 'enhanced_results' not in st.session_state or not st.session_state.enhanced_results:
        st.info("💡 Run an enhanced analysis to see detailed environment heat maps and explanations.")
        return
    
    results = st.session_state.enhanced_results
    analyzer = EnhancedEnvironmentAnalyzer()
    
    # vROPS Enhancement Indicator
    if results.get('vrops_enhanced'):
        st.info("📊 This analysis includes real performance data from vRealize Operations for enhanced accuracy.")
    
    # Environment overview cards
    st.markdown("#### Environment Complexity Overview")
    
    cols = st.columns(5)
    environments = ['DEV', 'QA', 'UAT', 'PREPROD', 'PROD']
    
    for i, env in enumerate(environments):
        with cols[i]:
            env_results = results['recommendations'].get(env, {})
            claude_analysis = env_results.get('claude_analysis', {})
            complexity = claude_analysis.get('complexity_score', 50)
            complexity_level = claude_analysis.get('complexity_level', 'MEDIUM')
            
            # Get detailed explanation
            complexity_explanation = analyzer.get_detailed_complexity_explanation(env, env_results)
            
            # Create expandable card
            with st.expander(f"{env} - {complexity:.0f}/100 ({complexity_level})", expanded=False):
                
                st.markdown(f"**Overall Complexity Score:** {complexity:.0f}/100")
                st.markdown(f"**Complexity Level:** {complexity_level}")
                
                st.markdown("**Key Complexity Factors:**")
                factors = complexity_explanation['factors']
                
                # Display factor scores
                factor_data = []
                for factor_name, factor_details in factors.items():
                    if isinstance(factor_details, dict):
                        score = factor_details.get('score', 0)
                        factor_data.append({'Factor': factor_name, 'Score': f"{score:.0f}/100"})
                
                if factor_data:
                    df_factors = pd.DataFrame(factor_data)
                    st.dataframe(df_factors, use_container_width=True, hide_index=True)
                
                st.markdown("**Why This Environment is Complex:**")
                for reason in complexity_explanation['detailed_reasons']:
                    st.markdown(f"• {reason}")
                
                st.markdown("**Specific Challenges:**")
                for challenge in complexity_explanation['specific_challenges'][:3]:
                    st.markdown(f"• {challenge}")
                
                st.markdown("**Mitigation Strategies:**")
                for strategy in complexity_explanation['mitigation_strategies'][:3]:
                    st.markdown(f"• {strategy}")
    
    # Heat map visualization
    st.markdown("#### Impact Heat Map Visualization")
    
    if 'heat_map_fig' in results:
        st.plotly_chart(results['heat_map_fig'], use_container_width=True)
    
    # Detailed complexity breakdown table
    st.markdown("#### Detailed Complexity Breakdown by Environment")
    
    detailed_data = []
    
    for env in environments:
        env_results = results['recommendations'].get(env, {})
        complexity_explanation = analyzer.get_detailed_complexity_explanation(env, env_results)
        
        factors = complexity_explanation['factors']
        
        detailed_data.append({
            'Environment': env,
            'Overall Score': f"{complexity_explanation['overall_score']:.0f}/100",
            'Complexity Level': complexity_explanation['complexity_level'],
            'Resource Intensity': f"{factors['Resource Intensity']['score']:.0f}/100",
            'Migration Risk': f"{factors['Migration Risk']['score']:.0f}/100",
            'Operational Complexity': f"{factors['Operational Complexity']['score']:.0f}/100",
            'Primary Reason': complexity_explanation['detailed_reasons'][0] if complexity_explanation['detailed_reasons'] else 'N/A'
        })
    
    df_detailed = pd.DataFrame(detailed_data)
    st.dataframe(df_detailed, use_container_width=True, hide_index=True)

def render_technical_recommendations_tab():
    """Render comprehensive technical recommendations tab with cost details."""
    
    st.markdown("### 🔧 Comprehensive Technical Recommendations by Environment")
    
    if 'enhanced_results' not in st.session_state or not st.session_state.enhanced_results:
        st.info("💡 Run an enhanced analysis to see detailed technical recommendations.")
        return
    
    results = st.session_state.enhanced_results
    analyzer = EnhancedEnvironmentAnalyzer()
    cost_calculator = AWSCostCalculator()
    
    # vROPS Enhancement Indicator
    if results.get('vrops_enhanced'):
        st.info("📊 Technical recommendations enhanced with vRealize Operations performance data.")
    
    # Environment selector
    selected_env = st.selectbox(
        "Select Environment for Detailed Technical Recommendations:",
        ['PROD', 'PREPROD', 'UAT', 'QA', 'DEV'],
        help="Choose an environment to see comprehensive technical specifications and costs"
    )
    
    env_results = results['recommendations'].get(selected_env, {})
    
    if not env_results:
        st.warning(f"No analysis results available for {selected_env} environment.")
        return
    
    # Get technical recommendations and costs
    tech_recs = analyzer.get_technical_recommendations(selected_env, env_results)
    requirements = env_results.get('requirements', {})
    service_costs = cost_calculator.calculate_service_costs(selected_env, tech_recs, requirements)
    
    st.markdown(f"## {selected_env} Environment - Technical Specifications & Costs")
    
    # Cost summary at the top
    st.markdown("### 💰 Cost Summary")
    
    col1, col2, col3, col4 = st.columns(4)
    
    with col1:
        total_monthly = service_costs['summary']['total_monthly']
        st.metric("Total Monthly Cost", f"${total_monthly:,.2f}")
    
    with col2:
        total_annual = service_costs['summary']['total_annual']
        st.metric("Total Annual Cost", f"${total_annual:,.2f}")
    
    with col3:
        largest_category = service_costs['summary']['largest_cost_category']
        largest_cost = service_costs[largest_category]['total']
        st.metric("Largest Cost Driver", f"{largest_category.title()}: ${largest_cost:.2f}")
    
    with col4:
        # Calculate cost per vCPU
        vcpus = requirements.get('vCPUs', 2)
        cost_per_vcpu = total_monthly / vcpus if vcpus > 0 else 0
        st.metric("Cost per vCPU/month", f"${cost_per_vcpu:.2f}")
    
    # Cost breakdown chart
    st.markdown("#### Cost Breakdown by Service Category")
    
    categories = ['compute', 'network', 'storage', 'database', 'security', 'monitoring']
    costs = [service_costs[cat]['total'] for cat in categories]
    
    # Filter out zero costs to avoid clutter
    filtered_data = [(cat.title(), cost) for cat, cost in zip(categories, costs) if cost > 0]
    
    if filtered_data:
        labels, values = zip(*filtered_data)
        
        # Create a more visually appealing pie chart
        fig_pie = go.Figure(data=[go.Pie(
            labels=labels,
            values=values,
            hole=.4,  # Donut chart
            textinfo='label+percent+value',
            textposition='auto',
            textfont=dict(size=14),
            marker=dict(
                colors=['#3b82f6', '#ef4444', '#10b981', '#f59e0b', '#8b5cf6', '#06b6d4'],
                line=dict(color='#FFFFFF', width=2)
            ),
            hovertemplate='<b>%{label}</b><br>' +
                         'Cost: $%{value:.2f}<br>' +
                         'Percentage: %{percent}<br>' +
                         '<extra></extra>',
            pull=[0.05 if max(values) == val else 0 for val in values]  # Pull out the largest slice
        )])
        
        fig_pie.update_layout(
            title=dict(
                text="Monthly Cost Distribution",
                x=0.5,
                font=dict(size=18, color='#1f2937')
            ),
            height=500,
            width=700,
            showlegend=True,
            legend=dict(
                orientation="v",
                yanchor="middle",
                y=0.5,
                xanchor="left",
                x=1.05,
                font=dict(size=12)
            ),
            margin=dict(l=20, r=120, t=60, b=20),
            annotations=[
                dict(
                    text=f"Total<br>${sum(values):.2f}/month",
                    x=0.5, y=0.5,
                    font_size=16,
                    font_color='#1f2937',
                    showarrow=False
                )
            ]
        )
        
        st.plotly_chart(fig_pie, use_container_width=True)
    else:
        st.warning("No cost data available for visualization.")
    
    # Create tabs for different technical areas with costs
    tech_tabs = st.tabs(["Compute", "Network", "Storage", "Database", "Security", "Monitoring"])
    
    # Compute tab with costs
    with tech_tabs[0]:
        st.markdown("#### 💻 Compute Configuration & Costs")
        
        compute_recs = tech_recs['compute']
        compute_costs = service_costs['compute']
        
        # Cost overview
        st.markdown(f"**Monthly Compute Cost: ${compute_costs['total']:.2f}**")
        
        col1, col2 = st.columns(2)
        
        with col1:
            st.markdown("**Primary Instance Recommendation**")
            
            primary_instance = compute_recs['primary_instance']
            instance_data = [
                {'Specification': 'Instance Type', 'Value': primary_instance['type']},
                {'Specification': 'vCPUs', 'Value': str(primary_instance['vcpus'])},
                {'Specification': 'Memory (GB)', 'Value': str(primary_instance['memory_gb'])},
                {'Specification': 'Rationale', 'Value': primary_instance['rationale']}
            ]
            
            df_instance = pd.DataFrame(instance_data)
            st.dataframe(df_instance, use_container_width=True, hide_index=True)
        
        with col2:
            st.markdown("**Compute Cost Breakdown**")
            
            cost_breakdown_data = []
            for service, details in compute_costs.items():
                if service != 'total' and service != 'optimization_notes':
                    cost_breakdown_data.append({
                        'Service': service.replace('_', ' ').title(),
                        'Monthly Cost': f"${details['cost']:.2f}",
                        'Details': details['details']
                    })
            
            df_compute_costs = pd.DataFrame(cost_breakdown_data)
            st.dataframe(df_compute_costs, use_container_width=True, hide_index=True)
        
        # Deployment configuration
        st.markdown("**Deployment Configuration**")
        
        deployment_data = [
            {'Configuration': 'Placement Strategy', 'Recommendation': compute_recs['placement_strategy']},
            {'Configuration': 'Auto Scaling', 'Recommendation': compute_recs['auto_scaling']},
            {'Configuration': 'Pricing Optimization', 'Recommendation': compute_recs['pricing_optimization']}
        ]
        
        df_deployment = pd.DataFrame(deployment_data)
        st.dataframe(df_deployment, use_container_width=True, hide_index=True)
        
        # Cost optimization notes
        st.markdown("**💡 Cost Optimization Recommendations**")
        for note in compute_costs.get('optimization_notes', []):
            st.markdown(note)
    
    # Network tab with costs
    with tech_tabs[1]:
        st.markdown("#### 🌐 Network Configuration & Costs")
        
        network_recs = tech_recs['network']
        network_costs = service_costs['network']
        
        # Cost overview
        st.markdown(f"**Monthly Network Cost: ${network_costs['total']:.2f}**")
        
        col1, col2 = st.columns(2)
        
        with col1:
            st.markdown("**Core Network Components**")
            
            core_network_data = [
                {'Component': 'VPC Design', 'Configuration': network_recs['vpc_design']},
                {'Component': 'Subnets', 'Configuration': network_recs['subnets']},
                {'Component': 'Security Groups', 'Configuration': network_recs['security_groups']},
                {'Component': 'Load Balancer', 'Configuration': network_recs['load_balancer']}
            ]
            
            df_core_network = pd.DataFrame(core_network_data)
            st.dataframe(df_core_network, use_container_width=True, hide_index=True)
        
        with col2:
            st.markdown("**Network Cost Breakdown**")
            
            network_cost_data = []
            for service, details in network_costs.items():
                if service != 'total' and service != 'optimization_notes':
                    network_cost_data.append({
                        'Service': service.replace('_', ' ').title(),
                        'Monthly Cost': f"${details['cost']:.2f}",
                        'Details': details['details']
                    })
            
            df_network_costs = pd.DataFrame(network_cost_data)
            st.dataframe(df_network_costs, use_container_width=True, hide_index=True)
        
        # Advanced network services
        st.markdown("**Advanced Network Services**")
        
        advanced_network_data = [
            {'Service': 'CDN', 'Configuration': network_recs['cdn']},
            {'Service': 'DNS', 'Configuration': network_recs['dns']},
            {'Service': 'NAT Gateway', 'Configuration': network_recs['nat_gateway']},
            {'Service': 'VPN', 'Configuration': network_recs['vpn']}
        ]
        
        df_advanced_network = pd.DataFrame(advanced_network_data)
        st.dataframe(df_advanced_network, use_container_width=True, hide_index=True)
        
        # Cost optimization notes
        st.markdown("**💡 Network Cost Optimization**")
        for note in network_costs.get('optimization_notes', []):
            st.markdown(note)
    
    # Storage tab with costs
    with tech_tabs[2]:
        st.markdown("#### 💾 Storage Configuration & Costs")
        
        storage_recs = tech_recs['storage']
        storage_costs = service_costs['storage']
        
        # Cost overview
        st.markdown(f"**Monthly Storage Cost: ${storage_costs['total']:.2f}**")
        
        col1, col2 = st.columns(2)
        
        with col1:
            st.markdown("**Primary Storage Configuration**")
            
            storage_config_data = [
                {'Setting': 'Storage Type', 'Value': storage_recs['primary_storage']},
                {'Setting': 'Recommended Size', 'Value': storage_recs['recommended_size']},
                {'Setting': 'IOPS', 'Value': storage_recs['iops_recommendation']},
                {'Setting': 'Throughput', 'Value': storage_recs['throughput_recommendation']}
            ]
            
            df_storage_config = pd.DataFrame(storage_config_data)
            st.dataframe(df_storage_config, use_container_width=True, hide_index=True)
        
        with col2:
            st.markdown("**Storage Cost Breakdown**")
            
            storage_cost_data = []
            for service, details in storage_costs.items():
                if service != 'total' and service != 'optimization_notes':
                    storage_cost_data.append({
                        'Storage Type': service.replace('_', ' ').title(),
                        'Monthly Cost': f"${details['cost']:.2f}",
                        'Details': details['details']
                    })
            
            df_storage_costs = pd.DataFrame(storage_cost_data)
            st.dataframe(df_storage_costs, use_container_width=True, hide_index=True)
        
        # Data protection
        st.markdown("**Data Protection & Management**")
        
        protection_data = [
            {'Feature': 'Backup Strategy', 'Configuration': storage_recs['backup_strategy']},
            {'Feature': 'Encryption', 'Configuration': storage_recs['encryption']},
            {'Feature': 'Performance', 'Configuration': storage_recs['performance']},
            {'Feature': 'Lifecycle Policy', 'Configuration': storage_recs['lifecycle_policy']}
        ]
        
        df_protection = pd.DataFrame(protection_data)
        st.dataframe(df_protection, use_container_width=True, hide_index=True)
        
        # Cost optimization notes
        st.markdown("**💡 Storage Cost Optimization**")
        for note in storage_costs.get('optimization_notes', []):
            st.markdown(note)
    
    # Database tab with costs
    with tech_tabs[3]:
        st.markdown("#### 🗄️ Database Configuration & Costs")
        
        db_recs = tech_recs['database']
        db_costs = service_costs['database']
        
        # Cost overview
        st.markdown(f"**Monthly Database Cost: ${db_costs['total']:.2f}**")
        
        col1, col2 = st.columns(2)
        
        with col1:
            st.markdown("**Database Engine & Instance**")
            
            db_config_data = [
                {'Setting': 'Database Engine', 'Value': db_recs['engine']},
                {'Setting': 'Instance Class', 'Value': db_recs['instance_class']},
                {'Setting': 'Multi-AZ', 'Value': 'Yes' if db_recs['multi_az'] else 'No'},
                {'Setting': 'Backup Retention', 'Value': db_recs['backup_retention']}
            ]
            
            df_db_config = pd.DataFrame(db_config_data)
            st.dataframe(df_db_config, use_container_width=True, hide_index=True)
        
        with col2:
            st.markdown("**Database Cost Breakdown**")
            
            db_cost_data = []
            for service, details in db_costs.items():
                if service != 'total' and service != 'optimization_notes':
                    db_cost_data.append({
                        'Database Component': service.replace('_', ' ').title(),
                        'Monthly Cost': f"${details['cost']:.2f}",
                        'Details': details['details']
                    })
            
            df_db_costs = pd.DataFrame(db_cost_data)
            st.dataframe(df_db_costs, use_container_width=True, hide_index=True)
        
        # Advanced database features
        st.markdown("**Advanced Database Features**")
        
        db_advanced_data = [
            {'Feature': 'Read Replicas', 'Configuration': db_recs['read_replicas']},
            {'Feature': 'Connection Pooling', 'Configuration': db_recs['connection_pooling']},
            {'Feature': 'Maintenance Window', 'Configuration': db_recs['maintenance_window']},
            {'Feature': 'Monitoring', 'Configuration': db_recs['monitoring']}
        ]
        
        df_db_advanced = pd.DataFrame(db_advanced_data)
        st.dataframe(df_db_advanced, use_container_width=True, hide_index=True)
        
        # Cost optimization notes
        st.markdown("**💡 Database Cost Optimization**")
        for note in db_costs.get('optimization_notes', []):
            st.markdown(note)
    
    # Security tab with costs
    with tech_tabs[4]:
        st.markdown("#### 🔒 Security Configuration & Costs")
        
        security_recs = tech_recs['security']
        security_costs = service_costs['security']
        
        # Cost overview
        st.markdown(f"**Monthly Security Cost: ${security_costs['total']:.2f}**")
        
        col1, col2 = st.columns(2)
        
        with col1:
            st.markdown("**Security Services Configuration**")
            
            security_data = []
            for key, value in security_recs.items():
                security_data.append({'Security Area': key.replace('_', ' ').title(), 'Configuration': value})
            
            df_security = pd.DataFrame(security_data)
            st.dataframe(df_security, use_container_width=True, hide_index=True)
        
        with col2:
            st.markdown("**Security Cost Breakdown**")
            
            security_cost_data = []
            for service, details in security_costs.items():
                if service != 'total' and service != 'optimization_notes':
                    security_cost_data.append({
                        'Security Service': service.replace('_', ' ').title(),
                        'Monthly Cost': f"${details['cost']:.2f}",
                        'Details': details['details']
                    })
            
            df_security_costs = pd.DataFrame(security_cost_data)
            st.dataframe(df_security_costs, use_container_width=True, hide_index=True)
        
        # Security best practices
        st.markdown("**Security Best Practices for this Environment:**")
        
        security_practices = [
            "🔐 Implement least privilege access principles",
            "🔍 Enable comprehensive audit logging",
            "🛡️ Use AWS Config for compliance monitoring",
            "🚨 Set up GuardDuty for threat detection",
            "📊 Regular security assessments and penetration testing"
        ]
        
        for practice in security_practices:
            st.markdown(practice)
        
        # Cost optimization notes
        st.markdown("**💡 Security Cost Optimization**")
        for note in security_costs.get('optimization_notes', []):
            st.markdown(note)
    
    # Monitoring tab with costs
    with tech_tabs[5]:
        st.markdown("#### 📊 Monitoring Configuration & Costs")
        
        monitoring_recs = tech_recs['monitoring']
        monitoring_costs = service_costs['monitoring']
        
        # Cost overview
        st.markdown(f"**Monthly Monitoring Cost: ${monitoring_costs['total']:.2f}**")
        
        col1, col2 = st.columns(2)
        
        with col1:
            st.markdown("**Core Monitoring Setup**")
            
            monitoring_core_data = [
                {'Component': 'CloudWatch', 'Configuration': monitoring_recs['cloudwatch']},
                {'Component': 'Alerting', 'Configuration': monitoring_recs['alerting']},
                {'Component': 'Dashboards', 'Configuration': monitoring_recs['dashboards']},
                {'Component': 'Log Retention', 'Configuration': monitoring_recs['log_retention']}
            ]
            
            df_monitoring_core = pd.DataFrame(monitoring_core_data)
            st.dataframe(df_monitoring_core, use_container_width=True, hide_index=True)
        
        with col2:
            st.markdown("**Monitoring Cost Breakdown**")
            
            monitoring_cost_data = []
            for service, details in monitoring_costs.items():
                if service != 'total' and service != 'optimization_notes':
                    monitoring_cost_data.append({
                        'Monitoring Service': service.replace('_', ' ').title(),
                        'Monthly Cost': f"${details['cost']:.2f}",
                        'Details': details['details']
                    })
            
            df_monitoring_costs = pd.DataFrame(monitoring_cost_data)
            st.dataframe(df_monitoring_costs, use_container_width=True, hide_index=True)
        
        # Advanced monitoring services
        st.markdown("**Advanced Monitoring Services**")
        
        monitoring_advanced_data = [
            {'Service': 'APM (X-Ray)', 'Configuration': monitoring_recs['apm']},
            {'Service': 'Synthetic Monitoring', 'Configuration': monitoring_recs['synthetic_monitoring']},
            {'Service': 'Cost Monitoring', 'Configuration': monitoring_recs['cost_monitoring']},
            {'Service': 'Health Checks', 'Configuration': monitoring_recs['health_checks']}
        ]
        
        df_monitoring_advanced = pd.DataFrame(monitoring_advanced_data)
        st.dataframe(df_monitoring_advanced, use_container_width=True, hide_index=True)
        
        # Cost optimization notes
        st.markdown("**💡 Monitoring Cost Optimization**")
        for note in monitoring_costs.get('optimization_notes', []):
            st.markdown(note)

# Bulk upload and reporting functions
def generate_bulk_template():
    """Downloadable CSV template for bulk upload."""
    import io

    sample_data = {
        "workload_name": ["App 1", "DB 1"],
        "cpu_cores": [4, 8],
        "ram_gb": [16, 32],
        "storage_gb": [200, 500],
        "workload_type": ["web_application", "database_server"],
        "operating_system": ["linux", "windows"],
        "peak_cpu_percent": [70, 85],
        "peak_ram_percent": [75, 90],
        "peak_iops": [3000, 6000],
        "business_criticality": ["medium", "high"],
        "region": ["us-east-1", "us-east-1"]
    }

    df = pd.DataFrame(sample_data)
    output = io.BytesIO()
    df.to_csv(output, index=False)
    output.seek(0)

    st.download_button(
        label="⬇️ Click to Download CSV Template",
        data=output,
        file_name="bulk_upload_template.csv",
        mime="text/csv"
    )

def render_bulk_upload_tab():
    """Render bulk upload tab."""
    st.markdown("### 📁 Bulk Workload Upload & Analysis")
    
    st.markdown("""
    Upload multiple workloads at once using CSV or Excel files. This feature allows you to:
    - Analyze dozens of workloads simultaneously
    - Get comprehensive cost and complexity analysis for all environments
    - Export consolidated reports
    - Compare workloads side-by-side
    """)
    
    # File upload section
    st.markdown("#### 📤 Upload Workloads File")
    
    col1, col2 = st.columns([2, 1])
    
    with col1:
        uploaded_file = st.file_uploader(
            "Choose a CSV or Excel file",
            type=['csv', 'xlsx', 'xls'],
            help="Upload a file containing multiple workloads to analyze"
        )
    
    with col2:
        # Create a button to trigger template download
        if st.button("📋 Download Template", key="bulk_template_download"):
            # Call the template generation function
            generate_bulk_template()
    
    # Show file format requirements
    with st.expander("📋 File Format Requirements", expanded=False):
        st.markdown("""
        **Required Columns** (case-insensitive):
        - `workload_name` or `name` - Name of the workload
        - `cpu_cores` or `cores` - Number of CPU cores
        - `ram_gb` or `memory_gb` - RAM in GB
        - `storage_gb` or `disk_gb` - Storage in GB
        
        **Optional Columns:**
        - `workload_type` - web_application, database_server, etc.
        - `operating_system` - linux, windows
        - `peak_cpu_percent` - Peak CPU utilization %
        - `peak_ram_percent` - Peak RAM utilization %
        - `peak_iops` - Peak IOPS
        - `business_criticality` - low, medium, high, critical
        - `region` - AWS region (default: us-east-1)
        
        **Example CSV:**
        ```
        workload_name,cpu_cores,ram_gb,storage_gb,workload_type,peak_cpu_percent
        Web App 1,4,16,200,web_application,75
        Database 1,8,32,500,database_server,85
        API Service,2,8,100,application_server,60
        ```
        """)
    
    # Process uploaded file
    if uploaded_file is not None:
        file_type = 'csv' if uploaded_file.name.endswith('.csv') else 'excel'
        
        if st.button("🚀 Analyze All Workloads", type="primary", key="bulk_analyze_button"):
            with st.spinner("🔄 Processing bulk workload analysis..."):
                bulk_analyzer = BulkWorkloadAnalyzer()
                results = bulk_analyzer.process_bulk_upload(uploaded_file, file_type)
                
                # Store results in session state
                st.session_state.bulk_results = results
                
                if 'error' in results:
                    st.error(f"❌ Error processing file: {results['error']}")
                else:
                    st.success(f"✅ Successfully analyzed {results['successful_analyses']} out of {results['total_workloads']} workloads!")
    
    # Display results
    if 'bulk_results' in st.session_state and st.session_state.bulk_results:
        render_bulk_results()
    
    # Add report generation section
    if 'bulk_results' in st.session_state and st.session_state.bulk_results:
        st.markdown("---")
        st.markdown("### 📋 Bulk Report Generation")
        col1, col2 = st.columns(2)
        with col1:
            if st.button("📊 Export to Excel", key="bulk_excel_export"):
                export_bulk_results_to_excel(st.session_state.bulk_results)
        with col2:
            if st.button("📄 Generate PDF Report", key="bulk_pdf_export"):
                export_bulk_results_to_pdf(st.session_state.bulk_results)

def render_bulk_results():
    """Render bulk analysis results."""
    results = st.session_state.bulk_results
    
    if 'error' in results:
        st.error(f"Error: {results['error']}")
        return
    
    st.markdown("---")
    st.markdown("### 📊 Bulk Analysis Results")    
    
    # Summary section
    summary = results.get('summary', {})
    if 'error' not in summary:
        col1, col2, col3, col4 = st.columns(4)
        
        with col1:
            st.metric("Total Workloads", summary.get('total_workloads_analyzed', 0))
        
        with col2:
            st.metric("Total Monthly Cost", f"${summary.get('total_monthly_cost', 0):,.0f}")
        
        with col3:
            st.metric("Average Complexity", f"{summary.get('average_complexity_score', 0):.1f}/100")
        
        with col4:
            st.metric("Most Common Instance", summary.get('most_common_instance_type', 'N/A'))
    
    # Add detailed tabs for bulk workloads
    if results['successful_analyses'] > 0:
        st.markdown("#### 🔍 Detailed Workload Analysis")
        selected_workload = st.selectbox(
            "Select Workload for Detailed View",
            [w['workload_name'] for w in results['workloads'] if w['status'] == 'success']
        )
        
        if selected_workload:
            workload_data = next(w for w in results['workloads'] 
                             if w['workload_name'] == selected_workload and w['status'] == 'success')
            
            # Create tabs for detailed analysis
            detailed_tabs = st.tabs(["Analysis", "Heat Map", "Recommendations"])
            
            with detailed_tabs[0]:
                render_workload_analysis(workload_data)
                
            with detailed_tabs[1]:
                render_workload_heatmaps(workload_data)
                
            with detailed_tabs[2]:
                render_workload_recommendations(workload_data)

def render_workload_analysis(workload_data):
    """Render analysis for a specific workload."""
    st.markdown(f"### 📊 Analysis for {workload_data['workload_name']}")
    
    try:
        prod_results = workload_data['analysis']['PROD']
        claude_analysis = prod_results.get('claude_analysis', {})
        tco_analysis = prod_results.get('tco_analysis', {})
        
        # Summary metrics
        col1, col2, col3, col4 = st.columns(4)
        
        with col1:
            complexity_score = claude_analysis.get('complexity_score', 50)
            complexity_level = claude_analysis.get('complexity_level', 'MEDIUM')
            
            st.markdown(f"""
            <div class="metric-card">
                <div style="font-size: 0.875rem; font-weight: 600; color: #6b7280; margin-bottom: 0.5rem;">🤖 Migration Complexity</div>
                <div style="font-size: 2rem; font-weight: 700; color: #1f2937; margin-bottom: 0.25rem;">{complexity_score:.0f}/100</div>
                <div style="font-size: 0.75rem; color: #9ca3af;">{complexity_level}</div>
            </div>
            """, unsafe_allow_html=True)
        
        with col2:
            monthly_cost = tco_analysis.get('monthly_cost', 0)
            
            st.markdown(f"""
            <div class="metric-card">
                <div style="font-size: 0.875rem; font-weight: 600; color: #6b7280; margin-bottom: 0.5rem;">☁️ AWS Monthly Cost</div>
                <div style="font-size: 2rem; font-weight: 700; color: #1f2937; margin-bottom: 0.25rem;">${monthly_cost:,.0f}</div>
                <div style="font-size: 0.75rem; color: #9ca3af;">Optimized Pricing</div>
            </div>
            """, unsafe_allow_html=True)
        
        with col3:
            timeline = claude_analysis.get('estimated_timeline', {})
            max_weeks = timeline.get('max_weeks', 8)
            
            st.markdown(f"""
            <div class="metric-card">
                <div style="font-size: 0.875rem; font-weight: 600; color: #6b7280; margin-bottom: 0.5rem;">⏱️ Migration Timeline</div>
                <div style="font-size: 2rem; font-weight: 700; color: #1f2937; margin-bottom: 0.25rem;">{max_weeks}</div>
                <div style="font-size: 0.75rem; color: #9ca3af;">Weeks (Estimated)</div>
            </div>
            """, unsafe_allow_html=True)
        
        with col4:
            st.markdown(f"""
            <div class="metric-card">
                <div style="font-size: 0.875rem; font-weight: 600; color: #6b7280; margin-bottom: 0.5rem;">🖥️ Instance Type</div>
                <div style="font-size: 1.5rem; font-weight: 700; color: #1f2937; margin-bottom: 0.25rem;">
                    {prod_results.get('cost_breakdown', {}).get('selected_instance', {}).get('type', 'N/A')}
                </div>
                <div style="font-size: 0.75rem; color: #9ca3af;">Recommended</div>
            </div>
            """, unsafe_allow_html=True)
        
        # Show key recommendations
        st.markdown("### 💡 Key Recommendations")
        recommendations = claude_analysis.get('recommendations', [])
        if recommendations:
            for i, rec in enumerate(recommendations[:5], 1):
                st.markdown(f"{i}. {rec}")
                
    except Exception as e:
        st.error(f"❌ Error displaying workload analysis: {str(e)}")
        logger.error(f"Error in render_workload_analysis: {e}")

def render_workload_heatmaps(workload_data):
    """Render heatmaps for a specific workload."""
    st.markdown(f"### 🌡️ Environment Heat Map for {workload_data['workload_name']}")
    
    # Generate heat map data
    heat_map_generator = EnvironmentHeatMapGenerator()
    heat_map_data = heat_map_generator.generate_heat_map_data(workload_data['analysis'])
    heat_map_fig = heat_map_generator.create_heat_map_visualization(heat_map_data)
    
    st.plotly_chart(heat_map_fig, use_container_width=True)

def render_workload_recommendations(workload_data):
    """Render recommendations for a specific workload."""
    st.markdown(f"### 🔧 Technical Recommendations for {workload_data['workload_name']}")
    
    try:
        analyzer = EnhancedEnvironmentAnalyzer()
        env = 'PROD'  # Focus on production environment
        env_results = workload_data['analysis'][env]
        tech_recs = analyzer.get_technical_recommendations(env, env_results)
        cost_calculator = AWSCostCalculator()
        requirements = env_results.get('requirements', {})
        service_costs = cost_calculator.calculate_service_costs(env, tech_recs, requirements)
        
        # Show key technical recommendations
        st.markdown("#### 💻 Compute Recommendations")
        compute_recs = tech_recs['compute']
        primary_instance = compute_recs['primary_instance']
        st.markdown(f"**Recommended Instance:** {primary_instance['type']} ({primary_instance['vcpus']} vCPUs, {primary_instance['memory_gb']} GB RAM)")
        st.markdown(f"**Rationale:** {primary_instance['rationale']}")
        
        st.markdown("#### 🌐 Network Recommendations")
        network_recs = tech_recs['network']
        st.markdown(f"**VPC Design:** {network_recs['vpc_design']}")
        st.markdown(f"**Load Balancer:** {network_recs['load_balancer']}")
        
        st.markdown("#### 💾 Storage Recommendations")
        storage_recs = tech_recs['storage']
        st.markdown(f"**Primary Storage:** {storage_recs['primary_storage']}")
        st.markdown(f"**Backup Strategy:** {storage_recs['backup_strategy']}")
        
        # Cost summary
        st.markdown("#### 💰 Cost Summary")
        total_cost = service_costs['summary']['total_monthly']
        st.markdown(f"**Total Monthly Cost:** ${total_cost:,.2f}")
        st.markdown(f"**Annual Cost:** ${total_cost * 12:,.2f}")
        
    except Exception as e:
        st.error(f"❌ Error displaying technical recommendations: {str(e)}")
        logger.error(f"Error in render_workload_recommendations: {e}")

def export_bulk_results_to_excel(results):
    """Export bulk results to Excel."""
    if not OPENPYXL_AVAILABLE:
        st.error("📊 openpyxl not available. Please install with: `pip install openpyxl`")
        return
    
    try:
        # Create Excel workbook
        wb = openpyxl.Workbook()
        ws_summary = wb.active
        ws_summary.title = "Summary"
        
        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
        data_font = Font(size=11)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Add summary data
        ws_summary['A1'] = "Bulk Workload Analysis Summary"
        ws_summary['A1'].font = Font(bold=True, size=16)
        ws_summary.merge_cells('A1:D1')
        
        summary = results.get('summary', {})
        if 'error' in summary:
            ws_summary['A3'] = "Error: " + summary['error']
            ws_summary['A3'].font = Font(color="FF0000")
        else:
            # Summary metrics
            summary_data = [
                ["Total Workloads", summary.get('total_workloads_analyzed', 0)],
                ["Total Monthly Cost", f"${summary.get('total_monthly_cost', 0):,.2f}"],
                ["Average Monthly Cost", f"${summary.get('average_monthly_cost', 0):,.2f}"],
                ["Average Complexity", f"{summary.get('average_complexity_score', 0):.1f}/100"],
                ["Most Common Instance", summary.get('most_common_instance_type', 'N/A')]
            ]
            
            # Write summary data
            for i, (metric, value) in enumerate(summary_data, 3):
                ws_summary[f'A{i}'] = metric
                ws_summary[f'B{i}'] = value
                
                if i == 3:  # Header row
                    ws_summary[f'A{i}'].font = header_font
                    ws_summary[f'A{i}'].fill = header_fill
                    ws_summary[f'B{i}'].font = header_font
                    ws_summary[f'B{i}'].fill = header_fill
                else:
                    ws_summary[f'A{i}'].font = data_font
                    ws_summary[f'B{i}'].font = data_font
                
                ws_summary[f'A{i}'].border = border
                ws_summary[f'B{i}'].border = border
        
        # Workloads sheet
        ws_workloads = wb.create_sheet("Workloads")
        
        # Headers
        headers = ["Workload", "Status", "Complexity", "Monthly Cost", "Instance Type", "Timeline (weeks)", "Migration Strategy"]
        for col_idx, header in enumerate(headers, 1):
            cell = ws_workloads.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center')
        
        # Workload data
        for row_idx, workload in enumerate(results.get('workloads', []), 2):
            if workload['status'] == 'success':
                prod_analysis = workload['analysis']['PROD']
                claude_analysis = prod_analysis.get('claude_analysis', {})
                tco_analysis = prod_analysis.get('tco_analysis', {})
                cost_breakdown = prod_analysis.get('cost_breakdown', {})
                selected_instance = cost_breakdown.get('selected_instance', {})
                
                ws_workloads.cell(row=row_idx, column=1, value=workload['workload_name'])
                ws_workloads.cell(row=row_idx, column=2, value="✅ Success")
                ws_workloads.cell(row=row_idx, column=3, value=f"{claude_analysis.get('complexity_score', 0):.0f}/100")
                ws_workloads.cell(row=row_idx, column=4, value=f"${tco_analysis.get('monthly_cost', 0):,.2f}")
                ws_workloads.cell(row=row_idx, column=5, value=selected_instance.get('type', 'N/A'))
                ws_workloads.cell(row=row_idx, column=6, value=claude_analysis.get('estimated_timeline', {}).get('max_weeks', 'N/A'))
                ws_workloads.cell(row=row_idx, column=7, value=claude_analysis.get('migration_strategy', {}).get('approach', 'N/A'))
            else:
                ws_workloads.cell(row=row_idx, column=1, value=workload['workload_name'])
                ws_workloads.cell(row=row_idx, column=2, value="❌ Failed")
                ws_workloads.cell(row=row_idx, column=3, value="N/A")
                ws_workloads.cell(row=row_idx, column=4, value="N/A")
                ws_workloads.cell(row=row_idx, column=5, value="N/A")
                ws_workloads.cell(row=row_idx, column=6, value="N/A")
                ws_workloads.cell(row=row_idx, column=7, value=workload.get('error', 'Analysis failed'))
            
            # Apply formatting
            for col_idx in range(1, len(headers) + 1):
                cell = ws_workloads.cell(row=row_idx, column=col_idx)
                cell.font = data_font
                cell.border = border
        
        # Auto-adjust column widths
        for sheet in wb:
            for column in sheet.columns:
                max_length = 0
                column_letter = openpyxl.utils.get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2) * 1.2
                sheet.column_dimensions[column_letter].width = adjusted_width
        
        # Save to BytesIO buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"bulk_workload_analysis_{timestamp}.xlsx"
        
        st.download_button(
            label="⬇️ Download Excel Report",
            data=buffer.getvalue(),
            file_name=filename,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="bulk_excel_report_download"
        )
        
        st.success("✅ Bulk Excel report generated successfully!")
        
    except Exception as e:
        st.error(f"Error generating Excel report: {str(e)}")
        logger.error(f"Error in bulk Excel generation: {e}")

def export_bulk_results_to_pdf(results):
    """Export enhanced bulk results to PDF with detailed analysis for each workload."""
    if not REPORTLAB_AVAILABLE:
        st.error("📄 ReportLab not available. Please install with: `pip install reportlab`")
        return
    
    try:
        # Create PDF content
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=1*inch, leftMargin=0.75*inch)
        
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Title'],
            fontSize=20,
            spaceAfter=20,
            alignment=TA_CENTER,
            textColor=colors.HexColor('#1f2937'),
            fontName='Helvetica-Bold'
        )
        
        subtitle_style = ParagraphStyle(
            'CustomSubtitle',
            parent=styles['Heading1'],
            fontSize=16,
            spaceAfter=15,
            textColor=colors.HexColor('#2563eb'),
            fontName='Helvetica-Bold'
        )
        
        story = []
        
        # Title
        story.append(Paragraph("Bulk Workload Analysis Report with vROPS Integration", title_style))
        story.append(Paragraph("Enterprise AWS Migration Analysis", styles['Normal']))
        story.append(Paragraph(f"Generated: {datetime.now().strftime('%B %d, %Y at %I:%M %p')}", styles['Normal']))
        story.append(Spacer(1, 0.5*inch))
        
        # Executive Summary
        story.append(Paragraph("Executive Summary", styles['Heading1']))
        
        summary = results.get('summary', {})
        if 'error' in summary:
            story.append(Paragraph(f"Error: {summary['error']}", styles['Normal']))
        else:
            summary_data = [
                ['Metric', 'Value'],
                ['Total Workloads Analyzed', str(summary.get('total_workloads_analyzed', 0))],
                ['Total Monthly Cost', f"${summary.get('total_monthly_cost', 0):,.2f}"],
                ['Total Annual Cost', f"${summary.get('total_annual_cost', 0):,.2f}"],
                ['Average Monthly Cost', f"${summary.get('average_monthly_cost', 0):,.2f}"],
                ['Average Complexity Score', f"{summary.get('average_complexity_score', 0):.1f}/100"],
                ['Most Common Instance Type', summary.get('most_common_instance_type', 'N/A')]
            ]
            
            summary_table = Table(summary_data, colWidths=[3*inch, 3*inch])
            summary_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#f3f4f6')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.HexColor('#1f2937')),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 11),
                ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#e5e7eb')),
                ('FONTSIZE', (0, 1), (-1, -1), 10)
            ]))
            
            story.append(summary_table)
            story.append(Spacer(1, 0.3*inch))
        
        # Footer
        story.append(Spacer(1, 0.3*inch))
        footer_text = f"Bulk Analysis Report generated by Enhanced AWS Migration Platform v7.0 with vROPS Integration on {datetime.now().strftime('%B %d, %Y')}"
        footer_style = ParagraphStyle('Footer', parent=styles['Normal'], fontSize=8, alignment=TA_CENTER, textColor=colors.HexColor('#6b7280'))
        story.append(Paragraph(footer_text, footer_style))
        
        # Build PDF
        doc.build(story)
        buffer.seek(0)
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"bulk_workload_analysis_vrops_{timestamp}.pdf"
        
        st.download_button(
            label="⬇️ Download PDF Report",
            data=buffer.getvalue(),
            file_name=filename,
            mime="application/pdf",
            key="bulk_pdf_report_download"
        )
        
        st.success("✅ Bulk PDF report generated successfully!")
        
    except Exception as e:
        st.error(f"Error generating PDF: {str(e)}")
        logger.error(f"Error in bulk PDF generation: {e}")

def generate_enhanced_excel_report():
    """Generate comprehensive Excel report with multiple sheets for enhanced analysis."""
    
    if 'enhanced_results' not in st.session_state or not st.session_state.enhanced_results:
        st.warning("No analysis results available for Excel generation.")
        return
    
    if not OPENPYXL_AVAILABLE:
        st.error("📊 openpyxl not available. Please install with: `pip install openpyxl`")
        return
    
    try:
        results = st.session_state.enhanced_results
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        
        # Create Excel workbook
        wb = openpyxl.Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
        title_font = Font(bold=True, size=14)
        data_font = Font(size=11)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        center_alignment = Alignment(horizontal='center')
        
        # Sheet 1: Executive Summary
        ws_summary = wb.create_sheet("Executive Summary")
        
        # Add title
        ws_summary['A1'] = "AWS Migration Analysis - Executive Summary"
        ws_summary['A1'].font = title_font
        ws_summary.merge_cells('A1:E1')
        
        # Add generation date
        ws_summary['A2'] = f"Generated: {datetime.now().strftime('%B %d, %Y at %I:%M %p')}"
        ws_summary['A2'].font = Font(size=11, italic=True)
        ws_summary.merge_cells('A2:E2')
        
        # vROPS enhancement note
        if results.get('vrops_enhanced'):
            ws_summary['A3'] = "Enhanced with vRealize Operations performance data"
            ws_summary['A3'].font = Font(size=11, italic=True, color="0F766E")
            ws_summary.merge_cells('A3:E3')
        
        ws_summary.append([])  # Empty row
        
        # Summary data from analysis
        prod_results = results['recommendations']['PROD']
        claude_analysis = prod_results.get('claude_analysis', {})
        tco_analysis = prod_results.get('tco_analysis', {})
        
        ws_summary.append(["Analysis Summary", ""])
        ws_summary.append(["Migration Complexity", 
                          f"{claude_analysis.get('complexity_level', 'MEDIUM')} ({claude_analysis.get('complexity_score', 50):.0f}/100)"])
        ws_summary.append(["Estimated Timeline", 
                          f"{claude_analysis.get('estimated_timeline', {}).get('max_weeks', 8)} weeks"])
        ws_summary.append(["Monthly Cost (PROD)", f"${tco_analysis.get('monthly_cost', 0):,.2f}"])
        ws_summary.append(["Annual Cost (PROD)", f"${tco_analysis.get('monthly_cost', 0) * 12:,.2f}"])
        ws_summary.append(["Best Pricing Option", 
                          tco_analysis.get('best_pricing_option', 'N/A').replace('_', ' ').title()])
        ws_summary.append([])  # Empty row
        
        # vROPS insights if available
        vrops_insights = claude_analysis.get('vrops_insights', {})
        if vrops_insights and results.get('vrops_enhanced'):
            ws_summary.append(["vROPS Performance Insights", ""])
            ws_summary.append(["Performance Impact", vrops_insights.get('performance_impact', 'N/A')])
            ws_summary.append(["Sizing Confidence", vrops_insights.get('sizing_confidence', 'N/A')])
            ws_summary.append(["Optimization Potential", vrops_insights.get('optimization_potential', 'N/A')])
            ws_summary.append(["Recommended Approach", vrops_insights.get('recommended_approach', 'N/A')])
        
        # Apply styles
        for row in ws_summary.iter_rows():
            for cell in row:
                cell.border = border
                if cell.row == 1:
                    cell.font = title_font
                elif cell.row in [5, 12 if results.get('vrops_enhanced') else 11]:  # Section headers
                    cell.font = Font(bold=True)
        
        # Auto-adjust column widths
        ws_summary.column_dimensions['A'].width = 25
        ws_summary.column_dimensions['B'].width = 40
        
        # Save to BytesIO buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        filename = f"Enhanced_AWS_Migration_Analysis_vROPS_{timestamp}.xlsx"
        
        st.download_button(
            label="⬇️ Download Excel Report",
            data=buffer.getvalue(),
            file_name=filename,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="excel_report_download"
        )
        
        st.success("✅ Enhanced Excel report with vROPS integration generated successfully!")
        
    except Exception as e:
        st.error(f"Error generating Excel report: {str(e)}")
        logger.error(f"Error in Excel generation: {e}")

def generate_enhanced_pdf_report():
    """Generate enhanced PDF report with vROPS integration."""
    
    if 'enhanced_results' not in st.session_state or not st.session_state.enhanced_results:
        st.warning("No analysis results available for PDF generation.")
        return
    
    if not REPORTLAB_AVAILABLE:
        st.warning("📄 ReportLab not available. Please install with: `pip install reportlab`")
        return
    
    try:
        results = st.session_state.enhanced_results
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        
        # Create PDF content
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=1*inch, leftMargin=0.75*inch)
        
        styles = getSampleStyleSheet()
        
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Title'],
            fontSize=20,
            spaceAfter=20,
            alignment=TA_CENTER,
            textColor=colors.HexColor('#1f2937'),
            fontName='Helvetica-Bold'
        )
        
        story = []
        
        # Title
        story.append(Paragraph("Enhanced AWS Migration Analysis with vROPS Integration", title_style))
        story.append(Paragraph("Enterprise Corporation", styles['Normal']))
        story.append(Paragraph(f"Generated: {datetime.now().strftime('%B %d, %Y at %I:%M %p')}", styles['Normal']))
        
        # vROPS enhancement note
        if results.get('vrops_enhanced'):
            story.append(Paragraph("Enhanced with vRealize Operations performance data", 
                                 ParagraphStyle('VROPSNote', parent=styles['Normal'], 
                                              fontSize=12, textColor=colors.HexColor('#0f766e'), 
                                              fontName='Helvetica-Bold')))
        
        story.append(Spacer(1, 0.5*inch))
        
        # Executive Summary
        prod_results = results['recommendations']['PROD']
        claude_analysis = prod_results.get('claude_analysis', {})
        tco_analysis = prod_results.get('tco_analysis', {})
        
        story.append(Paragraph("Executive Summary", styles['Heading1']))
        
        summary_data = [
            ['Metric', 'Value'],
            ['Workload Name', results['inputs']['workload_name']],
            ['Migration Complexity', f"{claude_analysis.get('complexity_level', 'MEDIUM')} ({claude_analysis.get('complexity_score', 50):.0f}/100)"],
            ['Estimated Timeline', f"{claude_analysis.get('estimated_timeline', {}).get('max_weeks', 8)} weeks"],
            ['Monthly Cost', f"${tco_analysis.get('monthly_cost', 0):,.2f}"],
            ['Best Pricing Option', tco_analysis.get('best_pricing_option', 'N/A').replace('_', ' ').title()]
        ]
        
        # Add vROPS insights if available
        vrops_insights = claude_analysis.get('vrops_insights', {})
        if vrops_insights and results.get('vrops_enhanced'):
            summary_data.extend([
                ['Performance Assessment', vrops_insights.get('performance_impact', 'N/A')],
                ['Sizing Confidence', vrops_insights.get('sizing_confidence', 'N/A')],
                ['Recommended Approach', vrops_insights.get('recommended_approach', 'N/A')]
            ])
        
        summary_table = Table(summary_data, colWidths=[3*inch, 3*inch])
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#f3f4f6')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.HexColor('#1f2937')),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 11),
            ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#e5e7eb')),
            ('FONTSIZE', (0, 1), (-1, -1), 10)
        ]))
        
        story.append(summary_table)
        story.append(PageBreak())
        
        # Footer
        story.append(Spacer(1, 0.3*inch))
        footer_text = f"Report generated by Enhanced AWS Migration Platform v7.0 with vROPS Integration on {datetime.now().strftime('%B %d, %Y')}"
        footer_style = ParagraphStyle('Footer', parent=styles['Normal'], fontSize=8, alignment=TA_CENTER, textColor=colors.HexColor('#6b7280'))
        story.append(Paragraph(footer_text, footer_style))
        
        # Build PDF
        doc.build(story)
        buffer.seek(0)
        
        filename = f"Enhanced_AWS_Migration_Report_vROPS_{timestamp}.pdf"
        
        st.download_button(
            label="⬇️ Download Enhanced PDF Report",
            data=buffer.getvalue(),
            file_name=filename,
            mime="application/pdf",
            key="pdf_report_download"
        )
        
        st.success("✅ Enhanced PDF report with vROPS integration generated successfully!")
        
    except Exception as e:
        st.error(f"Error generating PDF: {str(e)}")

def get_env_characteristics(env: str) -> str:
    """Get environment characteristics for PDF report."""
    characteristics = {
        'DEV': 'Development, experimentation, low availability requirements',
        'QA': 'Testing, validation, automated test integration',
        'UAT': 'User acceptance, business validation, stakeholder testing',
        'PREPROD': 'Production simulation, disaster recovery testing',
        'PROD': 'Business critical, high availability, 24/7 operations'
    }
    return characteristics.get(env, 'Standard environment characteristics')

def show_aws_connection_status():
    """Show AWS connection status in the sidebar."""
    try:
        # Check AWS connection status through the cost calculator
        cost_calculator = AWSCostCalculator()
        
        if cost_calculator.aws_connected:
            aws_status = "🟢 Connected"
            aws_help = "AWS Pricing API connected for real-time pricing"
        else:
            aws_status = "🟡 Using Fallback"
            aws_help = "Using fallback pricing data. Configure AWS credentials for real-time pricing"
        
        st.markdown(f"**AWS Pricing API:** {aws_status}")
        st.markdown(f"*{aws_help}*")
        
    except Exception as e:
        st.markdown("**AWS Pricing API:** 🔴 Error")
        st.markdown("*Error checking AWS connection*")

def main():
    """Enhanced main application with vROPS integration and nested tab structure."""
    
    # Initialize session state
    initialize_enhanced_session_state()
    
    # Check if calculator is properly initialized
    if st.session_state.enhanced_calculator is None:
        st.error("⚠️ Application initialization failed. Please refresh the page.")
        if st.button("🔄 Retry Initialization", key="retry_init_button"):
            st.rerun()
        st.stop()
    
    # Enhanced header
    st.markdown("""
    <div class="main-header">
        <h1>🏢 Enhanced AWS Migration Platform v7.0 with vROPS Integration</h1>
        <p>Comprehensive environment analysis with real Claude AI integration, vRealize Operations metrics import, and detailed technical recommendations</p>
        <p style="font-size: 0.9rem; opacity: 0.9;">🤖 Real Anthropic Claude API • 📊 vROPS Performance Data Import • ☁️ AWS-Native Analysis • 🔧 Technical-Complete</p>
    </div>
    """, unsafe_allow_html=True)
    
    # Enhanced sidebar with vROPS status
    with st.sidebar:
        st.markdown("### 🔑 Integration Status")
        
        # Claude API Key configuration
        analyzer = ClaudeAIMigrationAnalyzer()
        api_key = analyzer._get_claude_api_key()
        
        if api_key:
            claude_status = "🟢 Connected"
            claude_help = "Claude AI is connected and ready for analysis"
        else:
            claude_status = "🔴 Not Connected"
            claude_help = "Add ANTHROPIC_API_KEY to Streamlit secrets or environment variables"
        
        st.markdown(f"**Claude AI:** {claude_status}")
        st.markdown(f"*{claude_help}*")
        
         # ADD THIS: AWS Connection Status
        show_aws_connection_status()
        
        # vROPS Connection Status
        vrops_status = st.session_state.vrops_connection_status
        if vrops_status['status'] == 'connected':
            vrops_display = "🟢 Connected"
            vrops_help = f"Connected to vROPS: {vrops_status.get('version', 'Unknown')}"
        elif vrops_status['status'] == 'error':
            vrops_display = "🔴 Error"
            vrops_help = f"Connection error: {vrops_status.get('message', 'Unknown error')}"
        else:
            vrops_display = "🟡 Not Connected"
            vrops_help = "Connect to vROPS for real performance metrics"
        
        st.markdown(f"**vROPS:** {vrops_display}")
        st.markdown(f"*{vrops_help}*")
        
        if not api_key:
            with st.expander("🔧 How to configure Claude API", expanded=False):
                st.markdown("""
                **Option 1: Streamlit Secrets (Recommended)**
                1. Create `.streamlit/secrets.toml` file
                2. Add: `ANTHROPIC_API_KEY = "your-api-key-here"`
                
                **Option 2: Environment Variable**
                1. Set environment variable: `ANTHROPIC_API_KEY=your-api-key-here`
                
                **Get API Key:**
                1. Visit [Anthropic Console](https://console.anthropic.com/)
                2. Create account and get API key
                3. Add to your configuration
                """)
        
        st.markdown("---")
        
        st.markdown("### 🚀 Enhanced Features")
        
        st.markdown("""
        **🤖 Claude AI Analysis:**
        - Migration complexity scoring
        - Risk assessment & mitigation
        - Intelligent migration strategies
        - Timeline estimation
        
        **📊 vROPS Integration:**
        - Real VM performance metrics
        - CPU, Memory, Storage analysis
        - Performance-based sizing
        - Migration readiness assessment
        
        **☁️ AWS Integration:**
        - Real-time pricing data
        - Instance recommendations
        - Cost optimization insights
        - Multi-environment analysis
        
        **🔧 Technical Specifications:**
        - Compute, Network, Storage configs
        - Database recommendations
        - Security & monitoring setup
        - Auto-scaling strategies
        """)
        
        # Quick stats if results available
        if st.session_state.enhanced_results:
            st.markdown("---")
            st.markdown("### 📈 Quick Stats")
            
            prod_results = st.session_state.enhanced_results['recommendations'].get('PROD', {})
            claude_analysis = prod_results.get('claude_analysis', {})
            tco_analysis = prod_results.get('tco_analysis', {})
            
            complexity_score = claude_analysis.get('complexity_score', 0)
            monthly_cost = tco_analysis.get('monthly_cost', 0)
            
            st.metric("Complexity Score", f"{complexity_score:.0f}/100")
            st.metric("Monthly Cost", f"${monthly_cost:,.0f}")
            
            if st.session_state.enhanced_results.get('vrops_enhanced'):
                st.markdown("📊 **Enhanced with vROPS data**")
    
    # MAIN TABS - Updated structure with vROPS Integration
    main_tabs = st.tabs(["Single Workload", "vROPS Connection", "Bulk Analysis", "Reports"])
    
    # SINGLE WORKLOAD TAB with nested sub-tabs
    with main_tabs[0]:
        st.markdown("### 🖥️ Single Workload Analysis")
        
        # Create sub-tabs under Single Workload
        single_workload_subtabs = st.tabs(["Configuration", "Results", "Heat Map", "Technical Reports"])
        
        # Configuration sub-tab
        with single_workload_subtabs[0]:
            render_enhanced_configuration()
        
        # Results sub-tab
        with single_workload_subtabs[1]:
            render_enhanced_results()
        
        # Heat Map sub-tab
        with single_workload_subtabs[2]:
            render_enhanced_environment_heatmap_tab()
        
        # Technical Reports sub-tab
        with single_workload_subtabs[3]:
            render_technical_recommendations_tab()
    
    # vROPS CONNECTION TAB
    with main_tabs[1]:
        render_vrops_connection_tab()
    
    # BULK ANALYSIS TAB
    with main_tabs[2]:
        st.markdown("### 📁 Bulk Workload Analysis")
        render_bulk_upload_tab()
    
    # REPORTS TAB
    with main_tabs[3]:
        st.markdown("### 📋 Enhanced Reports")
        
        # Add note about report context
        st.info("💡 Reports will be generated based on your current analysis context (Single Workload or Bulk Analysis)")
        
        # Check which type of results we have
        has_single_results = st.session_state.enhanced_results is not None
        has_bulk_results = ('bulk_results' in st.session_state and 
                           st.session_state.bulk_results is not None and 
                           'error' not in st.session_state.bulk_results)
        
        if has_single_results or has_bulk_results:
            # Show report type selector
            if has_single_results and has_bulk_results:
                report_type = st.radio(
                    "Select Report Type:",
                    ["Single Workload Reports", "Bulk Analysis Reports"],
                    help="Choose which analysis results to include in your reports"
                )
            elif has_single_results:
                report_type = "Single Workload Reports"
                st.info("📊 Single workload analysis available for reporting")
            else:
                report_type = "Bulk Analysis Reports"
                st.info("📊 Bulk workload analysis available for reporting")
            
            # Single Workload Reports
            if report_type == "Single Workload Reports" and has_single_results:
                st.markdown("#### Single Workload Reports")
                
                # vROPS enhancement indicator
                if st.session_state.enhanced_results.get('vrops_enhanced'):
                    st.success("📊 Reports will include vRealize Operations performance insights")
                
                col1, col2, col3 = st.columns(3)
                
                with col1:
                    if st.button("📄 Generate PDF Report", type="primary", key="reports_pdf_generate"):
                        generate_enhanced_pdf_report()
                
                with col2:
                    if st.button("📊 Export to Excel", key="reports_tab_excel"):
                        generate_enhanced_excel_report()
                
                with col3:
                    if st.button("📈 Generate Heat Map CSV", key="reports_heatmap_csv"):
                        if 'heat_map_data' in st.session_state.enhanced_results:
                            csv_data = st.session_state.enhanced_results['heat_map_data'].to_csv(index=False)
                            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                            st.download_button(
                                "⬇️ Download Heat Map CSV",
                                csv_data,
                                f"environment_heatmap_{timestamp}.csv",
                                "text/csv",
                                key="heatmap_csv_download"
                            )
                
                # Report preview for single workload
                st.markdown("#### Report Preview")
                prod_results = st.session_state.enhanced_results['recommendations']['PROD']
                claude_analysis = prod_results.get('claude_analysis', {})
                
                st.markdown("**Executive Summary Preview:**")
                
                summary_preview = f"""
                **Workload:** {st.session_state.enhanced_results['inputs']['workload_name']}
                
                **Migration Complexity:** {claude_analysis.get('complexity_level', 'MEDIUM')} ({claude_analysis.get('complexity_score', 50):.0f}/100)
                
                **Recommended Strategy:** {claude_analysis.get('migration_strategy', {}).get('approach', 'Standard Migration')}
                
                **Estimated Timeline:** {claude_analysis.get('estimated_timeline', {}).get('max_weeks', 8)} weeks
                
                **Key Recommendations:**
                """
                
                st.markdown(summary_preview)
                
                recommendations = claude_analysis.get('recommendations', [])
                for i, rec in enumerate(recommendations[:3], 1):
                    st.markdown(f"{i}. {rec}")
                
                if len(recommendations) > 3:
                    st.markdown(f"... and {len(recommendations) - 3} more recommendations in the full report")
                
                # vROPS insights preview
                vrops_insights = claude_analysis.get('vrops_insights', {})
                if vrops_insights and st.session_state.enhanced_results.get('vrops_enhanced'):
                    st.markdown("**vROPS Performance Insights:**")
                    st.markdown(f"• **Performance Impact:** {vrops_insights.get('performance_impact', 'N/A')}")
                    st.markdown(f"• **Sizing Confidence:** {vrops_insights.get('sizing_confidence', 'N/A')}")
            
            # Bulk Analysis Reports
            elif report_type == "Bulk Analysis Reports" and has_bulk_results:
                st.markdown("#### Bulk Analysis Reports")
                
                col1, col2 = st.columns(2)
                with col1:
                    if st.button("📊 Export to Excel", key="bulk_excel_export_reports"):
                        export_bulk_results_to_excel(st.session_state.bulk_results)
                with col2:
                    if st.button("📄 Generate PDF Report", key="bulk_pdf_export_reports"):
                        export_bulk_results_to_pdf(st.session_state.bulk_results)
                
                # Show bulk summary
                bulk_results = st.session_state.bulk_results
                summary = bulk_results.get('summary', {})
                
                if 'error' not in summary:
                    st.markdown("#### Bulk Analysis Summary")
                    
                    col1, col2, col3, col4 = st.columns(4)
                    
                    with col1:
                        st.metric("Total Workloads", summary.get('total_workloads_analyzed', 0))
                    
                    with col2:
                        st.metric("Total Monthly Cost", f"${summary.get('total_monthly_cost', 0):,.0f}")
                    
                    with col3:
                        st.metric("Average Complexity", f"{summary.get('average_complexity_score', 0):.1f}/100")
                    
                    with col4:
                        st.metric("Most Common Instance", summary.get('most_common_instance_type', 'N/A'))
        else:
            st.info("💡 Run an analysis (Single Workload or Bulk Upload) to generate comprehensive reports.")
            st.markdown("""
            **Available Report Types:**
            - **Single Workload Analysis:** Configure workload parameters and run analysis for detailed reports
            - **Bulk Analysis:** Upload CSV/Excel files with multiple workloads for consolidated reporting
            - **vROPS Enhanced:** Connect to vRealize Operations for performance-based analysis and reporting
            
            **Report Formats:**
            - **PDF Reports:** Executive summaries, technical recommendations, migration strategies
            - **Excel Reports:** Detailed cost breakdowns, environment comparisons, technical specifications
            - **CSV Exports:** Heat map data, cost analysis, environment metrics
            """)
    
    # Enhanced footer
    st.markdown("---")
    st.markdown("""
    <div style="text-align: center; color: #6b7280; font-size: 0.875rem; padding: 2rem 0;">
        <strong>Enhanced AWS Migration Platform v7.0 with vROPS Integration</strong><br>
        Now powered by <strong>Real Anthropic Claude AI API</strong> and <strong>vRealize Operations Performance Metrics</strong><br>
        for intelligent migration analysis and comprehensive technical recommendations<br>
        <em>🤖 Real AI-Enhanced • 📊 vROPS Performance Data • ☁️ AWS-Native • 🔧 Technical-Complete • 📋 Excel/PDF Reports</em>
    </div>
    """, unsafe_allow_html=True)

if __name__ == "__main__":
    main()